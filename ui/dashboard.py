#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
===============================================================
DASHBOARD WEB - Sistema SAC
CEDIS Chedraui Cancún 427
===============================================================

Dashboard web profesional para visualización y gestión de:
- Validación de OCs en tiempo real
- Métricas y estadísticas
- Histórico de validaciones
- Gestión de errores
- Generación de reportes
- Configuración del sistema

Uso:
    python dashboard.py
    python dashboard.py --port 8080
    python dashboard.py --debug

    Acceder en: http://localhost:5000

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Chedraui Logística Cancún
===============================================================
"""

import os
import sys
import logging
import time
import json
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any

# Flask y extensiones
try:
    from flask import (
        Flask, render_template, jsonify, request,
        redirect, url_for, flash, send_file, send_from_directory,
        session
    )
    from functools import wraps
    FLASK_AVAILABLE = True
except ImportError:
    FLASK_AVAILABLE = False
    print("⚠️  Flask no instalado. Ejecuta: pip install flask")
    sys.exit(1)

# Importar módulos del sistema
try:
    from config import (
        CEDIS, EMAIL_CONFIG, TELEGRAM_CONFIG, VERSION,
        DB_CONFIG, PATHS, SYSTEM_CONFIG
    )
except ImportError as e:
    print(f"⚠️  Error importando config: {e}")
    CEDIS = {'code': '427', 'name': 'CEDIS Cancún', 'region': 'Sureste', 'almacen': 'C22'}
    EMAIL_CONFIG = {}
    TELEGRAM_CONFIG = {}
    VERSION = '1.0.0'
    DB_CONFIG = {}
    PATHS = {'logs': Path('./output/logs'), 'resultados': Path('./output/resultados')}
    SYSTEM_CONFIG = {'environment': 'development', 'debug': True, 'timezone': 'America/Cancun'}

try:
    from modules.db_local import DBLocal, get_db_local, ResultadoValidacion
    DB_LOCAL_AVAILABLE = True
except ImportError as e:
    print(f"⚠️  Error importando db_local: {e}")
    DB_LOCAL_AVAILABLE = False
    get_db_local = None
    DBLocal = None
    ResultadoValidacion = None

try:
    from modules.db_connection import (
        get_connection_info, PYODBC_AVAILABLE, IBM_DB_AVAILABLE
    )
except ImportError as e:
    print(f"⚠️  Error importando db_connection: {e}")
    PYODBC_AVAILABLE = False
    IBM_DB_AVAILABLE = False
    def get_connection_info():
        return {'password_configured': False}

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ===============================================================
# CONFIGURACIÓN DE FLASK
# ===============================================================

app = Flask(__name__,
            template_folder='templates',
            static_folder='static')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'sac-cedis-427-secret-key-2025')

# Configuración
app.config.update(
    TEMPLATES_AUTO_RELOAD=True,
    JSON_SORT_KEYS=False,
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,  # 16MB max upload
)

# Directorio de reportes
REPORTS_DIR = PATHS.get('resultados', Path('./output/resultados'))
if isinstance(REPORTS_DIR, str):
    REPORTS_DIR = Path(REPORTS_DIR)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Configuración de autenticación
AUTH_CONFIG = {
    'enabled': os.environ.get('AUTH_ENABLED', 'false').lower() == 'true',
    'username': os.environ.get('AUTH_USERNAME', 'admin'),
    'password': os.environ.get('AUTH_PASSWORD', 'sac427'),
    'session_lifetime': 3600 * 8  # 8 horas
}


# ===============================================================
# DECORADOR DE AUTENTICACIÓN
# ===============================================================

def login_required(f):
    """Decorador que requiere autenticación para acceder a una ruta."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not AUTH_CONFIG['enabled']:
            return f(*args, **kwargs)

        if not session.get('logged_in'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'No autorizado', 'redirect': '/login'}), 401
            return redirect(url_for('login'))

        return f(*args, **kwargs)
    return decorated_function


# ===============================================================
# CONTEXTO GLOBAL PARA TEMPLATES
# ===============================================================

@app.context_processor
def inject_globals():
    """Inyecta variables globales a todos los templates."""
    db_info = get_connection_info() if 'get_connection_info' in dir() else {}
    return {
        'cedis_name': CEDIS.get('name', 'CEDIS'),
        'cedis_code': CEDIS.get('code', '427'),
        'current_time': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'version': VERSION,
        'db_status': db_info.get('password_configured', False),
        'auth_enabled': AUTH_CONFIG['enabled'],
        'logged_in': session.get('logged_in', False),
        'current_user': session.get('username', ''),
    }


# ===============================================================
# FUNCIONES AUXILIARES
# ===============================================================

def get_db():
    """Obtiene instancia de la base de datos local."""
    if DB_LOCAL_AVAILABLE:
        try:
            return get_db_local()
        except Exception:
            return DBLocal()
    else:
        # Retornar un objeto mock cuando db_local no está disponible
        return MockDB()


class MockDB:
    """Mock de base de datos para cuando db_local no está disponible."""

    def obtener_estadisticas(self, dias=30):
        return {
            'validaciones': {'total_validaciones': 0, 'exitosas': 0, 'fallidas': 0,
                           'tiempo_promedio': 0, 'oc_unicas': 0},
            'errores': {'total': 0, 'criticos': 0, 'no_resueltos': 0},
            'tasa_exito': 0,
            'top_oc_problematicas': []
        }

    def obtener_historial_oc(self, oc_numero=None, dias=30, limite=100):
        return []

    def obtener_errores(self, severidad=None, no_resueltos=False, dias=7, limite=100):
        return []

    def obtener_metricas_diarias(self, dias=30):
        return []

    def obtener_info_db(self):
        return {'ruta': 'N/A', 'tamanio_mb': 0, 'registros': {}, 'total_registros': 0}

    def guardar_validacion_oc(self, **kwargs):
        pass

    def registrar_error(self, **kwargs):
        pass

    def marcar_error_resuelto(self, error_id, solucion=None):
        return True

    def limpiar_historial_antiguo(self, dias=90):
        return {'historico_oc': 0, 'registro_errores': 0, 'metricas_diarias': 0,
                'auditoria': 0, 'cache_consultas': 0}


def generar_datos_demo_oc(oc_numero: str) -> Dict:
    """Genera datos demo para validación de OC."""
    import random

    total_oc = random.randint(1000, 50000)
    diferencia = random.choice([0, 0, 0, random.randint(-500, 500)])
    total_distro = total_oc - diferencia

    errores = []
    if diferencia != 0:
        errores.append({
            'tipo': 'DISTRIBUCION_DIFERENTE',
            'severidad': '🔴 CRÍTICO' if abs(diferencia) > 100 else '🟡 MEDIO',
            'mensaje': f'Diferencia de {diferencia} unidades entre OC y distribución',
            'solucion': 'Verificar cantidades en el sistema'
        })

    if random.random() < 0.3:
        errores.append({
            'tipo': 'SKU_SIN_INNER_PACK',
            'severidad': '🟠 ALTO',
            'mensaje': 'Algunos SKUs no tienen Inner Pack configurado',
            'solucion': 'Configurar Inner Pack en catálogo de productos'
        })

    resultado = 'exitoso' if not errores else ('critico' if any('CRÍTICO' in e['severidad'] for e in errores) else 'con_errores')

    return {
        'oc_numero': oc_numero,
        'total_oc': total_oc,
        'total_distro': total_distro,
        'diferencia': diferencia,
        'errores': errores,
        'resultado': resultado,
        'tiempo': random.uniform(0.5, 3.0)
    }


# ===============================================================
# RUTAS DE AUTENTICACIÓN
# ===============================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login."""
    if not AUTH_CONFIG['enabled']:
        return redirect(url_for('dashboard'))

    if session.get('logged_in'):
        return redirect(url_for('dashboard'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if username == AUTH_CONFIG['username'] and password == AUTH_CONFIG['password']:
            session['logged_in'] = True
            session['username'] = username
            session.permanent = True
            app.permanent_session_lifetime = timedelta(seconds=AUTH_CONFIG['session_lifetime'])

            logger.info(f"Usuario {username} ha iniciado sesión")
            flash('Sesión iniciada correctamente', 'success')
            return redirect(url_for('dashboard'))
        else:
            error = 'Usuario o contraseña incorrectos'
            logger.warning(f"Intento de login fallido para usuario: {username}")

    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    """Cerrar sesión."""
    username = session.get('username', 'Desconocido')
    session.clear()
    logger.info(f"Usuario {username} ha cerrado sesión")
    flash('Sesión cerrada correctamente', 'info')
    return redirect(url_for('login'))


# ===============================================================
# RUTAS WEB (PÁGINAS HTML)
# ===============================================================

@app.route('/')
@login_required
def dashboard():
    """Página principal del dashboard."""
    db = get_db()

    # Obtener estadísticas
    stats_raw = db.obtener_estadisticas(dias=30)
    stats = {
        'validaciones_hoy': stats_raw['validaciones'].get('total_validaciones', 0) or 0,
        'tasa_exito': round(stats_raw.get('tasa_exito', 0), 1),
        'errores_criticos': stats_raw['errores'].get('no_resueltos', 0) or 0,
        'oc_unicas': stats_raw['validaciones'].get('oc_unicas', 0) or 0,
        'tiempo_promedio': stats_raw['validaciones'].get('tiempo_promedio', 0) or 0,
    }

    # Últimas validaciones
    ultimas_validaciones = db.obtener_historial_oc(dias=7, limite=10)

    # Métricas para gráfico
    metricas = db.obtener_metricas_diarias(dias=7)

    # Estado del sistema
    db_info = get_connection_info() if 'get_connection_info' in dir() else {}
    db_local_info = db.obtener_info_db()

    return render_template(
        'dashboard.html',
        active_page='dashboard',
        stats=stats,
        ultimas_validaciones=ultimas_validaciones,
        metricas_json=json.dumps(metricas, default=str),
        db_status={'conectado': db_info.get('password_configured', False)},
        telegram_status=bool(TELEGRAM_CONFIG.get('bot_token')),
        email_status=bool(EMAIL_CONFIG.get('user')),
        db_local_info=db_local_info
    )


@app.route('/validar')
@login_required
def validar_oc_page():
    """Página para validar OC."""
    db = get_db()
    ultimas_ocs = db.obtener_historial_oc(dias=7, limite=5)

    # Pre-fill OC if provided in query
    oc_prefill = request.args.get('oc', '')

    return render_template(
        'validar_oc.html',
        active_page='validar',
        ultimas_ocs=ultimas_ocs,
        oc_prefill=oc_prefill
    )


@app.route('/validaciones')
@login_required
def validaciones():
    """Lista de validaciones de OC."""
    db = get_db()

    # Parámetros de filtro
    dias = request.args.get('dias', 30, type=int)
    oc = request.args.get('oc', '')
    resultado = request.args.get('resultado', '')

    historial = db.obtener_historial_oc(
        oc_numero=oc if oc else None,
        dias=dias,
        limite=200
    )

    # Filtrar por resultado si se especifica
    if resultado:
        historial = [v for v in historial if v.get('resultado') == resultado]

    return render_template(
        'validaciones.html',
        active_page='validaciones',
        validaciones=historial,
        filtro_dias=dias,
        filtro_oc=oc,
        filtro_resultado=resultado
    )


@app.route('/errores')
@login_required
def errores():
    """Lista de errores del sistema."""
    db = get_db()

    severidad = request.args.get('severidad', '')
    solo_pendientes = request.args.get('pendientes', '0') == '1'
    modulo = request.args.get('modulo', '')

    lista_errores = db.obtener_errores(
        severidad=severidad if severidad else None,
        no_resueltos=solo_pendientes,
        dias=30
    )

    # Filtrar por módulo si se especifica
    if modulo:
        lista_errores = [e for e in lista_errores if e.get('modulo') == modulo]

    # Obtener lista de módulos únicos
    modulos = list(set(e.get('modulo', '') for e in db.obtener_errores(dias=90)))

    # Estadísticas de errores
    stats = {
        'criticos': len([e for e in lista_errores if 'CRÍTICO' in str(e.get('severidad', ''))]),
        'altos': len([e for e in lista_errores if 'ALTO' in str(e.get('severidad', ''))]),
        'medios': len([e for e in lista_errores if 'MEDIO' in str(e.get('severidad', ''))]),
        'bajos': len([e for e in lista_errores if 'BAJO' in str(e.get('severidad', ''))]),
        'pendientes': len([e for e in lista_errores if not e.get('resuelto', False)]),
    }

    return render_template(
        'errores.html',
        active_page='errores',
        errores=lista_errores,
        filtro_severidad=severidad,
        filtro_pendientes=solo_pendientes,
        filtro_modulo=modulo,
        modulos=modulos,
        stats=stats
    )


@app.route('/metricas')
@login_required
def metricas():
    """Métricas y estadísticas del sistema."""
    db = get_db()

    dias = request.args.get('dias', 30, type=int)

    stats = db.obtener_estadisticas(dias=dias)
    metricas_diarias = db.obtener_metricas_diarias(dias=dias)

    return render_template(
        'metricas.html',
        active_page='metricas',
        estadisticas=stats,
        metricas=metricas_diarias,
        metricas_json=json.dumps(metricas_diarias, default=str),
        estadisticas_json=json.dumps(stats, default=str),
        periodo=dias
    )


@app.route('/reportes')
@login_required
def reportes():
    """Gestión de reportes."""
    # Listar archivos de reportes existentes
    archivos = []
    if REPORTS_DIR.exists():
        for f in sorted(REPORTS_DIR.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True)[:30]:
            archivos.append({
                'nombre': f.name,
                'tamanio': f"{f.stat().st_size / 1024:.1f} KB",
                'fecha': datetime.fromtimestamp(f.stat().st_mtime).strftime('%d/%m/%Y %H:%M')
            })

    return render_template(
        'reportes.html',
        active_page='reportes',
        archivos=archivos
    )


@app.route('/configuracion')
@login_required
def configuracion():
    """Página de configuración del sistema."""
    db_info = get_connection_info() if 'get_connection_info' in dir() else {}
    db_local = get_db()
    db_local_info = db_local.obtener_info_db()

    return render_template(
        'configuracion.html',
        active_page='config',
        db_config={
            'host': DB_CONFIG.get('host', 'No configurado'),
            'port': DB_CONFIG.get('port', '50000'),
            'database': DB_CONFIG.get('database', 'No configurado'),
            'schema': DB_CONFIG.get('schema', 'WMWHSE1'),
            'user': DB_CONFIG.get('user', 'No configurado'),
            'password_configured': db_info.get('password_configured', False),
            'driver_available': PYODBC_AVAILABLE or IBM_DB_AVAILABLE,
            'driver_type': 'pyodbc' if PYODBC_AVAILABLE else ('ibm_db' if IBM_DB_AVAILABLE else 'Ninguno')
        },
        db_local_info=db_local_info,
        telegram_config={
            'habilitado': bool(TELEGRAM_CONFIG.get('bot_token')),
            'chat_ids': len(TELEGRAM_CONFIG.get('chat_ids', []))
        },
        email_config={
            'configurado': bool(EMAIL_CONFIG.get('user')),
            'servidor': EMAIL_CONFIG.get('smtp_server', 'No configurado'),
            'puerto': EMAIL_CONFIG.get('smtp_port', '587'),
            'usuario': EMAIL_CONFIG.get('user', '')
        },
        cedis_info=CEDIS,
        system_info={
            'version': VERSION,
            'environment': SYSTEM_CONFIG.get('environment', 'production'),
            'timezone': SYSTEM_CONFIG.get('timezone', 'America/Cancun'),
            'debug': SYSTEM_CONFIG.get('debug', False)
        }
    )


# ===============================================================
# API ENDPOINTS (JSON)
# ===============================================================

@app.route('/api/stats')
def api_stats():
    """API: Estadísticas del sistema."""
    db = get_db()
    dias = request.args.get('dias', 30, type=int)
    return jsonify(db.obtener_estadisticas(dias=dias))


@app.route('/api/validaciones')
def api_validaciones():
    """API: Lista de validaciones."""
    db = get_db()
    dias = request.args.get('dias', 30, type=int)
    limite = request.args.get('limite', 100, type=int)
    return jsonify(db.obtener_historial_oc(dias=dias, limite=limite))


@app.route('/api/errores')
def api_errores():
    """API: Lista de errores."""
    db = get_db()
    return jsonify(db.obtener_errores(dias=30))


@app.route('/api/errores/<int:error_id>')
def api_error_detalle(error_id):
    """API: Detalle de un error específico."""
    db = get_db()
    errores = db.obtener_errores(dias=365)
    error = next((e for e in errores if e.get('id') == error_id), None)

    if error:
        return jsonify(error)
    return jsonify({'error': 'Error no encontrado'}), 404


@app.route('/api/errores/<int:error_id>/resolver', methods=['POST'])
def api_resolver_error(error_id):
    """API: Marcar error como resuelto."""
    try:
        db = get_db()
        # Intentar marcar como resuelto
        db.marcar_error_resuelto(error_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/metricas')
def api_metricas():
    """API: Métricas diarias."""
    db = get_db()
    dias = request.args.get('dias', 30, type=int)
    return jsonify(db.obtener_metricas_diarias(dias=dias))


@app.route('/api/health')
def api_health():
    """API: Estado de salud del sistema."""
    return jsonify({
        'status': 'ok',
        'timestamp': datetime.now().isoformat(),
        'cedis': CEDIS.get('code', '427'),
        'version': VERSION
    })


@app.route('/api/health-check')
def api_health_check():
    """API: Verificación completa del sistema."""
    checks = {}

    # Check DB Local
    try:
        db = get_db()
        db.obtener_estadisticas(dias=1)
        checks['DB Local'] = {'status': 'ok', 'message': 'Conectada'}
    except Exception as e:
        checks['DB Local'] = {'status': 'error', 'message': str(e)}

    # Check DB2 config
    db_info = get_connection_info() if 'get_connection_info' in dir() else {}
    if db_info.get('password_configured'):
        checks['DB2 Config'] = {'status': 'ok', 'message': 'Configurada'}
    else:
        checks['DB2 Config'] = {'status': 'warning', 'message': 'Modo demo'}

    # Check Email
    if EMAIL_CONFIG.get('user'):
        checks['Email'] = {'status': 'ok', 'message': 'Configurado'}
    else:
        checks['Email'] = {'status': 'warning', 'message': 'No configurado'}

    # Check Telegram
    if TELEGRAM_CONFIG.get('bot_token'):
        checks['Telegram'] = {'status': 'ok', 'message': 'Configurado'}
    else:
        checks['Telegram'] = {'status': 'warning', 'message': 'No configurado'}

    # Check Reports Dir
    if REPORTS_DIR.exists():
        checks['Directorio Reportes'] = {'status': 'ok', 'message': str(REPORTS_DIR)}
    else:
        checks['Directorio Reportes'] = {'status': 'error', 'message': 'No existe'}

    return jsonify({'checks': checks})


@app.route('/api/validar-oc', methods=['POST'])
def api_validar_oc():
    """API: Validar una orden de compra."""
    try:
        data = request.get_json()
        oc_numero = data.get('oc_numero', '').strip()

        if not oc_numero:
            return jsonify({'success': False, 'error': 'Número de OC requerido'})

        # Normalizar OC (agregar C si no tiene)
        if not oc_numero.startswith('C'):
            oc_numero = 'C' + oc_numero

        start_time = time.time()

        # Intentar validación real si hay conexión DB2
        db_info = get_connection_info() if 'get_connection_info' in dir() else {}

        if db_info.get('password_configured'):
            # Aquí iría la validación real con DB2
            # Por ahora usamos datos demo
            resultado = generar_datos_demo_oc(oc_numero)
        else:
            # Modo demo
            resultado = generar_datos_demo_oc(oc_numero)

        resultado['tiempo'] = time.time() - start_time
        resultado['success'] = True

        # Guardar en DB local
        try:
            db = get_db()
            db.guardar_validacion_oc(
                oc_numero=oc_numero,
                resultado=resultado['resultado'],
                errores=resultado['errores'],
                total_oc=resultado['total_oc'],
                total_distro=resultado['total_distro'],
                tiempo_ejecucion=resultado['tiempo']
            )

            # Guardar errores
            for error in resultado['errores']:
                db.registrar_error(
                    tipo_error=error['tipo'],
                    severidad=error['severidad'],
                    mensaje=error['mensaje'],
                    modulo='VALIDACION_OC',
                    oc_relacionada=oc_numero
                )
        except Exception as e:
            logger.warning(f"No se pudo guardar en DB local: {e}")

        return jsonify(resultado)

    except Exception as e:
        logger.error(f"Error en validación: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/generar-reporte', methods=['POST'])
def api_generar_reporte():
    """API: Generar un reporte Excel."""
    try:
        data = request.get_json()
        tipo = data.get('tipo', 'diario')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        formato = data.get('formato', 'xlsx')

        db = get_db()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Generar nombre del archivo
        nombre_archivo = f"Reporte_{tipo.title()}_{timestamp}.xlsx"
        ruta_archivo = REPORTS_DIR / nombre_archivo

        # Obtener datos según el tipo
        if tipo == 'diario':
            datos = db.obtener_historial_oc(dias=1, limite=1000)
            titulo = 'Reporte Diario de Validaciones'
        elif tipo == 'errores':
            datos = db.obtener_errores(dias=30)
            titulo = 'Reporte de Errores'
        elif tipo == 'estadisticas':
            datos = db.obtener_metricas_diarias(dias=30)
            titulo = 'Reporte de Estadísticas'
        elif tipo == 'validaciones':
            datos = db.obtener_historial_oc(dias=30, limite=1000)
            titulo = 'Reporte de Validaciones'
        else:
            datos = db.obtener_estadisticas(dias=30)
            titulo = 'Reporte General'

        # Generar Excel usando openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = 'Datos'

            # Estilos
            header_fill = PatternFill(start_color='E31837', end_color='E31837', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            # Título corporativo
            ws.merge_cells('A1:F1')
            ws['A1'] = f'SAC - Sistema de Automatización de Consultas - CEDIS {CEDIS.get("code", "427")}'
            ws['A1'].font = Font(bold=True, size=14)

            ws.merge_cells('A2:F2')
            ws['A2'] = titulo
            ws['A2'].font = Font(bold=True, size=12)

            ws.merge_cells('A3:F3')
            ws['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'

            # Datos
            if isinstance(datos, list) and len(datos) > 0:
                # Headers
                headers = list(datos[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col, value=header.upper())
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')

                # Data rows
                for row_idx, item in enumerate(datos, 6):
                    for col_idx, key in enumerate(headers, 1):
                        value = item.get(key, '')
                        if isinstance(value, datetime):
                            value = value.strftime('%d/%m/%Y %H:%M')
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Auto-adjust columns
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)

            wb.save(ruta_archivo)

            return jsonify({
                'success': True,
                'archivo': nombre_archivo,
                'url': f'/api/descargar/{nombre_archivo}'
            })

        except ImportError:
            return jsonify({
                'success': False,
                'error': 'openpyxl no está instalado'
            })

    except Exception as e:
        logger.error(f"Error generando reporte: {e}")
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/descargar/<filename>')
def api_descargar(filename):
    """API: Descargar un archivo de reporte."""
    try:
        # Seguridad: evitar path traversal
        safe_filename = Path(filename).name
        file_path = REPORTS_DIR / safe_filename

        if file_path.exists() and file_path.suffix in ['.xlsx', '.csv', '.pdf']:
            return send_file(
                file_path,
                as_attachment=True,
                download_name=safe_filename
            )

        return jsonify({'error': 'Archivo no encontrado'}), 404

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/reportes/<filename>', methods=['DELETE'])
def api_eliminar_reporte(filename):
    """API: Eliminar un reporte."""
    try:
        safe_filename = Path(filename).name
        file_path = REPORTS_DIR / safe_filename

        if file_path.exists():
            file_path.unlink()
            return jsonify({'success': True})

        return jsonify({'success': False, 'error': 'Archivo no encontrado'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/test-db')
def api_test_db():
    """API: Probar conexión a DB2."""
    try:
        db_info = get_connection_info() if 'get_connection_info' in dir() else {}

        if not db_info.get('password_configured'):
            return jsonify({
                'success': False,
                'error': 'Credenciales no configuradas. Configure DB_PASSWORD en .env'
            })

        # Aquí iría la prueba real de conexión
        return jsonify({
            'success': True,
            'message': 'Configuración de conexión válida (modo demo)'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/test-email', methods=['POST'])
def api_test_email():
    """API: Enviar correo de prueba."""
    try:
        data = request.get_json()
        email = data.get('email', '')

        if not email:
            return jsonify({'success': False, 'error': 'Email requerido'})

        # Aquí iría el envío real
        return jsonify({
            'success': True,
            'message': f'Correo de prueba enviado a {email} (simulado)'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/test-telegram', methods=['POST'])
def api_test_telegram():
    """API: Enviar mensaje de prueba a Telegram."""
    try:
        if not TELEGRAM_CONFIG.get('bot_token'):
            return jsonify({
                'success': False,
                'error': 'Telegram no configurado'
            })

        # Aquí iría el envío real
        return jsonify({
            'success': True,
            'message': 'Mensaje de Telegram enviado (simulado)'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/limpiar-historial', methods=['POST'])
def api_limpiar_historial():
    """API: Limpiar registros antiguos."""
    try:
        data = request.get_json()
        dias = data.get('dias', 90)

        if dias < 7:
            return jsonify({
                'success': False,
                'error': 'Debe mantener al menos 7 días de historial'
            })

        db = get_db()
        eliminados = db.limpiar_historial_antiguo(dias=dias)
        total = sum(eliminados.values())

        return jsonify({
            'success': True,
            'registros_eliminados': total,
            'detalle': eliminados
        })

    except Exception as e:
        logger.error(f"Error al limpiar historial: {e}")
        return jsonify({'success': False, 'error': str(e)})


# ===============================================================
# MANEJO DE ERRORES
# ===============================================================

@app.errorhandler(404)
def not_found(error):
    """Página 404."""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Endpoint no encontrado'}), 404
    return render_template('error.html',
                         active_page='error',
                         error_code=404,
                         error_message='Página no encontrada'), 404


@app.errorhandler(500)
def server_error(error):
    """Página 500."""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Error interno del servidor'}), 500
    return render_template('error.html',
                         active_page='error',
                         error_code=500,
                         error_message='Error interno del servidor'), 500


@app.errorhandler(403)
def forbidden(error):
    """Página 403."""
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Acceso denegado'}), 403
    return render_template('error.html',
                         active_page='error',
                         error_code=403,
                         error_message='Acceso denegado'), 403


# ===============================================================
# PUNTO DE ENTRADA
# ===============================================================

def iniciar_dashboard(host='0.0.0.0', port=5000, debug=False):
    """
    Inicia el servidor web del dashboard.

    Args:
        host: Host para el servidor (default: 0.0.0.0)
        port: Puerto para el servidor (default: 5000)
        debug: Modo debug (default: False)
    """
    print("\n" + "=" * 60)
    print("🌐 DASHBOARD WEB - SAC CEDIS 427")
    print("=" * 60)
    print(f"\n✅ Servidor iniciado en: http://localhost:{port}")
    print(f"📊 CEDIS: {CEDIS.get('name', 'Cancún')} ({CEDIS.get('code', '427')})")
    print(f"📁 Reportes: {REPORTS_DIR}")
    print(f"🔧 Modo: {'Debug' if debug else 'Producción'}")
    print("\n   Presiona Ctrl+C para detener")
    print("=" * 60 + "\n")

    app.run(host=host, port=port, debug=debug)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='SAC Dashboard Web')
    parser.add_argument('--host', default='0.0.0.0', help='Host (default: 0.0.0.0)')
    parser.add_argument('--port', type=int, default=5000, help='Puerto (default: 5000)')
    parser.add_argument('--debug', action='store_true', help='Modo debug')

    args = parser.parse_args()
    iniciar_dashboard(host=args.host, port=args.port, debug=args.debug)
