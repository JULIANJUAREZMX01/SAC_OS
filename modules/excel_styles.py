"""
═══════════════════════════════════════════════════════════════
MÓDULO DE ESTILOS EXCEL - SISTEMA SAC
Estilos corporativos Chedraui para reportes Excel
═══════════════════════════════════════════════════════════════

Este módulo define todos los estilos corporativos Chedraui para
la generación de reportes Excel profesionales.

IMPORTANTE: Los colores se importan desde config.py (fuente única).

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
    numbers, Protection
)
from typing import Dict, Optional
import logging

# Importar colores desde la configuración centralizada
from config import ExcelColors

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# PALETA DE COLORES CORPORATIVOS CHEDRAUI
# ═══════════════════════════════════════════════════════════════

# Alias para compatibilidad - ChedrauiColors ahora viene de config
ChedrauiColors = ExcelColors


# ═══════════════════════════════════════════════════════════════
# CLASE PRINCIPAL DE ESTILOS
# ═══════════════════════════════════════════════════════════════

class ChedrauiStyles:
    """
    Sistema de estilos corporativos Chedraui para Excel

    Proporciona estilos predefinidos para:
    - Encabezados y subencabezados
    - Datos y celdas
    - Formatos numéricos (moneda, porcentaje, fecha)
    - Estados y alertas
    """

    # Referencia a colores
    COLORS = ChedrauiColors

    # Estilos de borde
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    MEDIUM_BORDER = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    THICK_BOTTOM_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )

    # Formatos numéricos
    NUMBER_FORMAT_CURRENCY = '"$"#,##0.00'
    NUMBER_FORMAT_PERCENT = '0.00%'
    NUMBER_FORMAT_INTEGER = '#,##0'
    NUMBER_FORMAT_DECIMAL = '#,##0.00'
    NUMBER_FORMAT_DATE = 'DD/MM/YYYY'
    NUMBER_FORMAT_DATETIME = 'DD/MM/YYYY HH:MM'
    NUMBER_FORMAT_TIME = 'HH:MM:SS'

    _styles_registered = False

    @classmethod
    def get_header_style(cls) -> NamedStyle:
        """
        Obtiene estilo para encabezados principales

        Returns:
            NamedStyle: Estilo de encabezado corporativo
        """
        style = NamedStyle(name="chedraui_header")
        style.font = Font(
            name='Arial',
            size=12,
            bold=True,
            color=ChedrauiColors.WHITE
        )
        style.fill = PatternFill(
            start_color=ChedrauiColors.HEADER_RED,
            end_color=ChedrauiColors.HEADER_RED,
            fill_type='solid'
        )
        style.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_subheader_style(cls) -> NamedStyle:
        """
        Obtiene estilo para subencabezados

        Returns:
            NamedStyle: Estilo de subencabezado
        """
        style = NamedStyle(name="chedraui_subheader")
        style.font = Font(
            name='Arial',
            size=11,
            bold=True,
            color=ChedrauiColors.BLACK
        )
        style.fill = PatternFill(
            start_color=ChedrauiColors.SUBHEADER,
            end_color=ChedrauiColors.SUBHEADER,
            fill_type='solid'
        )
        style.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_data_style(cls) -> NamedStyle:
        """
        Obtiene estilo para celdas de datos normales

        Returns:
            NamedStyle: Estilo de datos
        """
        style = NamedStyle(name="chedraui_data")
        style.font = Font(
            name='Arial',
            size=10,
            color=ChedrauiColors.BLACK
        )
        style.alignment = Alignment(
            horizontal='left',
            vertical='center'
        )
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_data_center_style(cls) -> NamedStyle:
        """
        Obtiene estilo para celdas de datos centrados

        Returns:
            NamedStyle: Estilo de datos centrados
        """
        style = NamedStyle(name="chedraui_data_center")
        style.font = Font(
            name='Arial',
            size=10,
            color=ChedrauiColors.BLACK
        )
        style.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_currency_style(cls) -> NamedStyle:
        """
        Obtiene estilo para moneda

        Returns:
            NamedStyle: Estilo para valores de moneda
        """
        style = NamedStyle(name="chedraui_currency")
        style.font = Font(
            name='Arial',
            size=10,
            color=ChedrauiColors.BLACK
        )
        style.alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        style.number_format = cls.NUMBER_FORMAT_CURRENCY
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_percentage_style(cls) -> NamedStyle:
        """
        Obtiene estilo para porcentajes

        Returns:
            NamedStyle: Estilo para porcentajes
        """
        style = NamedStyle(name="chedraui_percentage")
        style.font = Font(
            name='Arial',
            size=10,
            color=ChedrauiColors.BLACK
        )
        style.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        style.number_format = cls.NUMBER_FORMAT_PERCENT
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_date_style(cls) -> NamedStyle:
        """
        Obtiene estilo para fechas

        Returns:
            NamedStyle: Estilo para fechas
        """
        style = NamedStyle(name="chedraui_date")
        style.font = Font(
            name='Arial',
            size=10,
            color=ChedrauiColors.BLACK
        )
        style.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        style.number_format = cls.NUMBER_FORMAT_DATE
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_datetime_style(cls) -> NamedStyle:
        """
        Obtiene estilo para fecha y hora

        Returns:
            NamedStyle: Estilo para fecha/hora
        """
        style = NamedStyle(name="chedraui_datetime")
        style.font = Font(
            name='Arial',
            size=10,
            color=ChedrauiColors.BLACK
        )
        style.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        style.number_format = cls.NUMBER_FORMAT_DATETIME
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_integer_style(cls) -> NamedStyle:
        """
        Obtiene estilo para números enteros

        Returns:
            NamedStyle: Estilo para enteros
        """
        style = NamedStyle(name="chedraui_integer")
        style.font = Font(
            name='Arial',
            size=10,
            color=ChedrauiColors.BLACK
        )
        style.alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        style.number_format = cls.NUMBER_FORMAT_INTEGER
        style.border = cls.THIN_BORDER
        return style

    @classmethod
    def get_status_style(cls, status: str) -> Dict:
        """
        Obtiene estilo según el estado

        Args:
            status: Estado del elemento (ok, warning, error, critical, info)

        Returns:
            Dict: Diccionario con font, fill, alignment, border
        """
        status_lower = status.lower()

        # Mapeo de estados a colores
        status_colors = {
            'ok': ChedrauiColors.OK_GREEN,
            'success': ChedrauiColors.OK_GREEN,
            'completo': ChedrauiColors.OK_GREEN,
            'warning': ChedrauiColors.WARNING_YELLOW,
            'alerta': ChedrauiColors.WARNING_YELLOW,
            'incompleta': ChedrauiColors.WARNING_YELLOW,
            'error': ChedrauiColors.ERROR_RED,
            'critical': ChedrauiColors.CRITICO,
            'critico': ChedrauiColors.CRITICO,
            'excedente': ChedrauiColors.CRITICO,
            'info': ChedrauiColors.INFO_BLUE,
            'pendiente': ChedrauiColors.INFO_BLUE,
        }

        # Determinar color
        fill_color = ChedrauiColors.LIGHT_GRAY
        for key, color in status_colors.items():
            if key in status_lower:
                fill_color = color
                break

        # Determinar color de texto
        font_color = ChedrauiColors.BLACK
        if fill_color in [ChedrauiColors.HEADER_RED, ChedrauiColors.CRITICO,
                          ChedrauiColors.ERROR_RED, ChedrauiColors.CHEDRAUI_BLUE]:
            font_color = ChedrauiColors.WHITE

        return {
            'font': Font(name='Arial', size=10, bold=True, color=font_color),
            'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': cls.THIN_BORDER
        }

    @classmethod
    def get_title_style(cls) -> Dict:
        """
        Obtiene estilo para títulos principales

        Returns:
            Dict: Estilo para título
        """
        return {
            'font': Font(name='Arial', size=18, bold=True, color=ChedrauiColors.WHITE),
            'fill': PatternFill(
                start_color=ChedrauiColors.HEADER_RED,
                end_color=ChedrauiColors.HEADER_RED,
                fill_type='solid'
            ),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': cls.MEDIUM_BORDER
        }

    @classmethod
    def get_subtitle_style(cls) -> Dict:
        """
        Obtiene estilo para subtítulos

        Returns:
            Dict: Estilo para subtítulo
        """
        return {
            'font': Font(name='Arial', size=14, bold=True, color=ChedrauiColors.BLACK),
            'alignment': Alignment(horizontal='left', vertical='center'),
        }

    @classmethod
    def get_total_style(cls) -> Dict:
        """
        Obtiene estilo para filas de totales

        Returns:
            Dict: Estilo para totales
        """
        return {
            'font': Font(name='Arial', size=11, bold=True, color=ChedrauiColors.BLACK),
            'fill': PatternFill(
                start_color=ChedrauiColors.LIGHT_GRAY,
                end_color=ChedrauiColors.LIGHT_GRAY,
                fill_type='solid'
            ),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': cls.THICK_BOTTOM_BORDER
        }

    @classmethod
    def get_alternating_row_style(cls, row_num: int) -> PatternFill:
        """
        Obtiene color de fondo para filas alternas

        Args:
            row_num: Número de fila

        Returns:
            PatternFill: Relleno de fondo
        """
        if row_num % 2 == 0:
            return PatternFill(
                start_color=ChedrauiColors.ALT_ROW,
                end_color=ChedrauiColors.ALT_ROW,
                fill_type='solid'
            )
        return PatternFill(fill_type=None)

    @classmethod
    def get_kpi_positive_style(cls) -> Dict:
        """
        Obtiene estilo para KPI positivo

        Returns:
            Dict: Estilo para KPI positivo
        """
        return {
            'font': Font(name='Arial', size=24, bold=True, color=ChedrauiColors.OK_GREEN),
            'alignment': Alignment(horizontal='center', vertical='center'),
        }

    @classmethod
    def get_kpi_negative_style(cls) -> Dict:
        """
        Obtiene estilo para KPI negativo

        Returns:
            Dict: Estilo para KPI negativo
        """
        return {
            'font': Font(name='Arial', size=24, bold=True, color=ChedrauiColors.ERROR_RED),
            'alignment': Alignment(horizontal='center', vertical='center'),
        }

    @classmethod
    def get_kpi_neutral_style(cls) -> Dict:
        """
        Obtiene estilo para KPI neutro

        Returns:
            Dict: Estilo para KPI neutro
        """
        return {
            'font': Font(name='Arial', size=24, bold=True, color=ChedrauiColors.CHEDRAUI_BLUE),
            'alignment': Alignment(horizontal='center', vertical='center'),
        }

    @classmethod
    def register_styles(cls, workbook) -> None:
        """
        Registra todos los estilos en un workbook

        Args:
            workbook: Workbook de openpyxl
        """
        if cls._styles_registered:
            return

        styles_to_register = [
            ('chedraui_header', cls.get_header_style),
            ('chedraui_subheader', cls.get_subheader_style),
            ('chedraui_data', cls.get_data_style),
            ('chedraui_data_center', cls.get_data_center_style),
            ('chedraui_currency', cls.get_currency_style),
            ('chedraui_percentage', cls.get_percentage_style),
            ('chedraui_date', cls.get_date_style),
            ('chedraui_datetime', cls.get_datetime_style),
            ('chedraui_integer', cls.get_integer_style),
        ]

        for style_name, style_getter in styles_to_register:
            try:
                if style_name not in workbook.named_styles:
                    style = style_getter()
                    style.name = style_name
                    workbook.add_named_style(style)
            except ValueError:
                # Style already exists
                pass

        cls._styles_registered = True
        logger.debug("✅ Estilos Chedraui registrados en workbook")

    @classmethod
    def apply_header_to_row(cls, worksheet, row_num: int, start_col: int = 1, end_col: int = None) -> None:
        """
        Aplica estilo de encabezado a una fila

        Args:
            worksheet: Hoja de trabajo
            row_num: Número de fila
            start_col: Columna inicial
            end_col: Columna final (None = última columna con datos)
        """
        if end_col is None:
            end_col = worksheet.max_column

        header_style = cls.get_header_style()

        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = header_style.font
            cell.fill = header_style.fill
            cell.alignment = header_style.alignment
            cell.border = header_style.border

    @classmethod
    def apply_data_to_range(cls, worksheet, start_row: int, end_row: int,
                           start_col: int = 1, end_col: int = None,
                           alternate_colors: bool = True) -> None:
        """
        Aplica estilo de datos a un rango de celdas

        Args:
            worksheet: Hoja de trabajo
            start_row: Fila inicial
            end_row: Fila final
            start_col: Columna inicial
            end_col: Columna final
            alternate_colors: Aplicar colores alternos
        """
        if end_col is None:
            end_col = worksheet.max_column

        data_style = cls.get_data_style()

        for row in range(start_row, end_row + 1):
            fill = cls.get_alternating_row_style(row) if alternate_colors else None

            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_style.font
                cell.alignment = data_style.alignment
                cell.border = data_style.border
                if fill and fill.fill_type:
                    cell.fill = fill


# ═══════════════════════════════════════════════════════════════
# FUNCIONES DE UTILIDAD
# ═══════════════════════════════════════════════════════════════

def apply_style_dict(cell, style_dict: Dict) -> None:
    """
    Aplica un diccionario de estilos a una celda

    Args:
        cell: Celda de openpyxl
        style_dict: Diccionario con font, fill, alignment, border
    """
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']
    if 'number_format' in style_dict:
        cell.number_format = style_dict['number_format']


def get_severity_color(severity: str) -> str:
    """
    Obtiene color según severidad

    Args:
        severity: Nivel de severidad

    Returns:
        str: Código de color hexadecimal
    """
    severity_lower = severity.lower()

    if 'crítico' in severity_lower or 'critico' in severity_lower:
        return ChedrauiColors.CRITICO
    elif 'alto' in severity_lower:
        return ChedrauiColors.ALTO
    elif 'medio' in severity_lower:
        return ChedrauiColors.MEDIO
    elif 'bajo' in severity_lower:
        return ChedrauiColors.BAJO
    else:
        return ChedrauiColors.INFO_BLUE


# ═══════════════════════════════════════════════════════════════
# EXPORTACIÓN
# ═══════════════════════════════════════════════════════════════

__all__ = [
    'ChedrauiColors',
    'ChedrauiStyles',
    'apply_style_dict',
    'get_severity_color',
]


if __name__ == "__main__":
    print("""
    ═══════════════════════════════════════════════════════════════
    SISTEMA DE ESTILOS EXCEL - CHEDRAUI
    ═══════════════════════════════════════════════════════════════

    Colores corporativos disponibles:
    """)

    print(f"  HEADER_RED:     #{ChedrauiColors.HEADER_RED}")
    print(f"  CHEDRAUI_BLUE:  #{ChedrauiColors.CHEDRAUI_BLUE}")
    print(f"  OK_GREEN:       #{ChedrauiColors.OK_GREEN}")
    print(f"  WARNING_YELLOW: #{ChedrauiColors.WARNING_YELLOW}")
    print(f"  ERROR_RED:      #{ChedrauiColors.ERROR_RED}")
    print(f"  INFO_BLUE:      #{ChedrauiColors.INFO_BLUE}")

    print("\n  Estilos disponibles:")
    print("  - get_header_style()")
    print("  - get_subheader_style()")
    print("  - get_data_style()")
    print("  - get_currency_style()")
    print("  - get_percentage_style()")
    print("  - get_date_style()")
    print("  - get_status_style(status)")
    print("  - get_kpi_positive_style()")
    print("  - get_kpi_negative_style()")

    print("""
    ═══════════════════════════════════════════════════════════════
    """)
