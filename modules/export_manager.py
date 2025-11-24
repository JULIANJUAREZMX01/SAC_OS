"""
═══════════════════════════════════════════════════════════════
GESTOR DE EXPORTACIÓN DE REPORTES
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Módulo para exportar reportes a múltiples formatos.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Union, Any
import zipfile
import json
import csv
import logging

from .excel_styles import ChedrauiStyles, ChedrauiColors
from .excel_templates.base_template import ExcelTemplate

logger = logging.getLogger(__name__)


class ExportConfig:
    """Configuración para exportación"""

    def __init__(
        self,
        output_dir: Path = None,
        timestamp_format: str = '%Y%m%d_%H%M%S',
        include_timestamp: bool = True,
        compression: bool = False,
        encoding: str = 'utf-8'
    ):
        """
        Inicializa configuración de exportación

        Args:
            output_dir: Directorio de salida
            timestamp_format: Formato de timestamp
            include_timestamp: Incluir timestamp en nombre de archivo
            compression: Comprimir archivos generados
            encoding: Codificación para archivos de texto
        """
        self.output_dir = output_dir or Path('output/resultados')
        self.timestamp_format = timestamp_format
        self.include_timestamp = include_timestamp
        self.compression = compression
        self.encoding = encoding

        # Crear directorio si no existe
        self.output_dir.mkdir(parents=True, exist_ok=True)


class ExportHistory:
    """Registro de exportaciones realizadas"""

    def __init__(self):
        self.exports: List[Dict[str, Any]] = []

    def add_export(
        self,
        filename: str,
        format_type: str,
        size_bytes: int,
        rows: int = 0,
        success: bool = True,
        error: str = None
    ) -> None:
        """Registra una exportación"""
        self.exports.append({
            'timestamp': datetime.now().isoformat(),
            'filename': filename,
            'format': format_type,
            'size_bytes': size_bytes,
            'rows': rows,
            'success': success,
            'error': error
        })

    def get_history(self) -> List[Dict[str, Any]]:
        """Obtiene historial de exportaciones"""
        return self.exports.copy()

    def get_summary(self) -> Dict[str, Any]:
        """Obtiene resumen de exportaciones"""
        total = len(self.exports)
        successful = sum(1 for e in self.exports if e['success'])
        failed = total - successful
        total_size = sum(e['size_bytes'] for e in self.exports if e['success'])

        return {
            'total_exports': total,
            'successful': successful,
            'failed': failed,
            'total_size_mb': round(total_size / (1024 * 1024), 2)
        }


class ExportManager:
    """
    Gestor de exportación de reportes

    Formatos soportados:
    - Excel (.xlsx)
    - CSV (.csv)
    - JSON (.json)
    - ZIP (compresión múltiple)
    """

    def __init__(self, config: ExportConfig = None):
        """
        Inicializa el gestor de exportación

        Args:
            config: Configuración de exportación
        """
        self.config = config or ExportConfig()
        self.history = ExportHistory()
        self.template = ExcelTemplate()

    def _generate_filename(
        self,
        base_name: str,
        extension: str
    ) -> Path:
        """Genera nombre de archivo con timestamp"""
        if self.config.include_timestamp:
            timestamp = datetime.now().strftime(self.config.timestamp_format)
            filename = f"{base_name}_{timestamp}.{extension}"
        else:
            filename = f"{base_name}.{extension}"

        return self.config.output_dir / filename

    def export_to_excel(
        self,
        df: pd.DataFrame,
        filename: str = None,
        sheet_name: str = "Datos",
        template_type: str = None,
        title: str = None
    ) -> str:
        """
        Exporta DataFrame a Excel con formato corporativo

        Args:
            df: DataFrame a exportar
            filename: Nombre del archivo (sin extensión)
            sheet_name: Nombre de la hoja
            template_type: Tipo de template a usar
            title: Título del reporte

        Returns:
            str: Ruta del archivo generado
        """
        try:
            filepath = self._generate_filename(filename or "Reporte", "xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Aplicar template si se especifica
            start_row = 1
            if title:
                template = ExcelTemplate()
                start_row = template.apply_corporate_header(
                    ws, title=title, subtitle="", merge_cols=len(df.columns)
                )

            # Escribir datos con formato
            template = ExcelTemplate()
            template.create_table(ws, df, start_row)
            template.auto_adjust_columns(ws)
            template.apply_chedraui_styles(ws)

            # Guardar
            wb.save(str(filepath))

            # Registrar exportación
            file_size = filepath.stat().st_size
            self.history.add_export(
                str(filepath), 'xlsx', file_size, len(df), True
            )

            logger.info(f"✅ Excel exportado: {filepath}")
            return str(filepath)

        except Exception as e:
            self.history.add_export(
                filename or "error", 'xlsx', 0, 0, False, str(e)
            )
            logger.error(f"❌ Error exportando Excel: {e}")
            raise

    def export_to_csv(
        self,
        df: pd.DataFrame,
        filename: str = None,
        separator: str = ',',
        include_index: bool = False
    ) -> str:
        """
        Exporta DataFrame a CSV

        Args:
            df: DataFrame a exportar
            filename: Nombre del archivo
            separator: Separador de campos
            include_index: Incluir índice

        Returns:
            str: Ruta del archivo generado
        """
        try:
            filepath = self._generate_filename(filename or "Reporte", "csv")

            df.to_csv(
                filepath,
                sep=separator,
                index=include_index,
                encoding=self.config.encoding
            )

            file_size = filepath.stat().st_size
            self.history.add_export(
                str(filepath), 'csv', file_size, len(df), True
            )

            logger.info(f"✅ CSV exportado: {filepath}")
            return str(filepath)

        except Exception as e:
            self.history.add_export(
                filename or "error", 'csv', 0, 0, False, str(e)
            )
            logger.error(f"❌ Error exportando CSV: {e}")
            raise

    def export_to_json(
        self,
        df: pd.DataFrame,
        filename: str = None,
        orient: str = 'records',
        indent: int = 2
    ) -> str:
        """
        Exporta DataFrame a JSON

        Args:
            df: DataFrame a exportar
            filename: Nombre del archivo
            orient: Orientación del JSON ('records', 'columns', 'index')
            indent: Indentación

        Returns:
            str: Ruta del archivo generado
        """
        try:
            filepath = self._generate_filename(filename or "Reporte", "json")

            df.to_json(
                filepath,
                orient=orient,
                indent=indent,
                force_ascii=False
            )

            file_size = filepath.stat().st_size
            self.history.add_export(
                str(filepath), 'json', file_size, len(df), True
            )

            logger.info(f"✅ JSON exportado: {filepath}")
            return str(filepath)

        except Exception as e:
            self.history.add_export(
                filename or "error", 'json', 0, 0, False, str(e)
            )
            logger.error(f"❌ Error exportando JSON: {e}")
            raise

    def export_multiple_sheets(
        self,
        dfs_dict: Dict[str, pd.DataFrame],
        filename: str = None,
        title: str = None
    ) -> str:
        """
        Exporta múltiples DataFrames a un Excel con varias hojas

        Args:
            dfs_dict: Diccionario {nombre_hoja: DataFrame}
            filename: Nombre del archivo
            title: Título para primera hoja

        Returns:
            str: Ruta del archivo generado
        """
        try:
            filepath = self._generate_filename(filename or "Reporte_Multiple", "xlsx")

            wb = Workbook()
            template = ExcelTemplate()
            first_sheet = True

            for sheet_name, df in dfs_dict.items():
                if first_sheet:
                    ws = wb.active
                    ws.title = sheet_name[:31]  # Excel limita a 31 caracteres
                    first_sheet = False
                else:
                    ws = wb.create_sheet(sheet_name[:31])

                if df.empty:
                    ws['A1'] = f"Sin datos para {sheet_name}"
                    continue

                # Aplicar formato
                start_row = 1
                if title and sheet_name == list(dfs_dict.keys())[0]:
                    start_row = template.apply_corporate_header(
                        ws, title=title, merge_cols=len(df.columns)
                    )

                template.create_table(ws, df, start_row)
                template.auto_adjust_columns(ws)

            wb.save(str(filepath))

            total_rows = sum(len(df) for df in dfs_dict.values())
            file_size = filepath.stat().st_size
            self.history.add_export(
                str(filepath), 'xlsx', file_size, total_rows, True
            )

            logger.info(f"✅ Excel multi-hoja exportado: {filepath} ({len(dfs_dict)} hojas)")
            return str(filepath)

        except Exception as e:
            self.history.add_export(
                filename or "error", 'xlsx', 0, 0, False, str(e)
            )
            logger.error(f"❌ Error exportando Excel multi-hoja: {e}")
            raise

    def compress_reports(
        self,
        files: List[str],
        zip_name: str = None
    ) -> str:
        """
        Comprime múltiples archivos en un ZIP

        Args:
            files: Lista de rutas de archivos
            zip_name: Nombre del archivo ZIP

        Returns:
            str: Ruta del archivo ZIP
        """
        try:
            zip_path = self._generate_filename(zip_name or "Reportes", "zip")

            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in files:
                    path = Path(file_path)
                    if path.exists():
                        zipf.write(path, path.name)
                        logger.debug(f"📦 Agregado a ZIP: {path.name}")

            file_size = zip_path.stat().st_size
            self.history.add_export(
                str(zip_path), 'zip', file_size, len(files), True
            )

            logger.info(f"✅ ZIP creado: {zip_path} ({len(files)} archivos)")
            return str(zip_path)

        except Exception as e:
            self.history.add_export(
                zip_name or "error", 'zip', 0, 0, False, str(e)
            )
            logger.error(f"❌ Error creando ZIP: {e}")
            raise

    def export_to_multiple_formats(
        self,
        df: pd.DataFrame,
        base_name: str,
        formats: List[str] = None
    ) -> Dict[str, str]:
        """
        Exporta DataFrame a múltiples formatos

        Args:
            df: DataFrame a exportar
            base_name: Nombre base del archivo
            formats: Lista de formatos ('xlsx', 'csv', 'json')

        Returns:
            Dict: {formato: ruta_archivo}
        """
        formats = formats or ['xlsx', 'csv']
        results = {}

        for fmt in formats:
            try:
                if fmt == 'xlsx':
                    results['xlsx'] = self.export_to_excel(df, base_name)
                elif fmt == 'csv':
                    results['csv'] = self.export_to_csv(df, base_name)
                elif fmt == 'json':
                    results['json'] = self.export_to_json(df, base_name)
            except Exception as e:
                results[fmt] = f"Error: {e}"
                logger.error(f"❌ Error exportando a {fmt}: {e}")

        return results

    def get_export_history(self) -> List[Dict[str, Any]]:
        """
        Obtiene historial de exportaciones

        Returns:
            List: Historial de exportaciones
        """
        return self.history.get_history()

    def get_export_summary(self) -> Dict[str, Any]:
        """
        Obtiene resumen de exportaciones

        Returns:
            Dict: Resumen estadístico
        """
        return self.history.get_summary()

    def cleanup_old_exports(
        self,
        days: int = 30,
        pattern: str = "*.xlsx"
    ) -> int:
        """
        Limpia exportaciones antiguas

        Args:
            days: Días de antigüedad máxima
            pattern: Patrón de archivos a limpiar

        Returns:
            int: Número de archivos eliminados
        """
        from datetime import timedelta

        cutoff = datetime.now() - timedelta(days=days)
        deleted = 0

        for file_path in self.config.output_dir.glob(pattern):
            if file_path.is_file():
                mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                if mtime < cutoff:
                    try:
                        file_path.unlink()
                        deleted += 1
                        logger.debug(f"🗑️ Eliminado: {file_path.name}")
                    except Exception as e:
                        logger.warning(f"⚠️ No se pudo eliminar {file_path}: {e}")

        logger.info(f"🧹 Limpieza completada: {deleted} archivos eliminados")
        return deleted


# ═══════════════════════════════════════════════════════════════
# FUNCIONES DE UTILIDAD
# ═══════════════════════════════════════════════════════════════

def quick_export(
    df: pd.DataFrame,
    filename: str,
    format_type: str = 'xlsx'
) -> str:
    """
    Función rápida para exportar DataFrames

    Args:
        df: DataFrame a exportar
        filename: Nombre del archivo
        format_type: Formato de exportación

    Returns:
        str: Ruta del archivo generado
    """
    manager = ExportManager()

    if format_type == 'xlsx':
        return manager.export_to_excel(df, filename)
    elif format_type == 'csv':
        return manager.export_to_csv(df, filename)
    elif format_type == 'json':
        return manager.export_to_json(df, filename)
    else:
        raise ValueError(f"Formato no soportado: {format_type}")


def export_with_compression(
    df: pd.DataFrame,
    filename: str,
    formats: List[str] = None
) -> str:
    """
    Exporta a múltiples formatos y comprime

    Args:
        df: DataFrame a exportar
        filename: Nombre base
        formats: Formatos a exportar

    Returns:
        str: Ruta del archivo ZIP
    """
    manager = ExportManager()
    formats = formats or ['xlsx', 'csv']

    results = manager.export_to_multiple_formats(df, filename, formats)
    files = [v for v in results.values() if not v.startswith('Error')]

    return manager.compress_reports(files, f"{filename}_paquete")


# ═══════════════════════════════════════════════════════════════
# EXPORTACIÓN
# ═══════════════════════════════════════════════════════════════

__all__ = [
    'ExportConfig',
    'ExportHistory',
    'ExportManager',
    'quick_export',
    'export_with_compression',
]


if __name__ == "__main__":
    print("""
    ═══════════════════════════════════════════════════════════════
    GESTOR DE EXPORTACIÓN - CHEDRAUI
    ═══════════════════════════════════════════════════════════════

    Formatos soportados:
    - Excel (.xlsx) con formato corporativo
    - CSV (.csv)
    - JSON (.json)
    - ZIP (compresión)

    Uso:
        from modules.export_manager import ExportManager

        manager = ExportManager()
        manager.export_to_excel(df, "mi_reporte", title="Mi Reporte")
        manager.export_to_csv(df, "datos")
        manager.compress_reports([archivo1, archivo2], "paquete")

    ═══════════════════════════════════════════════════════════════
    """)
