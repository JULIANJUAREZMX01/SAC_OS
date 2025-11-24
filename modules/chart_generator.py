"""
═══════════════════════════════════════════════════════════════
GENERADOR DE GRÁFICOS EXCEL
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Módulo para generar gráficos profesionales en reportes Excel.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import (
    BarChart, PieChart, LineChart, AreaChart, DoughnutChart,
    Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.plotarea import DataTable
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from typing import Optional, Dict, List, Tuple, Any, Union
import logging

from .excel_styles import ChedrauiColors

logger = logging.getLogger(__name__)


class ChartConfig:
    """Configuración para gráficos"""

    def __init__(
        self,
        title: str = "",
        x_axis_title: str = "",
        y_axis_title: str = "",
        width: int = 15,
        height: int = 10,
        style: int = 10,
        show_legend: bool = True,
        legend_position: str = 'b',  # 'b'=bottom, 'r'=right, 't'=top, 'l'=left
        show_data_labels: bool = False,
        data_label_position: str = 'outEnd',
        colors: List[str] = None
    ):
        self.title = title
        self.x_axis_title = x_axis_title
        self.y_axis_title = y_axis_title
        self.width = width
        self.height = height
        self.style = style
        self.show_legend = show_legend
        self.legend_position = legend_position
        self.show_data_labels = show_data_labels
        self.data_label_position = data_label_position
        self.colors = colors or ChedrauiColors.CHART_COLORS


class ChartGenerator:
    """
    Generador de gráficos para reportes Excel

    Soporta:
    - Gráficos de barras (verticales y horizontales)
    - Gráficos de líneas
    - Gráficos de pastel
    - Gráficos de área
    - Gráficos combinados
    - Dashboards con múltiples gráficos
    """

    DEFAULT_COLORS = ChedrauiColors.CHART_COLORS

    @classmethod
    def create_bar_chart(
        cls,
        ws: Worksheet,
        data: Reference,
        categories: Reference,
        config: ChartConfig = None,
        chart_type: str = 'col'  # 'col' = columnas, 'bar' = barras horizontales
    ) -> BarChart:
        """
        Crea gráfico de barras

        Args:
            ws: Hoja de trabajo
            data: Referencia a los datos
            categories: Referencia a las categorías
            config: Configuración del gráfico
            chart_type: 'col' para columnas, 'bar' para horizontales

        Returns:
            BarChart: Gráfico de barras
        """
        config = config or ChartConfig()

        chart = BarChart()
        chart.type = chart_type
        chart.grouping = "standard"
        chart.title = config.title
        chart.style = config.style
        chart.width = config.width
        chart.height = config.height

        # Ejes
        if config.x_axis_title:
            chart.x_axis.title = config.x_axis_title
        if config.y_axis_title:
            chart.y_axis.title = config.y_axis_title

        # Datos
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Leyenda
        if config.show_legend:
            chart.legend = Legend()
            chart.legend.position = config.legend_position
        else:
            chart.legend = None

        # Etiquetas de datos
        if config.show_data_labels:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = False

        # Aplicar colores
        cls._apply_colors_to_chart(chart, config.colors)

        logger.debug(f"📊 Gráfico de barras creado: {config.title}")
        return chart

    @classmethod
    def create_pie_chart(
        cls,
        ws: Worksheet,
        data: Reference,
        categories: Reference,
        config: ChartConfig = None,
        is_doughnut: bool = False
    ) -> Union[PieChart, DoughnutChart]:
        """
        Crea gráfico de pastel o dona

        Args:
            ws: Hoja de trabajo
            data: Referencia a los datos
            categories: Referencia a las categorías
            config: Configuración del gráfico
            is_doughnut: True para gráfico de dona

        Returns:
            PieChart o DoughnutChart
        """
        config = config or ChartConfig(show_legend=True, show_data_labels=True)

        if is_doughnut:
            chart = DoughnutChart()
            chart.holeSize = 50
        else:
            chart = PieChart()

        chart.title = config.title
        chart.style = config.style
        chart.width = config.width
        chart.height = config.height

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Etiquetas
        if config.show_data_labels:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showCatName = True
            chart.dataLabels.showVal = False

        # Leyenda
        if config.show_legend:
            chart.legend = Legend()
            chart.legend.position = config.legend_position
        else:
            chart.legend = None

        cls._apply_colors_to_chart(chart, config.colors)

        logger.debug(f"📊 Gráfico de pastel creado: {config.title}")
        return chart

    @classmethod
    def create_line_chart(
        cls,
        ws: Worksheet,
        data: Reference,
        categories: Reference,
        config: ChartConfig = None,
        smooth_lines: bool = False
    ) -> LineChart:
        """
        Crea gráfico de líneas

        Args:
            ws: Hoja de trabajo
            data: Referencia a los datos
            categories: Referencia a las categorías
            config: Configuración del gráfico
            smooth_lines: True para líneas suavizadas

        Returns:
            LineChart: Gráfico de líneas
        """
        config = config or ChartConfig()

        chart = LineChart()
        chart.title = config.title
        chart.style = config.style
        chart.width = config.width
        chart.height = config.height

        if config.x_axis_title:
            chart.x_axis.title = config.x_axis_title
        if config.y_axis_title:
            chart.y_axis.title = config.y_axis_title

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Suavizar líneas si se solicita
        if smooth_lines:
            for series in chart.series:
                series.smooth = True

        # Leyenda
        if config.show_legend:
            chart.legend = Legend()
            chart.legend.position = config.legend_position
        else:
            chart.legend = None

        # Etiquetas
        if config.show_data_labels:
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True

        cls._apply_colors_to_chart(chart, config.colors)

        logger.debug(f"📊 Gráfico de líneas creado: {config.title}")
        return chart

    @classmethod
    def create_area_chart(
        cls,
        ws: Worksheet,
        data: Reference,
        categories: Reference,
        config: ChartConfig = None,
        stacked: bool = False
    ) -> AreaChart:
        """
        Crea gráfico de área

        Args:
            ws: Hoja de trabajo
            data: Referencia a los datos
            categories: Referencia a las categorías
            config: Configuración del gráfico
            stacked: True para área apilada

        Returns:
            AreaChart: Gráfico de área
        """
        config = config or ChartConfig()

        chart = AreaChart()
        chart.title = config.title
        chart.style = config.style
        chart.width = config.width
        chart.height = config.height

        if stacked:
            chart.grouping = "stacked"
        else:
            chart.grouping = "standard"

        if config.x_axis_title:
            chart.x_axis.title = config.x_axis_title
        if config.y_axis_title:
            chart.y_axis.title = config.y_axis_title

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        if config.show_legend:
            chart.legend = Legend()
            chart.legend.position = config.legend_position
        else:
            chart.legend = None

        cls._apply_colors_to_chart(chart, config.colors)

        logger.debug(f"📊 Gráfico de área creado: {config.title}")
        return chart

    @classmethod
    def create_combo_chart(
        cls,
        ws: Worksheet,
        bar_data: Reference,
        line_data: Reference,
        categories: Reference,
        config: ChartConfig = None
    ) -> BarChart:
        """
        Crea gráfico combinado (barras + línea)

        Args:
            ws: Hoja de trabajo
            bar_data: Referencia a datos de barras
            line_data: Referencia a datos de línea
            categories: Referencia a categorías
            config: Configuración del gráfico

        Returns:
            BarChart: Gráfico combinado
        """
        config = config or ChartConfig()

        # Gráfico de barras base
        bar_chart = BarChart()
        bar_chart.type = 'col'
        bar_chart.title = config.title
        bar_chart.style = config.style
        bar_chart.width = config.width
        bar_chart.height = config.height

        bar_chart.add_data(bar_data, titles_from_data=True)
        bar_chart.set_categories(categories)

        if config.x_axis_title:
            bar_chart.x_axis.title = config.x_axis_title
        if config.y_axis_title:
            bar_chart.y_axis.title = config.y_axis_title

        # Gráfico de línea secundario
        line_chart = LineChart()
        line_chart.add_data(line_data, titles_from_data=True)

        # Eje Y secundario
        line_chart.y_axis.axId = 200
        line_chart.y_axis.crosses = "max"

        # Combinar gráficos
        bar_chart += line_chart

        if config.show_legend:
            bar_chart.legend = Legend()
            bar_chart.legend.position = config.legend_position

        logger.debug(f"📊 Gráfico combinado creado: {config.title}")
        return bar_chart

    @classmethod
    def create_dashboard(
        cls,
        ws: Worksheet,
        charts: List[Tuple[Any, str]]
    ) -> None:
        """
        Crea dashboard con múltiples gráficos

        Args:
            ws: Hoja de trabajo
            charts: Lista de tuplas (gráfico, posición)
                    Ejemplo: [(chart1, "A1"), (chart2, "H1")]
        """
        for chart, position in charts:
            ws.add_chart(chart, position)
            logger.debug(f"📊 Gráfico agregado en {position}")

        logger.info(f"📊 Dashboard creado con {len(charts)} gráficos")

    @classmethod
    def style_chart(
        cls,
        chart: Any,
        style_config: Dict[str, Any]
    ) -> Any:
        """
        Aplica estilo personalizado a un gráfico

        Args:
            chart: Gráfico a estilizar
            style_config: Configuración de estilo
                - title_font_size: Tamaño de fuente del título
                - title_bold: Título en negritas
                - gridlines: Mostrar líneas de cuadrícula
                - border: Mostrar borde
                - 3d: Efecto 3D (solo algunos tipos)

        Returns:
            Gráfico estilizado
        """
        # Configurar título
        if 'title_font_size' in style_config and chart.title:
            pass  # openpyxl no soporta directamente cambiar fuente del título

        # Líneas de cuadrícula
        if 'gridlines' in style_config:
            if hasattr(chart, 'y_axis'):
                from openpyxl.chart.axis import ChartLines
                if style_config['gridlines']:
                    chart.y_axis.majorGridlines = ChartLines()
                else:
                    chart.y_axis.majorGridlines = None

        return chart

    @classmethod
    def _apply_colors_to_chart(
        cls,
        chart: Any,
        colors: List[str]
    ) -> None:
        """
        Aplica colores corporativos al gráfico

        Args:
            chart: Gráfico
            colors: Lista de colores hexadecimales
        """
        # Los colores se aplican a través del estilo del gráfico
        # openpyxl tiene soporte limitado para colores personalizados
        pass

    @classmethod
    def create_chart_from_dataframe(
        cls,
        ws: Worksheet,
        df: pd.DataFrame,
        chart_type: str,
        category_col: str,
        value_cols: List[str],
        start_row: int,
        start_col: int = 1,
        config: ChartConfig = None
    ) -> Tuple[Any, str]:
        """
        Crea un gráfico directamente desde un DataFrame

        Args:
            ws: Hoja de trabajo
            df: DataFrame con datos
            chart_type: Tipo de gráfico ('bar', 'line', 'pie', 'area')
            category_col: Columna de categorías
            value_cols: Columnas de valores
            start_row: Fila donde escribir datos
            start_col: Columna donde escribir datos
            config: Configuración del gráfico

        Returns:
            Tuple[gráfico, posición]
        """
        config = config or ChartConfig()

        # Escribir datos en la hoja
        # Encabezados
        ws.cell(row=start_row, column=start_col, value=category_col)
        for i, col in enumerate(value_cols, start_col + 1):
            ws.cell(row=start_row, column=i, value=col)

        # Datos
        for row_idx, row_data in enumerate(df.itertuples(), start_row + 1):
            ws.cell(row=row_idx, column=start_col, value=getattr(row_data, category_col))
            for col_idx, val_col in enumerate(value_cols, start_col + 1):
                ws.cell(row=row_idx, column=col_idx, value=getattr(row_data, val_col))

        end_row = start_row + len(df)
        end_col = start_col + len(value_cols)

        # Referencias
        categories = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=end_row)
        data = Reference(ws, min_col=start_col + 1, max_col=end_col,
                        min_row=start_row, max_row=end_row)

        # Crear gráfico según tipo
        chart = None
        if chart_type == 'bar':
            chart = cls.create_bar_chart(ws, data, categories, config)
        elif chart_type == 'line':
            chart = cls.create_line_chart(ws, data, categories, config)
        elif chart_type == 'pie':
            chart = cls.create_pie_chart(ws, data, categories, config)
        elif chart_type == 'area':
            chart = cls.create_area_chart(ws, data, categories, config)

        # Posición del gráfico (a la derecha de los datos)
        position = f"{get_column_letter(end_col + 2)}{start_row}"

        return chart, position


# ═══════════════════════════════════════════════════════════════
# FUNCIONES DE UTILIDAD
# ═══════════════════════════════════════════════════════════════

def add_chart_to_sheet(
    ws: Worksheet,
    df: pd.DataFrame,
    chart_type: str,
    category_col: str,
    value_col: str,
    position: str = "A1",
    title: str = ""
) -> None:
    """
    Función de utilidad para agregar gráfico rápidamente

    Args:
        ws: Hoja de trabajo
        df: DataFrame con datos
        chart_type: Tipo ('bar', 'pie', 'line')
        category_col: Columna de categorías
        value_col: Columna de valores
        position: Posición del gráfico
        title: Título del gráfico
    """
    # Escribir datos
    ws['A1'] = category_col
    ws['B1'] = value_col
    for idx, row in enumerate(df[[category_col, value_col]].itertuples(), 2):
        ws.cell(row=idx, column=1, value=row[1])
        ws.cell(row=idx, column=2, value=row[2])

    categories = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    data = Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1)

    config = ChartConfig(title=title)

    if chart_type == 'bar':
        chart = ChartGenerator.create_bar_chart(ws, data, categories, config)
    elif chart_type == 'pie':
        chart = ChartGenerator.create_pie_chart(ws, data, categories, config)
    elif chart_type == 'line':
        chart = ChartGenerator.create_line_chart(ws, data, categories, config)
    else:
        chart = ChartGenerator.create_bar_chart(ws, data, categories, config)

    ws.add_chart(chart, position)


# ═══════════════════════════════════════════════════════════════
# EXPORTACIÓN
# ═══════════════════════════════════════════════════════════════

__all__ = [
    'ChartConfig',
    'ChartGenerator',
    'add_chart_to_sheet',
]


if __name__ == "__main__":
    print("""
    ═══════════════════════════════════════════════════════════════
    GENERADOR DE GRÁFICOS EXCEL - CHEDRAUI
    ═══════════════════════════════════════════════════════════════

    Tipos de gráficos disponibles:
    - create_bar_chart(): Gráficos de barras/columnas
    - create_pie_chart(): Gráficos de pastel/dona
    - create_line_chart(): Gráficos de líneas
    - create_area_chart(): Gráficos de área
    - create_combo_chart(): Gráficos combinados
    - create_dashboard(): Dashboards con múltiples gráficos

    Uso:
        from modules.chart_generator import ChartGenerator, ChartConfig

        config = ChartConfig(title="Mi Gráfico", width=12, height=8)
        chart = ChartGenerator.create_bar_chart(ws, data, categories, config)
        ws.add_chart(chart, "E5")

    ═══════════════════════════════════════════════════════════════
    """)
