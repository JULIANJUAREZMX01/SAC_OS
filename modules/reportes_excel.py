"""
═══════════════════════════════════════════════════════════════
MÓDULO DE GENERACIÓN DE REPORTES EN EXCEL
Sistema de Gestión de Órdenes de Compra - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Este módulo genera reportes profesionales en Excel:
- Formato corporativo Chedraui
- Tablas dinámicas y gráficos
- Validaciones y resaltado de errores
- Exportación automática
- 10 tipos de reportes especializados

Desarrollado por: Julián Alexander Juárez Alvarado (ADM)
Analista de Sistemas - CEDIS Chedraui Logística Cancún
═══════════════════════════════════════════════════════════════
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime
import logging
from typing import Optional, Dict, List, Any, Tuple
from pathlib import Path

# Importar nuevos módulos del sistema de reportes
from .excel_styles import ChedrauiStyles, ChedrauiColors, apply_style_dict
from .chart_generator import ChartGenerator, ChartConfig
from .pivot_generator import PivotGenerator, PivotConfig

logger = logging.getLogger(__name__)


class GeneradorReportesExcel:
    """
    Generador de reportes profesionales en Excel con formato Chedraui

    Los colores se importan desde config.py (fuente única de verdad).
    """

    # Colores corporativos - importados desde ChedrauiColors (config.py)
    COLOR_HEADER = ChedrauiColors.HEADER_RED
    COLOR_SUBHEADER = ChedrauiColors.SUBHEADER
    COLOR_CRITICO = ChedrauiColors.CRITICO
    COLOR_ALERTA = ChedrauiColors.ALTO
    COLOR_OK = ChedrauiColors.OK_GREEN
    COLOR_INFO = ChedrauiColors.INFO_BLUE

    def __init__(self, cedis: str = "CANCÚN"):
        self.cedis = cedis
        
    def crear_reporte_validacion_oc(
        self, 
        df_validacion: pd.DataFrame, 
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte de Validación OC vs Distribuciones
        
        Args:
            df_validacion: DataFrame con datos de validación
            nombre_archivo: Nombre del archivo (opcional)
            
        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Validacion_OC_Distros_{timestamp}.xlsx"
        
        logger.info(f"📊 Generando reporte: {nombre_archivo}")
        
        # Crear el archivo Excel
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            # Escribir datos principales
            df_validacion.to_excel(writer, sheet_name='Validación OC', index=False, startrow=4)
            
            # Obtener el workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Validación OC']
            
            # Agregar encabezado corporativo
            self._agregar_encabezado(
                worksheet, 
                titulo="REPORTE DE VALIDACIÓN",
                subtitulo="Órdenes de Compra vs Distribuciones",
                cedis=self.cedis
            )
            
            # Formatear la tabla
            self._formatear_tabla_validacion(worksheet, df_validacion, start_row=5)
            
            # Agregar resumen
            self._agregar_resumen_validacion(worksheet, df_validacion)
            
            # Crear hoja de estadísticas
            self._crear_hoja_estadisticas(workbook, df_validacion)
            
            # Ajustar anchos de columna
            self._ajustar_columnas(worksheet)
        
        logger.info(f"✅ Reporte generado: {nombre_archivo}")
        return nombre_archivo
    
    def crear_reporte_distribuciones(
        self, 
        df_distribuciones: pd.DataFrame,
        oc_numero: str,
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte detallado de distribuciones por OC
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Distribuciones_OC_{oc_numero}_{timestamp}.xlsx"
        
        logger.info(f"📊 Generando reporte de distribuciones: {nombre_archivo}")
        
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            # Hoja principal de distribuciones
            df_distribuciones.to_excel(writer, sheet_name='Distribuciones', index=False, startrow=4)
            
            workbook = writer.book
            worksheet = writer.sheets['Distribuciones']
            
            # Encabezado
            self._agregar_encabezado(
                worksheet,
                titulo=f"DISTRIBUCIONES - OC {oc_numero}",
                subtitulo="Detalle por Tienda y SKU",
                cedis=self.cedis
            )
            
            # Formatear tabla
            self._formatear_tabla_distribuciones(worksheet, df_distribuciones, start_row=5)
            
            # Crear pivot por tienda
            if not df_distribuciones.empty:
                self._crear_pivot_tiendas(workbook, df_distribuciones)
            
            # Ajustar columnas
            self._ajustar_columnas(worksheet)
        
        logger.info(f"✅ Reporte de distribuciones generado")
        return nombre_archivo
    
    def crear_reporte_planning_diario(
        self,
        df_oc: pd.DataFrame,
        df_asn: pd.DataFrame,
        df_errores: pd.DataFrame,
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte diario de Planning completo
        """
        if nombre_archivo is None:
            fecha = datetime.now().strftime('%Y%m%d')
            nombre_archivo = f"Planning_Diario_{self.cedis}_{fecha}.xlsx"
        
        logger.info(f"📊 Generando reporte diario de Planning")
        
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            workbook = writer.book
            
            # HOJA 1: Resumen Ejecutivo
            self._crear_hoja_resumen_ejecutivo(workbook, df_oc, df_asn, df_errores)
            
            # HOJA 2: Órdenes de Compra
            if not df_oc.empty:
                df_oc.to_excel(writer, sheet_name='Órdenes de Compra', index=False, startrow=3)
                ws_oc = writer.sheets['Órdenes de Compra']
                self._agregar_encabezado_simple(ws_oc, "ÓRDENES DE COMPRA DEL DÍA")
                self._formatear_tabla_generica(ws_oc, df_oc, start_row=4)
            
            # HOJA 3: ASN Status
            if not df_asn.empty:
                df_asn.to_excel(writer, sheet_name='Status ASN', index=False, startrow=3)
                ws_asn = writer.sheets['Status ASN']
                self._agregar_encabezado_simple(ws_asn, "STATUS DE ASN")
                self._formatear_tabla_generica(ws_asn, df_asn, start_row=4)
            
            # HOJA 4: Errores Detectados
            if not df_errores.empty:
                df_errores.to_excel(writer, sheet_name='Errores', index=False, startrow=3)
                ws_errores = writer.sheets['Errores']
                self._agregar_encabezado_simple(ws_errores, "ERRORES Y ALERTAS")
                self._formatear_tabla_errores(ws_errores, df_errores, start_row=4)
        
        logger.info(f"✅ Reporte diario generado: {nombre_archivo}")
        return nombre_archivo
    
    def _agregar_encabezado(self, worksheet, titulo: str, subtitulo: str, cedis: str):
        """Agrega encabezado corporativo Chedraui"""
        # Fila 1: Logo y título principal
        worksheet['A1'] = 'CHEDRAUI'
        worksheet['A1'].font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
        worksheet['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, 
                                           end_color=self.COLOR_HEADER, 
                                           fill_type='solid')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge para el logo
        worksheet.merge_cells('A1:D1')
        
        # Fila 2: Título del reporte
        worksheet['A2'] = titulo
        worksheet['A2'].font = Font(name='Arial', size=14, bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='left')
        worksheet.merge_cells('A2:D2')
        
        # Fila 3: Subtítulo e información
        worksheet['A3'] = subtitulo
        worksheet['A3'].font = Font(name='Arial', size=11, italic=True)
        
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        worksheet['E3'] = f"CEDIS: {cedis}"
        worksheet['E3'].font = Font(name='Arial', size=10)
        worksheet['E3'].alignment = Alignment(horizontal='right')
        
        worksheet['G3'] = f"Fecha: {fecha_actual}"
        worksheet['G3'].font = Font(name='Arial', size=10)
        worksheet['G3'].alignment = Alignment(horizontal='right')
    
    def _agregar_encabezado_simple(self, worksheet, titulo: str):
        """Agrega encabezado simple"""
        worksheet['A1'] = titulo
        worksheet['A1'].font = Font(name='Arial', size=14, bold=True)
        worksheet['A1'].fill = PatternFill(start_color=self.COLOR_HEADER,
                                          end_color=self.COLOR_HEADER,
                                          fill_type='solid')
        worksheet['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        worksheet.merge_cells('A1:F1')
        
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        worksheet['A2'] = f"Generado: {fecha_actual}"
        worksheet['A2'].font = Font(name='Arial', size=9, italic=True)
    
    def _formatear_tabla_validacion(self, worksheet, df: pd.DataFrame, start_row: int):
        """Formatea tabla de validación con colores según status"""
        # Formatear encabezados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Formatear datos según status
        if 'STATUS' in df.columns:
            status_col = df.columns.get_loc('STATUS') + 1
            
            for row_num in range(start_row + 1, start_row + len(df) + 1):
                status_cell = worksheet.cell(row=row_num, column=status_col)
                status_value = status_cell.value
                
                # Colorear según status
                if status_value and 'excedente' in str(status_value).lower():
                    fill_color = self.COLOR_CRITICO
                elif status_value and 'incompleta' in str(status_value).lower():
                    fill_color = self.COLOR_ALERTA
                elif status_value and 'sin distro' in str(status_value).lower():
                    fill_color = self.COLOR_CRITICO
                elif status_value and 'ok' in str(status_value).lower():
                    fill_color = self.COLOR_OK
                else:
                    fill_color = self.COLOR_INFO
                
                # Aplicar color a toda la fila
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color=fill_color,
                                          end_color=fill_color,
                                          fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
    
    def _formatear_tabla_distribuciones(self, worksheet, df: pd.DataFrame, start_row: int):
        """Formatea tabla de distribuciones"""
        # Encabezados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar bordes a toda la tabla
        for row in range(start_row, start_row + len(df) + 1):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    def _formatear_tabla_generica(self, worksheet, df: pd.DataFrame, start_row: int):
        """Formato genérico para tablas"""
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_SUBHEADER,
                                   end_color=self.COLOR_SUBHEADER,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
    
    def _formatear_tabla_errores(self, worksheet, df: pd.DataFrame, start_row: int):
        """Formatea tabla de errores con colores según severidad"""
        # Encabezados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type='solid')
        
        # Colorear según severidad
        if 'Severidad' in df.columns:
            sev_col = df.columns.get_loc('Severidad') + 1
            
            for row_num in range(start_row + 1, start_row + len(df) + 1):
                sev_cell = worksheet.cell(row=row_num, column=sev_col)
                severidad = str(sev_cell.value).lower()
                
                if 'crítico' in severidad:
                    color = self.COLOR_CRITICO
                elif 'alto' in severidad:
                    color = self.COLOR_ALERTA
                elif 'medio' in severidad:
                    color = 'FFE699'
                else:
                    color = self.COLOR_INFO
                
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color=color,
                                          end_color=color,
                                          fill_type='solid')
    
    def _agregar_resumen_validacion(self, worksheet, df: pd.DataFrame):
        """Agrega resumen de validación en la parte superior"""
        if 'STATUS' not in df.columns:
            return
        
        # Calcular estadísticas
        total_oc = len(df)
        ok = len(df[df['STATUS'].str.contains('OK', case=False, na=False)])
        incompletas = len(df[df['STATUS'].str.contains('incompleta', case=False, na=False)])
        excedentes = len(df[df['STATUS'].str.contains('excedente', case=False, na=False)])
        sin_distro = len(df[df['STATUS'].str.contains('sin distro', case=False, na=False)])
        
        # Agregar resumen a la derecha
        col_resumen = len(df.columns) + 2
        
        worksheet.cell(row=5, column=col_resumen, value="RESUMEN")
        worksheet.cell(row=5, column=col_resumen).font = Font(bold=True, size=12)
        
        worksheet.cell(row=6, column=col_resumen, value="Total OC:")
        worksheet.cell(row=6, column=col_resumen + 1, value=total_oc)
        
        worksheet.cell(row=7, column=col_resumen, value="✅ OK:")
        worksheet.cell(row=7, column=col_resumen + 1, value=ok)
        worksheet.cell(row=7, column=col_resumen + 1).fill = PatternFill(
            start_color=self.COLOR_OK, end_color=self.COLOR_OK, fill_type='solid')
        
        worksheet.cell(row=8, column=col_resumen, value="⚠️ Incompletas:")
        worksheet.cell(row=8, column=col_resumen + 1, value=incompletas)
        worksheet.cell(row=8, column=col_resumen + 1).fill = PatternFill(
            start_color=self.COLOR_ALERTA, end_color=self.COLOR_ALERTA, fill_type='solid')
        
        worksheet.cell(row=9, column=col_resumen, value="🔴 Excedentes:")
        worksheet.cell(row=9, column=col_resumen + 1, value=excedentes)
        worksheet.cell(row=9, column=col_resumen + 1).fill = PatternFill(
            start_color=self.COLOR_CRITICO, end_color=self.COLOR_CRITICO, fill_type='solid')
        
        worksheet.cell(row=10, column=col_resumen, value="❌ Sin Distro:")
        worksheet.cell(row=10, column=col_resumen + 1, value=sin_distro)
        worksheet.cell(row=10, column=col_resumen + 1).fill = PatternFill(
            start_color=self.COLOR_CRITICO, end_color=self.COLOR_CRITICO, fill_type='solid')
    
    def _crear_hoja_estadisticas(self, workbook, df: pd.DataFrame):
        """Crea hoja con estadísticas y gráficos"""
        ws_stats = workbook.create_sheet('Estadísticas')
        
        # Título
        ws_stats['A1'] = 'ESTADÍSTICAS DE VALIDACIÓN'
        ws_stats['A1'].font = Font(name='Arial', size=16, bold=True)
        
        # Agregar estadísticas aquí
        # (Se pueden agregar más análisis según necesidad)
    
    def _crear_pivot_tiendas(self, workbook, df: pd.DataFrame):
        """Crea pivot de distribuciones por tienda"""
        if 'TIENDA' not in df.columns or 'DISTR_QTY' not in df.columns:
            return
        
        pivot = df.groupby('TIENDA')['DISTR_QTY'].sum().reset_index()
        pivot.columns = ['Tienda', 'Total Unidades']
        
        ws_pivot = workbook.create_sheet('Por Tienda')
        
        for row in dataframe_to_rows(pivot, index=False, header=True):
            ws_pivot.append(row)
        
        # Formatear
        for cell in ws_pivot[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.COLOR_SUBHEADER,
                                   end_color=self.COLOR_SUBHEADER,
                                   fill_type='solid')
    
    def _crear_hoja_resumen_ejecutivo(self, workbook, df_oc, df_asn, df_errores):
        """Crea hoja de resumen ejecutivo"""
        ws = workbook.create_sheet('Resumen Ejecutivo', 0)
        
        # Título principal
        ws['A1'] = '📊 RESUMEN EJECUTIVO - PLANNING DIARIO'
        ws['A1'].font = Font(name='Arial', size=18, bold=True)
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type='solid')
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
        ws.merge_cells('A1:F1')
        
        # Información general
        row = 3
        ws[f'A{row}'] = f"CEDIS: {self.cedis}"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        ws[f'A{row}'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        row += 2
        ws[f'A{row}'] = "INDICADORES PRINCIPALES"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].fill = PatternFill(start_color=self.COLOR_SUBHEADER,
                                         end_color=self.COLOR_SUBHEADER,
                                         fill_type='solid')
        
        row += 1
        ws[f'A{row}'] = "Total Órdenes de Compra:"
        ws[f'B{row}'] = len(df_oc) if not df_oc.empty else 0
        ws[f'B{row}'].font = Font(bold=True, size=12)
        
        row += 1
        ws[f'A{row}'] = "Total ASN Activos:"
        ws[f'B{row}'] = len(df_asn) if not df_asn.empty else 0
        ws[f'B{row}'].font = Font(bold=True, size=12)
        
        row += 1
        ws[f'A{row}'] = "Errores Detectados:"
        ws[f'B{row}'] = len(df_errores) if not df_errores.empty else 0
        if len(df_errores) > 0:
            ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLOR_CRITICO)
        else:
            ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLOR_OK)
        
        # Pie de página con créditos
        row += 10
        ws[f'A{row}'] = "─" * 80
        row += 1
        ws[f'A{row}'] = "Sistema desarrollado por: Julián Alexander Juárez Alvarado (ADM)"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 1
        ws[f'A{row}'] = "Analista de Sistemas - CEDIS Chedraui Logística Cancún"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 1
        ws[f'A{row}'] = '"Las máquinas y los sistemas al servicio de los analistas"'
        ws[f'A{row}'].font = Font(italic=True, size=9, color='666666')
    
    def _ajustar_columnas(self, worksheet):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (ValueError, TypeError, AttributeError):
                    # Valor None o tipo no convertible a string - continuar con siguiente celda
                    continue

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # ═══════════════════════════════════════════════════════════════
    # NUEVOS TIPOS DE REPORTES
    # ═══════════════════════════════════════════════════════════════

    def crear_reporte_asn_status(
        self,
        df_asn: pd.DataFrame,
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte de status de ASN (Advanced Shipping Notices)

        Args:
            df_asn: DataFrame con datos de ASN
            nombre_archivo: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"ASN_Status_{self.cedis}_{timestamp}.xlsx"

        logger.info(f"📊 Generando reporte de ASN Status: {nombre_archivo}")

        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            # Hoja principal
            df_asn.to_excel(writer, sheet_name='Status ASN', index=False, startrow=4)

            workbook = writer.book
            worksheet = writer.sheets['Status ASN']

            # Encabezado corporativo
            self._agregar_encabezado(
                worksheet,
                titulo="REPORTE DE STATUS ASN",
                subtitulo="Advanced Shipping Notices - Estado Actual",
                cedis=self.cedis
            )

            # Formatear tabla
            self._formatear_tabla_asn(worksheet, df_asn, start_row=5)

            # Crear resumen por status
            if 'STATUS' in df_asn.columns and not df_asn.empty:
                self._crear_resumen_asn(workbook, df_asn)

            self._ajustar_columnas(worksheet)

        logger.info(f"✅ Reporte ASN Status generado: {nombre_archivo}")
        return nombre_archivo

    def _formatear_tabla_asn(self, worksheet, df: pd.DataFrame, start_row: int):
        """Formatea tabla de ASN con colores según status"""
        # Encabezados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Colorear según status si existe la columna
        if 'STATUS' in df.columns:
            status_col = df.columns.get_loc('STATUS') + 1

            for row_num in range(start_row + 1, start_row + len(df) + 1):
                status_cell = worksheet.cell(row=row_num, column=status_col)
                status_value = str(status_cell.value).lower() if status_cell.value else ""

                if 'recibido' in status_value or 'complete' in status_value:
                    fill_color = self.COLOR_OK
                elif 'pendiente' in status_value or 'pending' in status_value:
                    fill_color = self.COLOR_INFO
                elif 'error' in status_value or 'cancelado' in status_value:
                    fill_color = self.COLOR_CRITICO
                elif 'parcial' in status_value:
                    fill_color = self.COLOR_ALERTA
                else:
                    fill_color = 'FFFFFF'

                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if fill_color != 'FFFFFF':
                        cell.fill = PatternFill(start_color=fill_color,
                                              end_color=fill_color,
                                              fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

    def _crear_resumen_asn(self, workbook, df: pd.DataFrame):
        """Crea hoja de resumen de ASN"""
        ws_resumen = workbook.create_sheet('Resumen ASN')

        # Título
        ws_resumen['A1'] = 'RESUMEN DE ASN POR STATUS'
        ws_resumen['A1'].font = Font(name='Arial', size=14, bold=True)
        ws_resumen.merge_cells('A1:C1')

        # Estadísticas
        resumen = df['STATUS'].value_counts().reset_index()
        resumen.columns = ['Status', 'Cantidad']

        for idx, row in enumerate(resumen.itertuples(), 3):
            ws_resumen.cell(row=idx, column=1, value=row.Status)
            ws_resumen.cell(row=idx, column=2, value=row.Cantidad)

        # Gráfico de pastel si hay datos
        if len(resumen) > 0:
            pie = PieChart()
            labels = Reference(ws_resumen, min_col=1, min_row=3, max_row=2 + len(resumen))
            data = Reference(ws_resumen, min_col=2, min_row=2, max_row=2 + len(resumen))
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Distribución por Status"
            ws_resumen.add_chart(pie, "E3")

    def crear_reporte_inventario(
        self,
        df_inventario: pd.DataFrame,
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte de inventario en almacén

        Args:
            df_inventario: DataFrame con datos de inventario
            nombre_archivo: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Inventario_{self.cedis}_{timestamp}.xlsx"

        logger.info(f"📊 Generando reporte de inventario: {nombre_archivo}")

        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            df_inventario.to_excel(writer, sheet_name='Inventario', index=False, startrow=4)

            workbook = writer.book
            worksheet = writer.sheets['Inventario']

            self._agregar_encabezado(
                worksheet,
                titulo="REPORTE DE INVENTARIO",
                subtitulo="Estado actual del inventario en almacén",
                cedis=self.cedis
            )

            self._formatear_tabla_inventario(worksheet, df_inventario, start_row=5)

            # Agregar estadísticas
            if not df_inventario.empty:
                self._crear_estadisticas_inventario(workbook, df_inventario)

            self._ajustar_columnas(worksheet)

        logger.info(f"✅ Reporte de inventario generado: {nombre_archivo}")
        return nombre_archivo

    def _formatear_tabla_inventario(self, worksheet, df: pd.DataFrame, start_row: int):
        """Formatea tabla de inventario con alertas de stock"""
        # Encabezados
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=start_row, column=col_num)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLOR_HEADER,
                                   end_color=self.COLOR_HEADER,
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Buscar columna de stock
        stock_cols = ['STOCK', 'CANTIDAD', 'QTY', 'ON_HAND']
        stock_col_idx = None
        for col_name in stock_cols:
            if col_name in df.columns:
                stock_col_idx = df.columns.get_loc(col_name) + 1
                break

        # Aplicar formato a datos
        for row_num in range(start_row + 1, start_row + len(df) + 1):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Colorear por nivel de stock
            if stock_col_idx:
                stock_cell = worksheet.cell(row=row_num, column=stock_col_idx)
                try:
                    stock_value = float(stock_cell.value) if stock_cell.value else 0
                    if stock_value <= 0:
                        stock_cell.fill = PatternFill(start_color=self.COLOR_CRITICO,
                                                     end_color=self.COLOR_CRITICO,
                                                     fill_type='solid')
                    elif stock_value < 10:
                        stock_cell.fill = PatternFill(start_color=self.COLOR_ALERTA,
                                                     end_color=self.COLOR_ALERTA,
                                                     fill_type='solid')
                except (ValueError, TypeError):
                    pass

    def _crear_estadisticas_inventario(self, workbook, df: pd.DataFrame):
        """Crea hoja de estadísticas de inventario"""
        ws_stats = workbook.create_sheet('Estadísticas')

        ws_stats['A1'] = 'ESTADÍSTICAS DE INVENTARIO'
        ws_stats['A1'].font = Font(name='Arial', size=14, bold=True)

        # Buscar columna de stock
        stock_col = None
        for col in ['STOCK', 'CANTIDAD', 'QTY', 'ON_HAND']:
            if col in df.columns:
                stock_col = col
                break

        if stock_col:
            ws_stats['A3'] = 'Total SKUs:'
            ws_stats['B3'] = len(df)

            ws_stats['A4'] = 'Stock Total:'
            ws_stats['B4'] = df[stock_col].sum()

            ws_stats['A5'] = 'Stock Promedio:'
            ws_stats['B5'] = round(df[stock_col].mean(), 2)

            ws_stats['A6'] = 'SKUs Sin Stock:'
            ws_stats['B6'] = len(df[df[stock_col] <= 0])

            ws_stats['A7'] = 'SKUs Stock Bajo (<10):'
            ws_stats['B7'] = len(df[(df[stock_col] > 0) & (df[stock_col] < 10)])

    def crear_reporte_recibo_programado(
        self,
        df_recibo: pd.DataFrame,
        fecha_programa: str = None,
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte de programa de recibo

        Args:
            df_recibo: DataFrame con datos de recibo programado
            fecha_programa: Fecha del programa
            nombre_archivo: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Recibo_Programado_{self.cedis}_{timestamp}.xlsx"

        if fecha_programa is None:
            fecha_programa = datetime.now().strftime('%d/%m/%Y')

        logger.info(f"📊 Generando programa de recibo: {nombre_archivo}")

        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            df_recibo.to_excel(writer, sheet_name='Programa Recibo', index=False, startrow=4)

            workbook = writer.book
            worksheet = writer.sheets['Programa Recibo']

            self._agregar_encabezado(
                worksheet,
                titulo=f"PROGRAMA DE RECIBO - {fecha_programa}",
                subtitulo="Entregas programadas del día",
                cedis=self.cedis
            )

            self._formatear_tabla_generica(worksheet, df_recibo, start_row=5)

            # Agregar resumen por proveedor si existe
            if 'PROVEEDOR' in df_recibo.columns and not df_recibo.empty:
                self._crear_resumen_proveedores(workbook, df_recibo)

            self._ajustar_columnas(worksheet)

        logger.info(f"✅ Programa de recibo generado: {nombre_archivo}")
        return nombre_archivo

    def _crear_resumen_proveedores(self, workbook, df: pd.DataFrame):
        """Crea resumen por proveedor"""
        ws = workbook.create_sheet('Por Proveedor')

        ws['A1'] = 'RESUMEN POR PROVEEDOR'
        ws['A1'].font = Font(name='Arial', size=14, bold=True)

        resumen = df.groupby('PROVEEDOR').size().reset_index(name='Entregas')
        resumen = resumen.sort_values('Entregas', ascending=False)

        ws['A3'] = 'Proveedor'
        ws['B3'] = 'Entregas'
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)

        for idx, row in enumerate(resumen.itertuples(), 4):
            ws.cell(row=idx, column=1, value=row.PROVEEDOR)
            ws.cell(row=idx, column=2, value=row.Entregas)

    def crear_reporte_kpis(
        self,
        kpis: Dict[str, Any],
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte de KPIs (Indicadores Clave de Desempeño)

        Args:
            kpis: Diccionario con KPIs
                  Formato: {'nombre_kpi': {'value': valor, 'target': meta, 'unit': unidad}}
            nombre_archivo: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"KPIs_{self.cedis}_{timestamp}.xlsx"

        logger.info(f"📊 Generando reporte de KPIs: {nombre_archivo}")

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'KPIs'

        # Encabezado
        worksheet.merge_cells('A1:F1')
        worksheet['A1'] = '📊 INDICADORES CLAVE DE DESEMPEÑO (KPIs)'
        worksheet['A1'].font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
        worksheet['A1'].fill = PatternFill(start_color=self.COLOR_HEADER,
                                           end_color=self.COLOR_HEADER,
                                           fill_type='solid')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[1].height = 40

        worksheet['A2'] = f"CEDIS: {self.cedis}"
        worksheet['A2'].font = Font(bold=True)
        worksheet['E2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        # Crear tarjetas de KPI
        row = 4
        col = 1

        for kpi_name, kpi_data in kpis.items():
            value = kpi_data.get('value', 0)
            target = kpi_data.get('target', None)
            unit = kpi_data.get('unit', '')
            trend = kpi_data.get('trend', 'neutral')

            # Nombre del KPI
            name_cell = worksheet.cell(row=row, column=col, value=kpi_name)
            name_cell.font = Font(name='Arial', size=11, bold=True)
            name_cell.fill = PatternFill(start_color=self.COLOR_SUBHEADER,
                                        end_color=self.COLOR_SUBHEADER,
                                        fill_type='solid')
            worksheet.merge_cells(start_row=row, start_column=col,
                                end_row=row, end_column=col + 1)

            # Valor del KPI
            row += 1
            value_cell = worksheet.cell(row=row, column=col, value=f"{value}{unit}")

            # Color según tendencia
            if trend == 'positive':
                value_cell.font = Font(name='Arial', size=24, bold=True,
                                      color=self.COLOR_OK)
            elif trend == 'negative':
                value_cell.font = Font(name='Arial', size=24, bold=True,
                                      color=self.COLOR_CRITICO)
            else:
                value_cell.font = Font(name='Arial', size=24, bold=True,
                                      color='003DA5')  # Azul Chedraui

            value_cell.alignment = Alignment(horizontal='center')
            worksheet.merge_cells(start_row=row, start_column=col,
                                end_row=row, end_column=col + 1)

            # Meta si existe
            if target is not None:
                row += 1
                meta_cell = worksheet.cell(row=row, column=col, value=f"Meta: {target}{unit}")
                meta_cell.font = Font(name='Arial', size=9, italic=True)
                meta_cell.alignment = Alignment(horizontal='center')

            row += 2

            # Siguiente columna después de 5 KPIs
            if row > 20:
                row = 4
                col += 3

        # Ajustar columnas
        for col_idx in range(1, 10):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 15

        workbook.save(nombre_archivo)
        logger.info(f"✅ Reporte de KPIs generado: {nombre_archivo}")
        return nombre_archivo

    def crear_reporte_auditoria(
        self,
        df_auditoria: pd.DataFrame,
        tipo_auditoria: str = "General",
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte de auditoría

        Args:
            df_auditoria: DataFrame con datos de auditoría
            tipo_auditoria: Tipo de auditoría
            nombre_archivo: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Auditoria_{tipo_auditoria}_{timestamp}.xlsx"

        logger.info(f"📊 Generando reporte de auditoría: {nombre_archivo}")

        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            df_auditoria.to_excel(writer, sheet_name='Auditoría', index=False, startrow=4)

            workbook = writer.book
            worksheet = writer.sheets['Auditoría']

            self._agregar_encabezado(
                worksheet,
                titulo=f"REPORTE DE AUDITORÍA - {tipo_auditoria.upper()}",
                subtitulo="Registro de acciones y cambios del sistema",
                cedis=self.cedis
            )

            self._formatear_tabla_generica(worksheet, df_auditoria, start_row=5)
            self._ajustar_columnas(worksheet)

        logger.info(f"✅ Reporte de auditoría generado: {nombre_archivo}")
        return nombre_archivo

    def crear_reporte_comparativo(
        self,
        df_actual: pd.DataFrame,
        df_anterior: pd.DataFrame,
        tipo_comparacion: str = "Período",
        nombre_archivo: str = None
    ) -> str:
        """
        Crea reporte comparativo entre dos períodos

        Args:
            df_actual: DataFrame con datos del período actual
            df_anterior: DataFrame con datos del período anterior
            tipo_comparacion: Tipo de comparación
            nombre_archivo: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Comparativo_{tipo_comparacion}_{timestamp}.xlsx"

        logger.info(f"📊 Generando reporte comparativo: {nombre_archivo}")

        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            workbook = writer.book

            # Hoja de datos actuales
            df_actual.to_excel(writer, sheet_name='Período Actual', index=False, startrow=4)
            ws_actual = writer.sheets['Período Actual']
            self._agregar_encabezado(
                ws_actual,
                titulo="PERÍODO ACTUAL",
                subtitulo=f"Datos de {tipo_comparacion}",
                cedis=self.cedis
            )
            self._formatear_tabla_generica(ws_actual, df_actual, start_row=5)
            self._ajustar_columnas(ws_actual)

            # Hoja de datos anteriores
            df_anterior.to_excel(writer, sheet_name='Período Anterior', index=False, startrow=4)
            ws_anterior = writer.sheets['Período Anterior']
            self._agregar_encabezado(
                ws_anterior,
                titulo="PERÍODO ANTERIOR",
                subtitulo=f"Datos de {tipo_comparacion}",
                cedis=self.cedis
            )
            self._formatear_tabla_generica(ws_anterior, df_anterior, start_row=5)
            self._ajustar_columnas(ws_anterior)

            # Hoja de análisis comparativo
            self._crear_analisis_comparativo(workbook, df_actual, df_anterior, tipo_comparacion)

        logger.info(f"✅ Reporte comparativo generado: {nombre_archivo}")
        return nombre_archivo

    def _crear_analisis_comparativo(self, workbook, df_actual: pd.DataFrame,
                                    df_anterior: pd.DataFrame, tipo: str):
        """Crea hoja de análisis comparativo"""
        ws = workbook.create_sheet('Análisis', 0)

        # Encabezado
        ws.merge_cells('A1:E1')
        ws['A1'] = f'📊 ANÁLISIS COMPARATIVO - {tipo.upper()}'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type='solid')
        ws.row_dimensions[1].height = 35

        ws['A3'] = 'Métrica'
        ws['B3'] = 'Actual'
        ws['C3'] = 'Anterior'
        ws['D3'] = 'Variación'
        ws['E3'] = 'Variación %'

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color=self.COLOR_SUBHEADER,
                                             end_color=self.COLOR_SUBHEADER,
                                             fill_type='solid')

        # Estadísticas básicas
        ws['A4'] = 'Total Registros'
        ws['B4'] = len(df_actual)
        ws['C4'] = len(df_anterior)
        ws['D4'] = len(df_actual) - len(df_anterior)
        if len(df_anterior) > 0:
            ws['E4'] = f"{((len(df_actual) - len(df_anterior)) / len(df_anterior) * 100):.1f}%"

        # Colorear variación
        if ws['D4'].value > 0:
            ws['D4'].font = Font(color=self.COLOR_OK)
        elif ws['D4'].value < 0:
            ws['D4'].font = Font(color=self.COLOR_CRITICO)

    def crear_dashboard_ejecutivo(
        self,
        kpis: Dict[str, Any],
        df_resumen: pd.DataFrame,
        df_errores: pd.DataFrame = None,
        nombre_archivo: str = None
    ) -> str:
        """
        Crea dashboard ejecutivo completo

        Args:
            kpis: Diccionario con KPIs principales
            df_resumen: DataFrame con resumen de datos
            df_errores: DataFrame con errores (opcional)
            nombre_archivo: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Dashboard_Ejecutivo_{self.cedis}_{timestamp}.xlsx"

        logger.info(f"📊 Generando dashboard ejecutivo: {nombre_archivo}")

        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = 'Dashboard'

        # Header grande
        ws.merge_cells('A1:L1')
        ws['A1'] = f'📊 DASHBOARD EJECUTIVO - CEDIS {self.cedis}'
        ws['A1'].font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 50

        ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True)

        # Sección de KPIs
        ws['A4'] = 'INDICADORES CLAVE'
        ws['A4'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws['A4'].fill = PatternFill(start_color=self.COLOR_HEADER,
                                    end_color=self.COLOR_HEADER,
                                    fill_type='solid')
        ws.merge_cells('A4:L4')

        # KPIs en fila horizontal
        col = 1
        for name, data in kpis.items():
            value = data.get('value', 0)
            unit = data.get('unit', '')
            trend = data.get('trend', 'neutral')

            # Nombre
            ws.cell(row=5, column=col, value=name)
            ws.cell(row=5, column=col).font = Font(size=10, bold=True)
            ws.cell(row=5, column=col).alignment = Alignment(horizontal='center')

            # Valor
            value_cell = ws.cell(row=6, column=col, value=f"{value}{unit}")
            if trend == 'positive':
                value_cell.font = Font(size=18, bold=True, color=self.COLOR_OK)
            elif trend == 'negative':
                value_cell.font = Font(size=18, bold=True, color=self.COLOR_CRITICO)
            else:
                value_cell.font = Font(size=18, bold=True, color='003DA5')
            value_cell.alignment = Alignment(horizontal='center')

            col += 2

        # Tabla de resumen
        if not df_resumen.empty:
            ws['A9'] = 'RESUMEN DE DATOS'
            ws['A9'].font = Font(name='Arial', size=12, bold=True)
            ws['A9'].fill = PatternFill(start_color=self.COLOR_SUBHEADER,
                                        end_color=self.COLOR_SUBHEADER,
                                        fill_type='solid')
            ws.merge_cells('A9:F9')

            # Escribir encabezados
            for col_idx, column in enumerate(df_resumen.columns, 1):
                cell = ws.cell(row=10, column=col_idx, value=column)
                cell.font = Font(bold=True)

            # Escribir datos
            for row_idx, row in enumerate(df_resumen.values, 11):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Errores si existen
        if df_errores is not None and not df_errores.empty:
            start_row = 10 + len(df_resumen) + 3
            ws.cell(row=start_row, column=1, value='ERRORES DETECTADOS')
            ws.cell(row=start_row, column=1).font = Font(name='Arial', size=12, bold=True,
                                                         color='FFFFFF')
            ws.cell(row=start_row, column=1).fill = PatternFill(start_color=self.COLOR_CRITICO,
                                                                end_color=self.COLOR_CRITICO,
                                                                fill_type='solid')

            for col_idx, column in enumerate(df_errores.columns, 1):
                ws.cell(row=start_row + 1, column=col_idx, value=column)

            for row_idx, row in enumerate(df_errores.head(5).values, start_row + 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Ajustar columnas
        for col_idx in range(1, 13):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        # Footer corporativo
        footer_row = ws.max_row + 3
        ws.cell(row=footer_row, column=1, value="─" * 80)
        ws.cell(row=footer_row + 1, column=1,
                value="Sistema desarrollado por: Julián Alexander Juárez Alvarado (ADM)")
        ws.cell(row=footer_row + 1, column=1).font = Font(italic=True, size=9)
        ws.cell(row=footer_row + 2, column=1,
                value='"Las máquinas y los sistemas al servicio de los analistas"')
        ws.cell(row=footer_row + 2, column=1).font = Font(italic=True, size=9, color='666666')

        workbook.save(nombre_archivo)
        logger.info(f"✅ Dashboard ejecutivo generado: {nombre_archivo}")
        return nombre_archivo


# ═══════════════════════════════════════════════════════════════
# FUNCIONES DE UTILIDAD
# ═══════════════════════════════════════════════════════════════

def generar_reporte_rapido(df: pd.DataFrame, nombre: str, titulo: str = "REPORTE") -> str:
    """
    Función rápida para generar un reporte simple
    """
    generador = GeneradorReportesExcel()
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"{nombre}_{timestamp}.xlsx"
    
    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Datos', index=False, startrow=2)
        
        worksheet = writer.sheets['Datos']
        worksheet['A1'] = titulo
        worksheet['A1'].font = Font(name='Arial', size=14, bold=True)
        
        generador._ajustar_columnas(worksheet)
    
    return nombre_archivo


# ═══════════════════════════════════════════════════════════════
# EJEMPLO DE USO
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("""
    ═══════════════════════════════════════════════════════════════
    GENERADOR DE REPORTES EXCEL - CHEDRAUI CEDIS
    ═══════════════════════════════════════════════════════════════
    
    Desarrollado con ❤️ por:
    Julián Alexander Juárez Alvarado (ADM)
    
    ═══════════════════════════════════════════════════════════════
    """)
    
    # Ejemplo de uso con datos de prueba
    datos_prueba = pd.DataFrame({
        'OC': ['OC001', 'OC002', 'OC003'],
        'Total_OC': [100, 200, 150],
        'Total_Distro': [100, 180, 150],
        'STATUS': ['OK', 'Distro incompleta', 'OK']
    })
    
    generador = GeneradorReportesExcel(cedis="CANCÚN")
    archivo = generador.crear_reporte_validacion_oc(datos_prueba, "ejemplo_reporte.xlsx")
    
    print(f"\n✅ Reporte de ejemplo generado: {archivo}\n")
