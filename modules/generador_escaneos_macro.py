"""
═══════════════════════════════════════════════════════════════
GENERADOR DE EXCEL CON MACROS PARA ESCANEOS
Sistema de Gestión de Órdenes de Compra - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Este módulo genera archivos Excel con macros habilitadas (.xlsm) para:
- Captura de escaneos de pallets, cartones, SKUs, LPNs, ASNs
- Organización automática de datos escaneados
- Selección de área de trabajo al inicio
- Envío por correo según el área
- Guardado en escritorio con nombre personalizado

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Chedraui Logística Cancún
═══════════════════════════════════════════════════════════════
"""

import os
import shutil
import zipfile
import tempfile
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, List
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
# CÓDIGO VBA PARA MACROS
# ═══════════════════════════════════════════════════════════════

VBA_MODULO_PRINCIPAL = '''
' ═══════════════════════════════════════════════════════════════
' MÓDULO DE ESCANEOS - CEDIS CHEDRAUI
' Sistema de Captura y Envío de Escaneos
' ═══════════════════════════════════════════════════════════════

Option Explicit

' Variables globales
Public AreaSeleccionada As String
Public NombreUsuario As String
Public NombreSesion As String
Public FechaInicio As Date

' ═══════════════════════════════════════════════════════════════
' INICIALIZACIÓN AL ABRIR
' ═══════════════════════════════════════════════════════════════

Sub Auto_Open()
    Call MostrarFormularioInicio
End Sub

Sub Workbook_Open()
    Call MostrarFormularioInicio
End Sub

' ═══════════════════════════════════════════════════════════════
' FORMULARIO DE INICIO
' ═══════════════════════════════════════════════════════════════

Sub MostrarFormularioInicio()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("CONFIG")

    ' Si ya está configurado, no mostrar
    If ws.Range("B2").Value <> "" And ws.Range("B3").Value <> "" Then
        Exit Sub
    End If

    ' Mostrar formulario de configuración
    frmConfigInicio.Show
End Sub

Sub ConfigurarSesion(ByVal Area As String, ByVal Usuario As String, ByVal Sesion As String)
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("CONFIG")

    AreaSeleccionada = Area
    NombreUsuario = Usuario
    NombreSesion = Sesion
    FechaInicio = Now

    ' Guardar en hoja CONFIG
    ws.Range("B2").Value = Area
    ws.Range("B3").Value = Usuario
    ws.Range("B4").Value = Sesion
    ws.Range("B5").Value = FechaInicio

    ' Actualizar encabezados
    Call ActualizarEncabezados

    MsgBox "✅ Sesión configurada correctamente" & vbCrLf & _
           "Área: " & Area & vbCrLf & _
           "Usuario: " & Usuario & vbCrLf & _
           "Sesión: " & Sesion, vbInformation, "CEDIS Chedraui - Escaneos"
End Sub

Sub ActualizarEncabezados()
    Dim ws As Worksheet
    Dim wsConfig As Worksheet
    Set wsConfig = ThisWorkbook.Sheets("CONFIG")

    For Each ws In ThisWorkbook.Worksheets
        If ws.Name <> "CONFIG" And ws.Name <> "INSTRUCCIONES" Then
            ws.Range("A1").Value = "CEDIS CHEDRAUI - " & wsConfig.Range("B2").Value
            ws.Range("A2").Value = "Usuario: " & wsConfig.Range("B3").Value & " | Sesión: " & wsConfig.Range("B4").Value
        End If
    Next ws
End Sub

' ═══════════════════════════════════════════════════════════════
' ORGANIZACIÓN AUTOMÁTICA DE ESCANEOS
' ═══════════════════════════════════════════════════════════════

Sub OrganizarEscaneos()
    Dim ws As Worksheet
    Set ws = ActiveSheet

    If ws.Name = "CONFIG" Or ws.Name = "INSTRUCCIONES" Then
        MsgBox "⚠️ Por favor seleccione una hoja de escaneos", vbExclamation
        Exit Sub
    End If

    Application.ScreenUpdating = False

    ' Eliminar duplicados
    Call EliminarDuplicados(ws)

    ' Ordenar datos
    Call OrdenarDatos(ws)

    ' Aplicar formato
    Call AplicarFormato(ws)

    ' Actualizar contador
    Call ActualizarContador(ws)

    Application.ScreenUpdating = True

    MsgBox "✅ Escaneos organizados correctamente" & vbCrLf & _
           "Total registros: " & ws.Range("E2").Value, vbInformation, "Organización Completa"
End Sub

Sub EliminarDuplicados(ws As Worksheet)
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row

    If lastRow <= 4 Then Exit Sub

    On Error Resume Next
    ws.Range("A5:D" & lastRow).RemoveDuplicates Columns:=1, Header:=xlNo
    On Error GoTo 0
End Sub

Sub OrdenarDatos(ws As Worksheet)
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row

    If lastRow <= 4 Then Exit Sub

    ws.Sort.SortFields.Clear
    ws.Sort.SortFields.Add Key:=ws.Range("A5:A" & lastRow), Order:=xlAscending

    With ws.Sort
        .SetRange ws.Range("A4:D" & lastRow)
        .Header = xlYes
        .Apply
    End With
End Sub

Sub AplicarFormato(ws As Worksheet)
    Dim lastRow As Long
    Dim rng As Range

    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
    If lastRow < 5 Then lastRow = 5

    Set rng = ws.Range("A5:D" & lastRow)

    ' Bordes
    With rng.Borders
        .LineStyle = xlContinuous
        .Weight = xlThin
    End With

    ' Alternar colores
    Dim i As Long
    For i = 5 To lastRow
        If i Mod 2 = 0 Then
            ws.Range("A" & i & ":D" & i).Interior.Color = RGB(248, 249, 250)
        Else
            ws.Range("A" & i & ":D" & i).Interior.Color = RGB(255, 255, 255)
        End If
    Next i
End Sub

Sub ActualizarContador(ws As Worksheet)
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row

    If lastRow < 5 Then
        ws.Range("E2").Value = 0
    Else
        ws.Range("E2").Value = lastRow - 4
    End If

    ws.Range("E3").Value = Now
End Sub

' ═══════════════════════════════════════════════════════════════
' PROCESAMIENTO DE ESCANEO EN TIEMPO REAL
' ═══════════════════════════════════════════════════════════════

Sub ProcesarEscaneo()
    Dim ws As Worksheet
    Dim valorEscaneado As String
    Dim tipoElemento As String
    Dim ultimaFila As Long

    Set ws = ActiveSheet

    If ws.Name = "CONFIG" Or ws.Name = "INSTRUCCIONES" Then
        Exit Sub
    End If

    ' Verificar si hay nuevo valor en celda de entrada
    valorEscaneado = Trim(ws.Range("G5").Value)
    If valorEscaneado = "" Then Exit Sub

    ' Detectar tipo de elemento
    tipoElemento = DetectarTipoElemento(valorEscaneado)

    ' Insertar en la tabla
    ultimaFila = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row + 1
    If ultimaFila < 5 Then ultimaFila = 5

    ws.Cells(ultimaFila, 1).Value = valorEscaneado
    ws.Cells(ultimaFila, 2).Value = tipoElemento
    ws.Cells(ultimaFila, 3).Value = Now
    ws.Cells(ultimaFila, 4).Value = ThisWorkbook.Sheets("CONFIG").Range("B3").Value

    ' Limpiar celda de entrada
    ws.Range("G5").Value = ""

    ' Actualizar contador
    Call ActualizarContador(ws)
End Sub

Function DetectarTipoElemento(valor As String) As String
    valor = UCase(Trim(valor))

    If Left(valor, 3) = "LPN" Then
        DetectarTipoElemento = "LPN"
    ElseIf Left(valor, 3) = "ASN" Then
        DetectarTipoElemento = "ASN"
    ElseIf Left(valor, 2) = "PL" Or Left(valor, 3) = "PAL" Then
        DetectarTipoElemento = "PALLET"
    ElseIf Left(valor, 2) = "CT" Or Left(valor, 3) = "CAR" Or Left(valor, 3) = "CTN" Then
        DetectarTipoElemento = "CARTON"
    ElseIf Len(valor) >= 8 And IsNumeric(valor) Then
        If Len(valor) = 13 Or Len(valor) = 14 Then
            DetectarTipoElemento = "SKU/EAN"
        ElseIf Left(valor, 3) = "750" Or Left(valor, 3) = "811" Or Left(valor, 2) = "40" Then
            DetectarTipoElemento = "OC"
        Else
            DetectarTipoElemento = "CODIGO"
        End If
    ElseIf Left(valor, 1) = "C" And IsNumeric(Mid(valor, 2)) Then
        DetectarTipoElemento = "OC"
    Else
        DetectarTipoElemento = "OTRO"
    End If
End Function

' ═══════════════════════════════════════════════════════════════
' ENVÍO POR CORREO
' ═══════════════════════════════════════════════════════════════

Sub EnviarPorCorreo()
    Dim wsConfig As Worksheet
    Dim ws As Worksheet
    Dim outlookApp As Object
    Dim outlookMail As Object
    Dim archivoTemp As String
    Dim destinatarios As String
    Dim asunto As String
    Dim cuerpo As String
    Dim area As String

    Set wsConfig = ThisWorkbook.Sheets("CONFIG")
    Set ws = ActiveSheet

    If ws.Name = "CONFIG" Or ws.Name = "INSTRUCCIONES" Then
        MsgBox "⚠️ Por favor seleccione una hoja de escaneos para enviar", vbExclamation
        Exit Sub
    End If

    area = wsConfig.Range("B2").Value

    ' Obtener destinatarios según área
    destinatarios = ObtenerDestinatariosArea(area)

    If destinatarios = "" Then
        destinatarios = InputBox("Ingrese el correo del destinatario:", "Enviar Escaneos", "")
        If destinatarios = "" Then Exit Sub
    End If

    ' Guardar copia temporal
    archivoTemp = Environ("TEMP") & "\\Escaneos_" & Format(Now, "yyyymmdd_hhmmss") & ".xlsx"
    ws.Copy
    ActiveWorkbook.SaveAs archivoTemp, FileFormat:=51
    ActiveWorkbook.Close False

    ' Preparar asunto y cuerpo
    asunto = "📊 Escaneos " & area & " - " & wsConfig.Range("B4").Value & " - " & Format(Now, "dd/mm/yyyy")

    cuerpo = "<html><body style='font-family: Calibri, sans-serif;'>" & _
             "<div style='background-color: #E31837; color: white; padding: 15px; text-align: center;'>" & _
             "<h2>🏪 CEDIS CHEDRAUI - ESCANEOS</h2></div>" & _
             "<div style='padding: 20px;'>" & _
             "<p><strong>Área:</strong> " & area & "</p>" & _
             "<p><strong>Usuario:</strong> " & wsConfig.Range("B3").Value & "</p>" & _
             "<p><strong>Sesión:</strong> " & wsConfig.Range("B4").Value & "</p>" & _
             "<p><strong>Total Escaneos:</strong> " & ws.Range("E2").Value & "</p>" & _
             "<p><strong>Fecha Envío:</strong> " & Format(Now, "dd/mm/yyyy hh:mm:ss") & "</p>" & _
             "<hr><p>Se adjunta archivo con los escaneos realizados.</p>" & _
             "</div>" & _
             "<div style='background-color: #f0f0f0; padding: 10px; text-align: center; font-size: 12px;'>" & _
             "Sistema SAC - CEDIS Chedraui Cancún 427</div></body></html>"

    ' Crear correo con Outlook
    On Error Resume Next
    Set outlookApp = GetObject(, "Outlook.Application")
    If outlookApp Is Nothing Then
        Set outlookApp = CreateObject("Outlook.Application")
    End If
    On Error GoTo ErrorHandler

    Set outlookMail = outlookApp.CreateItem(0)

    With outlookMail
        .To = destinatarios
        .Subject = asunto
        .HTMLBody = cuerpo
        .Attachments.Add archivoTemp
        .Display ' Mostrar para revisar antes de enviar
    End With

    MsgBox "✅ Correo preparado correctamente" & vbCrLf & _
           "Revise y envíe manualmente", vbInformation, "Envío de Escaneos"

    Exit Sub

ErrorHandler:
    MsgBox "❌ Error al crear el correo: " & Err.Description & vbCrLf & _
           "Asegúrese de tener Outlook instalado", vbCritical, "Error"
End Sub

Function ObtenerDestinatariosArea(area As String) As String
    Select Case UCase(area)
        Case "RECIBO"
            ObtenerDestinatariosArea = "recibo@chedraui.com.mx"
        Case "SURTIDO"
            ObtenerDestinatariosArea = "surtido@chedraui.com.mx"
        Case "EMBARQUES"
            ObtenerDestinatariosArea = "embarques@chedraui.com.mx"
        Case "INVENTARIOS"
            ObtenerDestinatariosArea = "inventarios@chedraui.com.mx"
        Case "PLANNING"
            ObtenerDestinatariosArea = "planning@chedraui.com.mx"
        Case "SISTEMAS"
            ObtenerDestinatariosArea = "sistemas@chedraui.com.mx"
        Case "CALIDAD"
            ObtenerDestinatariosArea = "calidad@chedraui.com.mx"
        Case Else
            ObtenerDestinatariosArea = ""
    End Select
End Function

' ═══════════════════════════════════════════════════════════════
' GUARDAR EN ESCRITORIO
' ═══════════════════════════════════════════════════════════════

Sub GuardarEnEscritorio()
    Dim wsConfig As Worksheet
    Dim rutaEscritorio As String
    Dim nombreArchivo As String
    Dim rutaCompleta As String

    Set wsConfig = ThisWorkbook.Sheets("CONFIG")

    ' Obtener ruta del escritorio
    rutaEscritorio = Environ("USERPROFILE") & "\\Desktop\\"

    ' Generar nombre de archivo
    nombreArchivo = "Escaneos_" & _
                    wsConfig.Range("B2").Value & "_" & _
                    wsConfig.Range("B4").Value & "_" & _
                    Format(Now, "yyyymmdd_hhmmss") & ".xlsx"

    ' Limpiar caracteres no válidos
    nombreArchivo = LimpiarNombreArchivo(nombreArchivo)

    rutaCompleta = rutaEscritorio & nombreArchivo

    ' Guardar copia
    ThisWorkbook.SaveCopyAs rutaCompleta

    MsgBox "✅ Archivo guardado exitosamente en:" & vbCrLf & vbCrLf & _
           rutaCompleta, vbInformation, "Guardado en Escritorio"
End Sub

Function LimpiarNombreArchivo(nombre As String) As String
    Dim caracteres As String
    Dim i As Integer

    caracteres = "\/:*?""<>|"

    For i = 1 To Len(caracteres)
        nombre = Replace(nombre, Mid(caracteres, i, 1), "_")
    Next i

    LimpiarNombreArchivo = nombre
End Function

' ═══════════════════════════════════════════════════════════════
' LIMPIAR DATOS
' ═══════════════════════════════════════════════════════════════

Sub LimpiarHoja()
    Dim ws As Worksheet
    Dim respuesta As VbMsgBoxResult

    Set ws = ActiveSheet

    If ws.Name = "CONFIG" Or ws.Name = "INSTRUCCIONES" Then
        MsgBox "⚠️ No se puede limpiar esta hoja", vbExclamation
        Exit Sub
    End If

    respuesta = MsgBox("⚠️ ¿Está seguro de eliminar todos los escaneos de esta hoja?" & vbCrLf & _
                       "Esta acción no se puede deshacer.", vbYesNo + vbExclamation, "Confirmar Limpieza")

    If respuesta = vbNo Then Exit Sub

    ' Limpiar datos (mantener encabezados)
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row

    If lastRow >= 5 Then
        ws.Range("A5:D" & lastRow).ClearContents
    End If

    ' Reiniciar contador
    ws.Range("E2").Value = 0
    ws.Range("E3").Value = Now

    MsgBox "✅ Hoja limpiada correctamente", vbInformation, "Limpieza Completa"
End Sub

' ═══════════════════════════════════════════════════════════════
' EXPORTAR RESUMEN
' ═══════════════════════════════════════════════════════════════

Sub ExportarResumen()
    Dim ws As Worksheet
    Dim wsResumen As Worksheet
    Dim fila As Long

    ' Crear hoja de resumen si no existe
    On Error Resume Next
    Set wsResumen = ThisWorkbook.Sheets("RESUMEN")
    On Error GoTo 0

    If wsResumen Is Nothing Then
        Set wsResumen = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
        wsResumen.Name = "RESUMEN"

        ' Encabezados
        wsResumen.Range("A1").Value = "RESUMEN DE ESCANEOS"
        wsResumen.Range("A3").Value = "Hoja"
        wsResumen.Range("B3").Value = "Total"
        wsResumen.Range("C3").Value = "Última Actualización"
    End If

    ' Limpiar datos anteriores
    wsResumen.Range("A4:C100").ClearContents

    fila = 4
    For Each ws In ThisWorkbook.Worksheets
        If ws.Name <> "CONFIG" And ws.Name <> "INSTRUCCIONES" And ws.Name <> "RESUMEN" Then
            wsResumen.Cells(fila, 1).Value = ws.Name
            wsResumen.Cells(fila, 2).Value = ws.Range("E2").Value
            wsResumen.Cells(fila, 3).Value = ws.Range("E3").Value
            fila = fila + 1
        End If
    Next ws

    ' Total general
    wsResumen.Cells(fila + 1, 1).Value = "TOTAL GENERAL"
    wsResumen.Cells(fila + 1, 2).Formula = "=SUM(B4:B" & fila - 1 & ")"

    wsResumen.Activate

    MsgBox "✅ Resumen generado correctamente", vbInformation, "Resumen de Escaneos"
End Sub

' ═══════════════════════════════════════════════════════════════
' REINICIAR CONFIGURACIÓN
' ═══════════════════════════════════════════════════════════════

Sub ReiniciarConfiguracion()
    Dim respuesta As VbMsgBoxResult
    Dim ws As Worksheet

    respuesta = MsgBox("⚠️ ¿Está seguro de reiniciar la configuración?" & vbCrLf & _
                       "Deberá configurar nuevamente área y usuario.", _
                       vbYesNo + vbExclamation, "Confirmar Reinicio")

    If respuesta = vbNo Then Exit Sub

    Set ws = ThisWorkbook.Sheets("CONFIG")
    ws.Range("B2:B5").ClearContents

    AreaSeleccionada = ""
    NombreUsuario = ""
    NombreSesion = ""

    MsgBox "✅ Configuración reiniciada" & vbCrLf & _
           "Por favor cierre y vuelva a abrir el archivo", vbInformation, "Reinicio Completo"
End Sub
'''

VBA_FORMULARIO_INICIO = '''
' ═══════════════════════════════════════════════════════════════
' FORMULARIO DE CONFIGURACIÓN INICIAL
' ═══════════════════════════════════════════════════════════════

Private Sub UserForm_Initialize()
    ' Cargar áreas en el combobox
    With cboArea
        .AddItem "RECIBO"
        .AddItem "SURTIDO"
        .AddItem "EMBARQUES"
        .AddItem "INVENTARIOS"
        .AddItem "PLANNING"
        .AddItem "SISTEMAS"
        .AddItem "CALIDAD"
        .AddItem "DEVOLUCIONES"
        .AddItem "MAQUILA"
        .AddItem "OTRO"
    End With

    ' Valores por defecto
    txtUsuario.Value = Environ("USERNAME")
    txtSesion.Value = "Sesion_" & Format(Now, "yyyymmdd")

    Me.Caption = "🏪 CEDIS Chedraui - Configuración de Escaneos"
End Sub

Private Sub btnAceptar_Click()
    ' Validar campos
    If cboArea.Value = "" Then
        MsgBox "⚠️ Por favor seleccione un área", vbExclamation
        cboArea.SetFocus
        Exit Sub
    End If

    If Trim(txtUsuario.Value) = "" Then
        MsgBox "⚠️ Por favor ingrese su nombre", vbExclamation
        txtUsuario.SetFocus
        Exit Sub
    End If

    If Trim(txtSesion.Value) = "" Then
        MsgBox "⚠️ Por favor ingrese un nombre para la sesión", vbExclamation
        txtSesion.SetFocus
        Exit Sub
    End If

    ' Configurar sesión
    Call ConfigurarSesion(cboArea.Value, txtUsuario.Value, txtSesion.Value)

    Unload Me
End Sub

Private Sub btnCancelar_Click()
    Unload Me
End Sub
'''


class GeneradorExcelMacros:
    """
    Generador de archivos Excel con macros habilitadas para escaneos

    Crea archivos .xlsm con:
    - Formulario de inicio para selección de área
    - Hojas para diferentes tipos de escaneos
    - Macros VBA para organización automática
    - Envío por correo según área
    - Guardado en escritorio
    """

    # Áreas disponibles
    AREAS = [
        "RECIBO",
        "SURTIDO",
        "EMBARQUES",
        "INVENTARIOS",
        "PLANNING",
        "SISTEMAS",
        "CALIDAD",
        "DEVOLUCIONES",
        "MAQUILA"
    ]

    # Tipos de escaneo
    TIPOS_ESCANEO = [
        ("PALLETS", "Escaneo de tarimas y pallets"),
        ("CARTONES", "Escaneo de cartones y cajas"),
        ("SKUs", "Escaneo de códigos de producto"),
        ("LPNs", "Escaneo de etiquetas LPN"),
        ("ASNs", "Escaneo de avisos de envío"),
        ("GENERAL", "Escaneos varios")
    ]

    # Colores corporativos
    COLOR_ROJO_CHEDRAUI = "E31837"
    COLOR_BLANCO = "FFFFFF"
    COLOR_GRIS_CLARO = "F8F9FA"
    COLOR_VERDE = "28A745"
    COLOR_AZUL = "007BFF"

    def __init__(self, cedis: str = "CANCÚN 427"):
        self.cedis = cedis
        self.output_dir = Path("output/escaneos")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generar_archivo_escaneos(
        self,
        nombre_archivo: str = None,
        incluir_instrucciones: bool = True,
        areas_personalizadas: List[str] = None,
        tipos_escaneo_personalizados: List[tuple] = None
    ) -> str:
        """
        Genera un archivo Excel con macros para escaneos

        Args:
            nombre_archivo: Nombre del archivo (sin extensión)
            incluir_instrucciones: Si incluir hoja de instrucciones
            areas_personalizadas: Lista personalizada de áreas
            tipos_escaneo_personalizados: Lista de tuplas (nombre, descripción)

        Returns:
            str: Ruta del archivo generado
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("❌ openpyxl no está instalado")
            raise ImportError("Requiere openpyxl: pip install openpyxl")

        if nombre_archivo is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_archivo = f"Escaneos_CEDIS_{timestamp}"

        # Usar áreas y tipos personalizados o por defecto
        areas = areas_personalizadas or self.AREAS
        tipos = tipos_escaneo_personalizados or self.TIPOS_ESCANEO

        logger.info(f"📊 Generando archivo de escaneos: {nombre_archivo}")

        # Crear workbook
        wb = Workbook()

        # Eliminar hoja por defecto
        ws_default = wb.active
        wb.remove(ws_default)

        # Crear hoja de configuración
        self._crear_hoja_config(wb, areas)

        # Crear hoja de instrucciones
        if incluir_instrucciones:
            self._crear_hoja_instrucciones(wb)

        # Crear hojas para cada tipo de escaneo
        for tipo_nombre, tipo_descripcion in tipos:
            self._crear_hoja_escaneo(wb, tipo_nombre, tipo_descripcion)

        # Guardar archivo
        ruta_archivo = self.output_dir / f"{nombre_archivo}.xlsx"
        wb.save(str(ruta_archivo))

        logger.info(f"✅ Archivo base generado: {ruta_archivo}")

        # Crear archivo con instrucciones VBA
        ruta_vba = self._guardar_codigo_vba(nombre_archivo)

        # Mensaje informativo
        self._mostrar_instrucciones_vba(ruta_archivo, ruta_vba)

        return str(ruta_archivo)

    def _crear_hoja_config(self, wb, areas: List[str]):
        """Crea la hoja de configuración"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation

        ws = wb.create_sheet("CONFIG", 0)

        # Estilos
        header_fill = PatternFill(start_color=self.COLOR_ROJO_CHEDRAUI,
                                  end_color=self.COLOR_ROJO_CHEDRAUI,
                                  fill_type="solid")
        header_font = Font(bold=True, color=self.COLOR_BLANCO, size=14)
        label_font = Font(bold=True, size=11)

        # Encabezado
        ws.merge_cells('A1:D1')
        ws['A1'] = "🏪 CONFIGURACIÓN DE SESIÓN - CEDIS CHEDRAUI"
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Campos de configuración
        campos = [
            ("A2", "Área de Trabajo:", "B2", ""),
            ("A3", "Nombre de Usuario:", "B3", ""),
            ("A4", "Nombre de Sesión:", "B4", ""),
            ("A5", "Fecha de Inicio:", "B5", ""),
        ]

        for label_cell, label_text, value_cell, default_value in campos:
            ws[label_cell] = label_text
            ws[label_cell].font = label_font
            ws[value_cell] = default_value
            ws[value_cell].alignment = Alignment(horizontal='left')

        # Validación de datos para área
        areas_str = ",".join(areas)
        dv = DataValidation(type="list", formula1=f'"{areas_str}"', allow_blank=False)
        dv.error = "Por favor seleccione un área válida"
        dv.errorTitle = "Área inválida"
        dv.prompt = "Seleccione su área de trabajo"
        dv.promptTitle = "Área"
        ws.add_data_validation(dv)
        dv.add(ws['B2'])

        # Ancho de columnas
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

        # Sección de estadísticas
        ws['A7'] = "📊 ESTADÍSTICAS DE LA SESIÓN"
        ws['A7'].font = Font(bold=True, size=12)
        ws.merge_cells('A7:D7')

        ws['A8'] = "Total Escaneos:"
        ws['B8'] = "=SUMPRODUCT((PALLETS!E2)+(CARTONES!E2)+(SKUs!E2)+(LPNs!E2)+(ASNs!E2)+(GENERAL!E2))"

        ws['A9'] = "Última Actividad:"
        ws['B9'] = "=MAX(PALLETS!E3,CARTONES!E3,SKUs!E3,LPNs!E3,ASNs!E3,GENERAL!E3)"
        ws['B9'].number_format = 'DD/MM/YYYY HH:MM:SS'

        # Sección de destinatarios
        ws['A11'] = "📧 CORREOS POR ÁREA"
        ws['A11'].font = Font(bold=True, size=12)
        ws.merge_cells('A11:D11')

        destinatarios = [
            ("RECIBO", "recibo@chedraui.com.mx"),
            ("SURTIDO", "surtido@chedraui.com.mx"),
            ("EMBARQUES", "embarques@chedraui.com.mx"),
            ("INVENTARIOS", "inventarios@chedraui.com.mx"),
            ("PLANNING", "planning@chedraui.com.mx"),
            ("SISTEMAS", "sistemas@chedraui.com.mx"),
        ]

        for i, (area, correo) in enumerate(destinatarios, start=12):
            ws[f'A{i}'] = area
            ws[f'B{i}'] = correo

    def _crear_hoja_instrucciones(self, wb):
        """Crea la hoja de instrucciones"""
        from openpyxl.styles import Font, PatternFill, Alignment

        ws = wb.create_sheet("INSTRUCCIONES", 1)

        header_fill = PatternFill(start_color=self.COLOR_ROJO_CHEDRAUI,
                                  end_color=self.COLOR_ROJO_CHEDRAUI,
                                  fill_type="solid")

        # Encabezado
        ws.merge_cells('A1:F1')
        ws['A1'] = "📋 INSTRUCCIONES DE USO - SISTEMA DE ESCANEOS"
        ws['A1'].fill = header_fill
        ws['A1'].font = Font(bold=True, color=self.COLOR_BLANCO, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        instrucciones = [
            "",
            "🚀 INICIO RÁPIDO:",
            "1. Al abrir el archivo, configure su área de trabajo en la hoja CONFIG",
            "2. Ingrese su nombre de usuario y un nombre para la sesión",
            "3. Seleccione la hoja correspondiente al tipo de elemento a escanear",
            "4. Pegue o escriba los códigos escaneados en la columna A",
            "",
            "📊 HOJAS DISPONIBLES:",
            "• PALLETS - Para escaneo de tarimas y pallets",
            "• CARTONES - Para escaneo de cajas y cartones",
            "• SKUs - Para escaneo de códigos de producto",
            "• LPNs - Para escaneo de etiquetas LPN",
            "• ASNs - Para escaneo de avisos de envío",
            "• GENERAL - Para escaneos varios",
            "",
            "🔧 FUNCIONES DISPONIBLES (requieren macros VBA):",
            "• Organizar Escaneos - Ordena y elimina duplicados",
            "• Enviar por Correo - Envía los escaneos al área correspondiente",
            "• Guardar en Escritorio - Guarda una copia en el escritorio",
            "• Limpiar Hoja - Elimina todos los escaneos de la hoja activa",
            "• Exportar Resumen - Genera un resumen de todas las hojas",
            "",
            "⚠️ IMPORTANTE:",
            "• Para habilitar las macros, guarde el archivo como .xlsm",
            "• Debe importar el código VBA desde el archivo proporcionado",
            "• Asegúrese de tener habilitadas las macros en Excel",
            "",
            f"📍 CEDIS: {self.cedis}",
            f"📅 Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            "",
            "Desarrollado por: Sistemas CEDIS Chedraui",
        ]

        for i, texto in enumerate(instrucciones, start=2):
            ws[f'A{i}'] = texto
            if texto.startswith("🚀") or texto.startswith("📊") or texto.startswith("🔧") or texto.startswith("⚠️"):
                ws[f'A{i}'].font = Font(bold=True, size=11)

        ws.column_dimensions['A'].width = 80

    def _crear_hoja_escaneo(self, wb, nombre: str, descripcion: str):
        """Crea una hoja para un tipo específico de escaneo"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ws = wb.create_sheet(nombre)

        # Estilos
        header_fill = PatternFill(start_color=self.COLOR_ROJO_CHEDRAUI,
                                  end_color=self.COLOR_ROJO_CHEDRAUI,
                                  fill_type="solid")
        col_header_fill = PatternFill(start_color="4472C4",
                                      end_color="4472C4",
                                      fill_type="solid")
        white_font = Font(bold=True, color=self.COLOR_BLANCO)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezado principal
        ws.merge_cells('A1:D1')
        ws['A1'] = f"🏪 CEDIS CHEDRAUI - {descripcion.upper()}"
        ws['A1'].fill = header_fill
        ws['A1'].font = Font(bold=True, color=self.COLOR_BLANCO, size=12)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Información de sesión
        ws.merge_cells('A2:D2')
        ws['A2'] = "Usuario: [Configurar en hoja CONFIG] | Sesión: [Configurar en hoja CONFIG]"
        ws['A2'].alignment = Alignment(horizontal='center')

        # Área de estadísticas (columna E)
        ws['E1'] = "📊 ESTADÍSTICAS"
        ws['E1'].font = Font(bold=True, size=10)
        ws['E2'] = 0  # Total escaneos
        ws['E3'] = ""  # Última actualización
        ws['F1'] = "Total:"
        ws['F2'] = "Actualizado:"

        # Encabezados de columnas
        headers = ["CÓDIGO ESCANEADO", "TIPO", "FECHA/HORA", "USUARIO"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = col_header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Área de entrada rápida
        ws['G1'] = "📥 ENTRADA RÁPIDA"
        ws['G1'].font = Font(bold=True, size=10)
        ws['G2'] = "Pegue aquí:"
        ws['G3'] = "(El sistema procesará automáticamente)"
        ws['G5'] = ""  # Celda de entrada
        ws['G5'].fill = PatternFill(start_color="FFFFCC",
                                    end_color="FFFFCC",
                                    fill_type="solid")

        # Ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25

        # Congelar paneles
        ws.freeze_panes = 'A5'

    def _guardar_codigo_vba(self, nombre_base: str) -> str:
        """Guarda el código VBA en un archivo de texto"""

        ruta_vba = self.output_dir / f"{nombre_base}_MACROS_VBA.txt"

        contenido_vba = f'''
' ═══════════════════════════════════════════════════════════════
' CÓDIGO VBA PARA EXCEL CON MACROS DE ESCANEOS
' CEDIS CHEDRAUI - Sistema SAC
' ═══════════════════════════════════════════════════════════════
'
' INSTRUCCIONES DE INSTALACIÓN:
' 1. Abra el archivo Excel generado (.xlsx)
' 2. Guárdelo como archivo habilitado para macros (.xlsm)
' 3. Presione Alt+F11 para abrir el editor de VBA
' 4. Inserte un nuevo módulo (Insertar > Módulo)
' 5. Copie y pegue el código del MÓDULO PRINCIPAL
' 6. Cree un UserForm llamado "frmConfigInicio" con:
'    - ComboBox: cboArea
'    - TextBox: txtUsuario
'    - TextBox: txtSesion
'    - CommandButton: btnAceptar
'    - CommandButton: btnCancelar
' 7. Copie el código del formulario en el UserForm
' 8. Guarde y cierre el editor
' 9. Cierre y vuelva a abrir el archivo para probar
'
' Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
' ═══════════════════════════════════════════════════════════════

' ═══════════════════════════════════════════════════════════════
' MÓDULO PRINCIPAL (Insertar en un Módulo nuevo)
' ═══════════════════════════════════════════════════════════════

{VBA_MODULO_PRINCIPAL}

' ═══════════════════════════════════════════════════════════════
' CÓDIGO DEL USERFORM (frmConfigInicio)
' ═══════════════════════════════════════════════════════════════

{VBA_FORMULARIO_INICIO}

' ═══════════════════════════════════════════════════════════════
' FIN DEL CÓDIGO VBA
' ═══════════════════════════════════════════════════════════════
'''

        with open(ruta_vba, 'w', encoding='utf-8') as f:
            f.write(contenido_vba)

        logger.info(f"✅ Código VBA guardado: {ruta_vba}")
        return str(ruta_vba)

    def _mostrar_instrucciones_vba(self, ruta_excel: str, ruta_vba: str):
        """Muestra instrucciones para agregar macros"""

        print(f"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                    📊 ARCHIVO DE ESCANEOS GENERADO                             ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║                                                                               ║
║  📁 Archivo Excel: {Path(ruta_excel).name:<54} ║
║  📝 Código VBA:    {Path(ruta_vba).name:<54} ║
║                                                                               ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  📋 INSTRUCCIONES PARA HABILITAR MACROS:                                      ║
║                                                                               ║
║  1. Abra el archivo .xlsx generado                                            ║
║  2. Guarde como "Excel habilitado para macros (*.xlsm)"                       ║
║  3. Presione Alt+F11 para abrir el editor de VBA                              ║
║  4. Vaya a Insertar → Módulo                                                  ║
║  5. Copie el código del archivo _MACROS_VBA.txt                               ║
║  6. Cree el UserForm siguiendo las instrucciones del archivo                  ║
║  7. Guarde, cierre y vuelva a abrir                                           ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
""")


# ═══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL PARA USO DIRECTO
# ═══════════════════════════════════════════════════════════════

def crear_excel_escaneos(
    nombre: str = None,
    cedis: str = "CANCÚN 427",
    areas: List[str] = None,
    tipos: List[tuple] = None
) -> str:
    """
    Función de conveniencia para crear archivo de escaneos

    Args:
        nombre: Nombre del archivo (opcional)
        cedis: Nombre del CEDIS
        areas: Lista de áreas personalizadas
        tipos: Lista de tuplas (nombre, descripcion) para tipos de escaneo

    Returns:
        str: Ruta del archivo generado
    """
    generador = GeneradorExcelMacros(cedis=cedis)
    return generador.generar_archivo_escaneos(
        nombre_archivo=nombre,
        areas_personalizadas=areas,
        tipos_escaneo_personalizados=tipos
    )


if __name__ == "__main__":
    # Ejemplo de uso
    print("═══════════════════════════════════════════════════════════════")
    print("    GENERADOR DE EXCEL CON MACROS PARA ESCANEOS")
    print("    CEDIS Chedraui - Sistema SAC")
    print("═══════════════════════════════════════════════════════════════")
    print()

    ruta = crear_excel_escaneos()
    print(f"\n✅ Archivo generado exitosamente: {ruta}")
