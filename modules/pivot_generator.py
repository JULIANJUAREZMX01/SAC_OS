"""
═══════════════════════════════════════════════════════════════
GENERADOR DE TABLAS DINÁMICAS (PIVOT)
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Módulo para generar tablas dinámicas y análisis de datos.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pandas as pd
import numpy as np
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from typing import Optional, Dict, List, Any, Union, Callable
import logging

from .excel_styles import ChedrauiStyles, ChedrauiColors

logger = logging.getLogger(__name__)


class PivotConfig:
    """Configuración para tablas pivot"""

    def __init__(
        self,
        aggfunc: Union[str, Callable] = 'sum',
        fill_value: Any = 0,
        margins: bool = False,
        margins_name: str = 'Total',
        sort_values: bool = True,
        sort_ascending: bool = False,
        top_n: int = None
    ):
        """
        Inicializa configuración de pivot

        Args:
            aggfunc: Función de agregación ('sum', 'mean', 'count', 'min', 'max')
            fill_value: Valor para celdas vacías
            margins: Incluir totales
            margins_name: Nombre para fila/columna de totales
            sort_values: Ordenar resultados
            sort_ascending: Orden ascendente/descendente
            top_n: Limitar a los primeros N resultados
        """
        self.aggfunc = aggfunc
        self.fill_value = fill_value
        self.margins = margins
        self.margins_name = margins_name
        self.sort_values = sort_values
        self.sort_ascending = sort_ascending
        self.top_n = top_n


class PivotGenerator:
    """
    Generador de tablas dinámicas para reportes Excel

    Funcionalidades:
    - Crear pivots desde DataFrames
    - Múltiples funciones de agregación
    - Exportar a Excel con formato
    - Generar gráficos desde pivots
    """

    @classmethod
    def create_pivot_table(
        cls,
        df: pd.DataFrame,
        rows: Union[str, List[str]],
        cols: Union[str, List[str]] = None,
        values: Union[str, List[str]] = None,
        config: PivotConfig = None
    ) -> pd.DataFrame:
        """
        Crea tabla pivot desde DataFrame

        Args:
            df: DataFrame origen
            rows: Columna(s) para filas
            cols: Columna(s) para columnas (opcional)
            values: Columna(s) de valores a agregar
            config: Configuración del pivot

        Returns:
            DataFrame: Tabla pivot resultante
        """
        config = config or PivotConfig()

        # Normalizar a listas
        if isinstance(rows, str):
            rows = [rows]
        if isinstance(cols, str) and cols:
            cols = [cols]
        if isinstance(values, str) and values:
            values = [values]

        try:
            pivot = pd.pivot_table(
                df,
                index=rows,
                columns=cols,
                values=values,
                aggfunc=config.aggfunc,
                fill_value=config.fill_value,
                margins=config.margins,
                margins_name=config.margins_name
            )

            # Aplanar columnas si es MultiIndex
            if isinstance(pivot.columns, pd.MultiIndex):
                pivot.columns = ['_'.join(map(str, col)).strip('_') for col in pivot.columns.values]

            # Ordenar si se solicita
            if config.sort_values and values:
                sort_col = values[0] if isinstance(values, list) else values
                if sort_col in pivot.columns:
                    pivot = pivot.sort_values(sort_col, ascending=config.sort_ascending)
                elif len(pivot.columns) > 0:
                    pivot = pivot.sort_values(pivot.columns[0], ascending=config.sort_ascending)

            # Limitar a top N
            if config.top_n and len(pivot) > config.top_n:
                pivot = pivot.head(config.top_n)

            logger.debug(f"📊 Pivot creado: {len(pivot)} filas x {len(pivot.columns)} columnas")
            return pivot.reset_index()

        except Exception as e:
            logger.error(f"❌ Error creando pivot: {e}")
            return pd.DataFrame()

    @classmethod
    def create_crosstab(
        cls,
        df: pd.DataFrame,
        row_col: str,
        col_col: str,
        values_col: str = None,
        aggfunc: str = 'count',
        normalize: Union[bool, str] = False
    ) -> pd.DataFrame:
        """
        Crea tabla de contingencia (crosstab)

        Args:
            df: DataFrame origen
            row_col: Columna para filas
            col_col: Columna para columnas
            values_col: Columna de valores (opcional)
            aggfunc: Función de agregación
            normalize: Normalizar ('all', 'index', 'columns', False)

        Returns:
            DataFrame: Tabla de contingencia
        """
        try:
            if values_col:
                crosstab = pd.crosstab(
                    df[row_col],
                    df[col_col],
                    values=df[values_col],
                    aggfunc=aggfunc,
                    normalize=normalize,
                    margins=True
                )
            else:
                crosstab = pd.crosstab(
                    df[row_col],
                    df[col_col],
                    normalize=normalize,
                    margins=True
                )

            logger.debug(f"📊 Crosstab creado: {crosstab.shape}")
            return crosstab.reset_index()

        except Exception as e:
            logger.error(f"❌ Error creando crosstab: {e}")
            return pd.DataFrame()

    @classmethod
    def create_summary_stats(
        cls,
        df: pd.DataFrame,
        group_by: Union[str, List[str]],
        numeric_cols: List[str] = None,
        stats: List[str] = None
    ) -> pd.DataFrame:
        """
        Crea resumen estadístico agrupado

        Args:
            df: DataFrame origen
            group_by: Columna(s) de agrupación
            numeric_cols: Columnas numéricas a analizar
            stats: Estadísticas a calcular ('sum', 'mean', 'count', 'std', 'min', 'max')

        Returns:
            DataFrame: Resumen estadístico
        """
        stats = stats or ['count', 'sum', 'mean', 'min', 'max']

        if isinstance(group_by, str):
            group_by = [group_by]

        if numeric_cols is None:
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()

        try:
            agg_dict = {col: stats for col in numeric_cols}
            summary = df.groupby(group_by).agg(agg_dict)

            # Aplanar columnas
            summary.columns = ['_'.join(col).strip() for col in summary.columns.values]
            summary = summary.reset_index()

            logger.debug(f"📊 Resumen estadístico creado: {summary.shape}")
            return summary

        except Exception as e:
            logger.error(f"❌ Error creando resumen estadístico: {e}")
            return pd.DataFrame()

    @classmethod
    def add_pivot_to_excel(
        cls,
        ws: Worksheet,
        pivot_df: pd.DataFrame,
        position: tuple = (1, 1),
        title: str = None,
        apply_formatting: bool = True
    ) -> tuple:
        """
        Agrega tabla pivot a una hoja Excel

        Args:
            ws: Hoja de trabajo
            pivot_df: DataFrame pivot a agregar
            position: (fila, columna) inicial
            title: Título de la tabla (opcional)
            apply_formatting: Aplicar formato corporativo

        Returns:
            tuple: (fila_final, columna_final)
        """
        start_row, start_col = position

        # Agregar título si existe
        if title:
            title_cell = ws.cell(row=start_row, column=start_col, value=title)
            title_cell.font = Font(name='Arial', size=12, bold=True)
            title_cell.fill = PatternFill(
                start_color=ChedrauiColors.SUBHEADER,
                end_color=ChedrauiColors.SUBHEADER,
                fill_type='solid'
            )
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=start_row,
                end_column=start_col + len(pivot_df.columns) - 1
            )
            start_row += 1

        # Escribir encabezados
        for col_idx, column in enumerate(pivot_df.columns, start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=column)
            if apply_formatting:
                cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                cell.fill = PatternFill(
                    start_color=ChedrauiColors.HEADER_RED,
                    end_color=ChedrauiColors.HEADER_RED,
                    fill_type='solid'
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = ChedrauiStyles.THIN_BORDER

        # Escribir datos
        for row_idx, row_data in enumerate(pivot_df.values, start_row + 1):
            for col_idx, value in enumerate(row_data, start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if apply_formatting:
                    cell.border = ChedrauiStyles.THIN_BORDER
                    cell.alignment = Alignment(vertical='center')

                    # Formato numérico
                    if isinstance(value, (int, float)):
                        if isinstance(value, float):
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                    # Filas alternas
                    if (row_idx - start_row) % 2 == 0:
                        cell.fill = PatternFill(
                            start_color=ChedrauiColors.ALT_ROW,
                            end_color=ChedrauiColors.ALT_ROW,
                            fill_type='solid'
                        )

        end_row = start_row + len(pivot_df)
        end_col = start_col + len(pivot_df.columns) - 1

        # Ajustar anchos de columna
        for col_idx in range(start_col, end_col + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in range(start_row, end_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

        logger.debug(f"📊 Pivot agregado a Excel en ({start_row}, {start_col})")
        return end_row, end_col

    @classmethod
    def create_pivot_chart(
        cls,
        ws: Worksheet,
        pivot_df: pd.DataFrame,
        chart_type: str = 'bar',
        position: str = "E1",
        title: str = None,
        width: int = 12,
        height: int = 8
    ) -> None:
        """
        Crea gráfico a partir de tabla pivot

        Args:
            ws: Hoja de trabajo (debe tener los datos del pivot)
            pivot_df: DataFrame pivot para referencia
            chart_type: 'bar', 'pie', 'line'
            position: Posición del gráfico
            title: Título del gráfico
            width: Ancho del gráfico
            height: Alto del gráfico
        """
        if pivot_df.empty:
            logger.warning("⚠️ Pivot vacío, no se puede crear gráfico")
            return

        # Encontrar datos en la hoja
        data_start_row = 1
        data_end_row = len(pivot_df) + 1
        data_end_col = len(pivot_df.columns)

        # Referencias
        categories = Reference(ws, min_col=1, min_row=2, max_row=data_end_row)
        data = Reference(ws, min_col=2, max_col=data_end_col,
                        min_row=1, max_row=data_end_row)

        # Crear gráfico según tipo
        if chart_type == 'bar':
            chart = BarChart()
            chart.type = 'col'
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            chart = BarChart()
            chart.type = 'col'

        chart.title = title or "Gráfico Pivot"
        chart.width = width
        chart.height = height

        chart.add_data(data, titles_from_data=True)
        if chart_type != 'pie':
            chart.set_categories(categories)

        ws.add_chart(chart, position)
        logger.debug(f"📊 Gráfico pivot creado en {position}")

    @classmethod
    def create_ranking(
        cls,
        df: pd.DataFrame,
        rank_col: str,
        value_col: str,
        top_n: int = 10,
        ascending: bool = False
    ) -> pd.DataFrame:
        """
        Crea ranking de elementos

        Args:
            df: DataFrame origen
            rank_col: Columna a rankear
            value_col: Columna de valores
            top_n: Número de elementos en ranking
            ascending: Orden ascendente/descendente

        Returns:
            DataFrame: Ranking con posición
        """
        try:
            ranking = df.groupby(rank_col)[value_col].sum().reset_index()
            ranking = ranking.sort_values(value_col, ascending=ascending)
            ranking = ranking.head(top_n)
            ranking['Posición'] = range(1, len(ranking) + 1)
            ranking = ranking[['Posición', rank_col, value_col]]

            logger.debug(f"📊 Ranking creado: Top {top_n}")
            return ranking

        except Exception as e:
            logger.error(f"❌ Error creando ranking: {e}")
            return pd.DataFrame()

    @classmethod
    def create_time_series_pivot(
        cls,
        df: pd.DataFrame,
        date_col: str,
        value_col: str,
        freq: str = 'D',
        aggfunc: str = 'sum'
    ) -> pd.DataFrame:
        """
        Crea pivot de series temporales

        Args:
            df: DataFrame origen
            date_col: Columna de fecha
            value_col: Columna de valores
            freq: Frecuencia ('D'=día, 'W'=semana, 'M'=mes, 'Q'=trimestre)
            aggfunc: Función de agregación

        Returns:
            DataFrame: Serie temporal agregada
        """
        try:
            df_copy = df.copy()
            df_copy[date_col] = pd.to_datetime(df_copy[date_col])

            if freq == 'D':
                df_copy['Período'] = df_copy[date_col].dt.date
            elif freq == 'W':
                df_copy['Período'] = df_copy[date_col].dt.to_period('W').astype(str)
            elif freq == 'M':
                df_copy['Período'] = df_copy[date_col].dt.to_period('M').astype(str)
            elif freq == 'Q':
                df_copy['Período'] = df_copy[date_col].dt.to_period('Q').astype(str)
            else:
                df_copy['Período'] = df_copy[date_col].dt.date

            time_series = df_copy.groupby('Período')[value_col].agg(aggfunc).reset_index()
            time_series.columns = ['Período', value_col]

            logger.debug(f"📊 Serie temporal creada: {len(time_series)} períodos")
            return time_series

        except Exception as e:
            logger.error(f"❌ Error creando serie temporal: {e}")
            return pd.DataFrame()


# ═══════════════════════════════════════════════════════════════
# FUNCIONES DE UTILIDAD
# ═══════════════════════════════════════════════════════════════

def quick_pivot(
    df: pd.DataFrame,
    group_by: str,
    value_col: str,
    aggfunc: str = 'sum',
    top_n: int = None
) -> pd.DataFrame:
    """
    Función rápida para crear pivots simples

    Args:
        df: DataFrame origen
        group_by: Columna de agrupación
        value_col: Columna de valores
        aggfunc: Función de agregación
        top_n: Limitar resultados

    Returns:
        DataFrame: Pivot simple
    """
    config = PivotConfig(aggfunc=aggfunc, top_n=top_n)
    return PivotGenerator.create_pivot_table(
        df, rows=group_by, values=value_col, config=config
    )


# ═══════════════════════════════════════════════════════════════
# EXPORTACIÓN
# ═══════════════════════════════════════════════════════════════

__all__ = [
    'PivotConfig',
    'PivotGenerator',
    'quick_pivot',
]


if __name__ == "__main__":
    print("""
    ═══════════════════════════════════════════════════════════════
    GENERADOR DE TABLAS DINÁMICAS - CHEDRAUI
    ═══════════════════════════════════════════════════════════════

    Funciones disponibles:
    - create_pivot_table(): Tablas pivot estándar
    - create_crosstab(): Tablas de contingencia
    - create_summary_stats(): Resúmenes estadísticos
    - create_ranking(): Rankings top N
    - create_time_series_pivot(): Series temporales
    - add_pivot_to_excel(): Exportar a Excel
    - create_pivot_chart(): Gráficos desde pivots

    Ejemplo:
        pivot = PivotGenerator.create_pivot_table(
            df,
            rows='TIENDA',
            values='CANTIDAD',
            config=PivotConfig(aggfunc='sum', margins=True)
        )

    ═══════════════════════════════════════════════════════════════
    """)
