"""
═══════════════════════════════════════════════════════════════
TEMPLATE BASE PARA REPORTES EXCEL
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Clase base para todos los templates de reportes Excel.
Proporciona funcionalidad común para encabezados, estilos y tablas.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, FormulaRule,
    CellIsRule, Rule
)
from openpyxl.drawing.image import Image
from datetime import datetime
from typing import Optional, Dict, List, Tuple, Any
from pathlib import Path
import logging

from ..excel_styles import ChedrauiStyles, ChedrauiColors, apply_style_dict

logger = logging.getLogger(__name__)


class ExcelTemplate:
    """
    Clase base para templates de reportes Excel corporativos Chedraui

    Proporciona funcionalidad común:
    - Encabezados corporativos
    - Pies de página
    - Estilos corporativos
    - Tablas formateadas
    - Formato condicional
    - Inserción de logo
    """

    # Configuración por defecto
    DEFAULT_ROW_HEIGHT = 20
    DEFAULT_HEADER_HEIGHT = 35
    DEFAULT_TITLE_HEIGHT = 40
    LOGO_PATH = Path(__file__).parent.parent.parent / 'assets' / 'logo_chedraui.png'

    def __init__(self, cedis: str = "CANCÚN", author: str = "Sistema SAC"):
        """
        Inicializa el template

        Args:
            cedis: Nombre del CEDIS
            author: Autor del reporte
        """
        self.cedis = cedis
        self.author = author
        self.workbook: Optional[Workbook] = None
        self.created_at = datetime.now()

    def create_workbook(self) -> Workbook:
        """
        Crea un nuevo workbook con estilos registrados

        Returns:
            Workbook: Nuevo workbook de Excel
        """
        self.workbook = Workbook()
        ChedrauiStyles.register_styles(self.workbook)
        return self.workbook

    def apply_corporate_header(
        self,
        ws: Worksheet,
        title: str,
        subtitle: str = "",
        merge_cols: int = 6
    ) -> int:
        """
        Aplica encabezado corporativo Chedraui

        Args:
            ws: Hoja de trabajo
            title: Título principal
            subtitle: Subtítulo (opcional)
            merge_cols: Número de columnas a combinar

        Returns:
            int: Fila donde terminó el encabezado
        """
        # Fila 1: Logo/Chedraui
        ws.merge_cells(f'A1:{get_column_letter(merge_cols)}1')
        cell_title = ws['A1']
        cell_title.value = "CHEDRAUI"
        apply_style_dict(cell_title, ChedrauiStyles.get_title_style())
        ws.row_dimensions[1].height = self.DEFAULT_TITLE_HEIGHT

        # Fila 2: Título del reporte
        ws.merge_cells(f'A2:{get_column_letter(merge_cols)}2')
        cell_report = ws['A2']
        cell_report.value = title
        cell_report.font = Font(name='Arial', size=14, bold=True)
        cell_report.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = 25

        # Fila 3: Subtítulo e información
        if subtitle:
            ws['A3'] = subtitle
            ws['A3'].font = Font(name='Arial', size=11, italic=True)

        # Información del CEDIS y fecha
        fecha_col = get_column_letter(merge_cols - 1)
        ws[f'{fecha_col}3'] = f"CEDIS: {self.cedis}"
        ws[f'{fecha_col}3'].font = Font(name='Arial', size=10)
        ws[f'{fecha_col}3'].alignment = Alignment(horizontal='right')

        last_col = get_column_letter(merge_cols)
        ws[f'{last_col}3'] = f"Fecha: {self.created_at.strftime('%d/%m/%Y %H:%M')}"
        ws[f'{last_col}3'].font = Font(name='Arial', size=10)
        ws[f'{last_col}3'].alignment = Alignment(horizontal='right')

        return 4  # Siguiente fila disponible

    def apply_corporate_footer(self, ws: Worksheet, row: int = None) -> None:
        """
        Aplica pie de página corporativo

        Args:
            ws: Hoja de trabajo
            row: Fila donde colocar el footer (None = última + 2)
        """
        if row is None:
            row = ws.max_row + 2

        # Línea separadora
        ws[f'A{row}'] = "─" * 80
        ws[f'A{row}'].font = Font(color='CCCCCC')

        row += 1
        ws[f'A{row}'] = f"Generado por: {self.author}"
        ws[f'A{row}'].font = Font(name='Arial', size=9, italic=True, color='666666')

        row += 1
        ws[f'A{row}'] = f"Sistema SAC - CEDIS {self.cedis}"
        ws[f'A{row}'].font = Font(name='Arial', size=9, italic=True, color='666666')

        row += 1
        ws[f'A{row}'] = '"Las máquinas y los sistemas al servicio de los analistas"'
        ws[f'A{row}'].font = Font(name='Arial', size=8, italic=True, color='999999')

    def apply_chedraui_styles(self, ws: Worksheet) -> None:
        """
        Aplica estilos corporativos generales a la hoja

        Args:
            ws: Hoja de trabajo
        """
        # Configurar vista de impresión
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Configurar márgenes
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        # Encabezado y pie de página para impresión
        ws.oddHeader.center.text = f"CHEDRAUI - CEDIS {self.cedis}"
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.center.size = 12

        ws.oddFooter.left.text = f"Generado: {self.created_at.strftime('%d/%m/%Y %H:%M')}"
        ws.oddFooter.center.text = "Página &P de &N"
        ws.oddFooter.right.text = "Sistema SAC"

    def add_logo(self, ws: Worksheet, position: str = "A1") -> bool:
        """
        Agrega logo de Chedraui a la hoja

        Args:
            ws: Hoja de trabajo
            position: Celda donde colocar el logo

        Returns:
            bool: True si se agregó exitosamente
        """
        try:
            if self.LOGO_PATH.exists():
                img = Image(str(self.LOGO_PATH))
                img.width = 150
                img.height = 50
                ws.add_image(img, position)
                logger.debug(f"✅ Logo agregado en {position}")
                return True
            else:
                logger.debug(f"⚠️ Logo no encontrado en {self.LOGO_PATH}")
                return False
        except Exception as e:
            logger.warning(f"⚠️ No se pudo agregar logo: {e}")
            return False

    def create_table(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        start_row: int,
        start_col: int = 1,
        table_name: str = None,
        table_style: str = "TableStyleMedium2"
    ) -> Tuple[int, int]:
        """
        Crea una tabla Excel desde un DataFrame

        Args:
            ws: Hoja de trabajo
            df: DataFrame con los datos
            start_row: Fila inicial
            start_col: Columna inicial
            table_name: Nombre de la tabla (opcional)
            table_style: Estilo de tabla Excel

        Returns:
            Tuple[int, int]: (fila_final, columna_final)
        """
        if df.empty:
            ws.cell(row=start_row, column=start_col, value="Sin datos disponibles")
            return start_row, start_col

        # Escribir encabezados
        for col_idx, column in enumerate(df.columns, start_col):
            cell = ws.cell(row=start_row, column=col_idx, value=column)
            apply_style_dict(cell, {
                'font': ChedrauiStyles.get_header_style().font,
                'fill': ChedrauiStyles.get_header_style().fill,
                'alignment': ChedrauiStyles.get_header_style().alignment,
                'border': ChedrauiStyles.THIN_BORDER
            })

        ws.row_dimensions[start_row].height = self.DEFAULT_HEADER_HEIGHT

        # Escribir datos
        for row_idx, row_data in enumerate(df.values, start_row + 1):
            for col_idx, value in enumerate(row_data, start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = ChedrauiStyles.THIN_BORDER
                cell.alignment = Alignment(vertical='center')

                # Aplicar filas alternas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=ChedrauiColors.ALT_ROW,
                        end_color=ChedrauiColors.ALT_ROW,
                        fill_type='solid'
                    )

        end_row = start_row + len(df)
        end_col = start_col + len(df.columns) - 1

        # Crear tabla Excel (opcional)
        if table_name:
            try:
                ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
                table = Table(displayName=table_name, ref=ref)
                style = TableStyleInfo(
                    name=table_style,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            except Exception as e:
                logger.warning(f"⚠️ No se pudo crear tabla: {e}")

        return end_row, end_col

    def add_conditional_formatting(
        self,
        ws: Worksheet,
        rules: List[Dict[str, Any]]
    ) -> None:
        """
        Agrega formato condicional a la hoja

        Args:
            ws: Hoja de trabajo
            rules: Lista de reglas de formato condicional
                   Cada regla es un dict con:
                   - range: "A1:Z100"
                   - type: "color_scale", "data_bar", "cell_is", "formula"
                   - params: Parámetros específicos del tipo
        """
        for rule in rules:
            range_str = rule.get('range')
            rule_type = rule.get('type')
            params = rule.get('params', {})

            if not range_str or not rule_type:
                continue

            try:
                if rule_type == 'color_scale':
                    cf_rule = ColorScaleRule(
                        start_type=params.get('start_type', 'min'),
                        start_color=params.get('start_color', ChedrauiColors.OK_GREEN),
                        mid_type=params.get('mid_type', 'percentile'),
                        mid_value=params.get('mid_value', 50),
                        mid_color=params.get('mid_color', ChedrauiColors.WARNING_YELLOW),
                        end_type=params.get('end_type', 'max'),
                        end_color=params.get('end_color', ChedrauiColors.ERROR_RED)
                    )
                    ws.conditional_formatting.add(range_str, cf_rule)

                elif rule_type == 'data_bar':
                    cf_rule = DataBarRule(
                        start_type=params.get('start_type', 'min'),
                        end_type=params.get('end_type', 'max'),
                        color=params.get('color', ChedrauiColors.CHEDRAUI_BLUE)
                    )
                    ws.conditional_formatting.add(range_str, cf_rule)

                elif rule_type == 'cell_is':
                    red_fill = PatternFill(
                        start_color=params.get('fill_color', ChedrauiColors.LIGHT_RED),
                        end_color=params.get('fill_color', ChedrauiColors.LIGHT_RED),
                        fill_type='solid'
                    )
                    cf_rule = CellIsRule(
                        operator=params.get('operator', 'lessThan'),
                        formula=[params.get('value', '0')],
                        fill=red_fill
                    )
                    ws.conditional_formatting.add(range_str, cf_rule)

                elif rule_type == 'formula':
                    fill_color = params.get('fill_color', ChedrauiColors.WARNING_YELLOW)
                    fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type='solid'
                    )
                    cf_rule = FormulaRule(
                        formula=[params.get('formula', 'FALSE')],
                        fill=fill
                    )
                    ws.conditional_formatting.add(range_str, cf_rule)

            except Exception as e:
                logger.warning(f"⚠️ Error aplicando formato condicional: {e}")

    def auto_adjust_columns(
        self,
        ws: Worksheet,
        min_width: int = 10,
        max_width: int = 50
    ) -> None:
        """
        Ajusta automáticamente el ancho de las columnas

        Args:
            ws: Hoja de trabajo
            min_width: Ancho mínimo
            max_width: Ancho máximo
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass

            adjusted_width = max(min_width, min(max_length + 2, max_width))
            ws.column_dimensions[column_letter].width = adjusted_width

    def freeze_panes(self, ws: Worksheet, cell: str = "A5") -> None:
        """
        Congela paneles en la celda especificada

        Args:
            ws: Hoja de trabajo
            cell: Celda donde congelar (default: A5 para dejar encabezado visible)
        """
        ws.freeze_panes = cell

    def add_autofilter(self, ws: Worksheet, range_str: str = None) -> None:
        """
        Agrega autofiltro a un rango

        Args:
            ws: Hoja de trabajo
            range_str: Rango para filtro (None = detectar automáticamente)
        """
        if range_str is None:
            # Detectar rango de datos
            max_row = ws.max_row
            max_col = ws.max_column
            range_str = f"A4:{get_column_letter(max_col)}{max_row}"

        ws.auto_filter.ref = range_str

    def add_summary_section(
        self,
        ws: Worksheet,
        summary_data: Dict[str, Any],
        start_row: int,
        start_col: int = 1,
        title: str = "RESUMEN"
    ) -> int:
        """
        Agrega sección de resumen

        Args:
            ws: Hoja de trabajo
            summary_data: Diccionario con datos de resumen
            start_row: Fila inicial
            start_col: Columna inicial
            title: Título de la sección

        Returns:
            int: Última fila utilizada
        """
        # Título del resumen
        title_cell = ws.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(name='Arial', size=12, bold=True)
        title_cell.fill = PatternFill(
            start_color=ChedrauiColors.SUBHEADER,
            end_color=ChedrauiColors.SUBHEADER,
            fill_type='solid'
        )
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=start_col + 1
        )

        row = start_row + 1
        for key, value in summary_data.items():
            ws.cell(row=row, column=start_col, value=key)
            ws.cell(row=row, column=start_col).font = Font(name='Arial', size=10)

            value_cell = ws.cell(row=row, column=start_col + 1, value=value)
            value_cell.font = Font(name='Arial', size=10, bold=True)

            # Colorear según tipo de valor
            if isinstance(value, (int, float)):
                if 'error' in key.lower() or 'crítico' in key.lower():
                    if value > 0:
                        value_cell.fill = PatternFill(
                            start_color=ChedrauiColors.LIGHT_RED,
                            end_color=ChedrauiColors.LIGHT_RED,
                            fill_type='solid'
                        )
                elif 'ok' in key.lower() or 'éxito' in key.lower():
                    value_cell.fill = PatternFill(
                        start_color=ChedrauiColors.LIGHT_GREEN,
                        end_color=ChedrauiColors.LIGHT_GREEN,
                        fill_type='solid'
                    )

            row += 1

        return row

    def save(self, filename: str) -> str:
        """
        Guarda el workbook en un archivo

        Args:
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo guardado
        """
        if self.workbook:
            self.workbook.save(filename)
            logger.info(f"✅ Reporte guardado: {filename}")
            return filename
        else:
            raise ValueError("No hay workbook para guardar. Llame create_workbook() primero.")


# ═══════════════════════════════════════════════════════════════
# EXPORTACIÓN
# ═══════════════════════════════════════════════════════════════

__all__ = ['ExcelTemplate']
