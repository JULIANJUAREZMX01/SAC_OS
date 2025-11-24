"""
═══════════════════════════════════════════════════════════════
TEMPLATES DE REPORTES ESPECÍFICOS
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Templates específicos para cada tipo de reporte del sistema SAC.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Optional, Dict, List, Any
import logging

from .base_template import ExcelTemplate
from ..excel_styles import ChedrauiStyles, ChedrauiColors, apply_style_dict

logger = logging.getLogger(__name__)


class OCValidationTemplate(ExcelTemplate):
    """
    Template para reportes de validación OC vs Distribuciones
    """

    def create_report(
        self,
        df_validation: pd.DataFrame,
        filename: str = None
    ) -> str:
        """
        Crea reporte de validación OC vs Distribuciones

        Args:
            df_validation: DataFrame con datos de validación
            filename: Nombre del archivo (opcional)

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Validacion_OC_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Validación OC"

        # Encabezado corporativo
        start_row = self.apply_corporate_header(
            ws,
            title="REPORTE DE VALIDACIÓN OC vs DISTRIBUCIONES",
            subtitle="Comparación de cantidades ordenadas vs distribuidas",
            merge_cols=max(8, len(df_validation.columns))
        )

        # Crear tabla de datos
        end_row, end_col = self.create_table(ws, df_validation, start_row)

        # Aplicar formato de estado si existe columna STATUS
        if 'STATUS' in df_validation.columns:
            self._apply_status_formatting(ws, df_validation, start_row)

        # Agregar resumen
        self._add_validation_summary(ws, df_validation, end_col + 2, start_row)

        # Formato condicional para diferencias
        if 'DIFERENCIA' in df_validation.columns:
            diff_col = df_validation.columns.get_loc('DIFERENCIA') + 1
            col_letter = get_column_letter(diff_col)
            self.add_conditional_formatting(ws, [{
                'range': f'{col_letter}{start_row + 1}:{col_letter}{end_row}',
                'type': 'cell_is',
                'params': {
                    'operator': 'notEqual',
                    'value': '0',
                    'fill_color': ChedrauiColors.WARNING_YELLOW
                }
            }])

        # Configuración final
        self.auto_adjust_columns(ws)
        self.freeze_panes(ws, f"A{start_row + 1}")
        self.add_autofilter(ws)
        self.apply_chedraui_styles(ws)
        self.apply_corporate_footer(ws)

        return self.save(filename)

    def _apply_status_formatting(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        start_row: int
    ) -> None:
        """Aplica formato según status"""
        status_col = df.columns.get_loc('STATUS') + 1

        for row_idx in range(start_row + 1, start_row + len(df) + 1):
            status_cell = ws.cell(row=row_idx, column=status_col)
            status_value = str(status_cell.value).lower() if status_cell.value else ""

            style = ChedrauiStyles.get_status_style(status_value)
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                apply_style_dict(cell, style)

    def _add_validation_summary(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        start_col: int,
        start_row: int
    ) -> None:
        """Agrega resumen de validación"""
        if 'STATUS' not in df.columns:
            return

        total = len(df)
        ok = len(df[df['STATUS'].str.contains('OK', case=False, na=False)])
        incomplete = len(df[df['STATUS'].str.contains('incompleta', case=False, na=False)])
        excess = len(df[df['STATUS'].str.contains('excedente', case=False, na=False)])
        no_distro = len(df[df['STATUS'].str.contains('sin distro', case=False, na=False)])

        summary_data = {
            'Total OC:': total,
            '✅ OK:': ok,
            '⚠️ Incompletas:': incomplete,
            '🔴 Excedentes:': excess,
            '❌ Sin Distro:': no_distro
        }

        self.add_summary_section(ws, summary_data, start_row, start_col, "RESUMEN")


class DailyPlanningTemplate(ExcelTemplate):
    """
    Template para reportes diarios de Planning
    """

    def create_report(
        self,
        df_oc: pd.DataFrame,
        df_asn: pd.DataFrame,
        df_errors: pd.DataFrame,
        filename: str = None
    ) -> str:
        """
        Crea reporte diario de Planning completo

        Args:
            df_oc: DataFrame de órdenes de compra
            df_asn: DataFrame de ASN
            df_errors: DataFrame de errores
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            fecha = datetime.now().strftime('%Y%m%d')
            filename = f"Planning_Diario_{self.cedis}_{fecha}.xlsx"

        wb = self.create_workbook()

        # Hoja 1: Resumen Ejecutivo
        self._create_executive_summary(wb, df_oc, df_asn, df_errors)

        # Hoja 2: Órdenes de Compra
        if not df_oc.empty:
            ws_oc = wb.create_sheet("Órdenes de Compra")
            start_row = self.apply_corporate_header(
                ws_oc, "ÓRDENES DE COMPRA DEL DÍA",
                merge_cols=len(df_oc.columns)
            )
            self.create_table(ws_oc, df_oc, start_row)
            self.auto_adjust_columns(ws_oc)
            self.freeze_panes(ws_oc, f"A{start_row + 1}")

        # Hoja 3: Status ASN
        if not df_asn.empty:
            ws_asn = wb.create_sheet("Status ASN")
            start_row = self.apply_corporate_header(
                ws_asn, "STATUS DE ASN",
                merge_cols=len(df_asn.columns)
            )
            self.create_table(ws_asn, df_asn, start_row)
            self.auto_adjust_columns(ws_asn)

        # Hoja 4: Errores
        if not df_errors.empty:
            ws_err = wb.create_sheet("Errores")
            start_row = self.apply_corporate_header(
                ws_err, "ERRORES Y ALERTAS DETECTADOS",
                merge_cols=len(df_errors.columns)
            )
            self.create_table(ws_err, df_errors, start_row)
            self._apply_severity_formatting(ws_err, df_errors, start_row)
            self.auto_adjust_columns(ws_err)

        return self.save(filename)

    def _create_executive_summary(
        self,
        wb: Workbook,
        df_oc: pd.DataFrame,
        df_asn: pd.DataFrame,
        df_errors: pd.DataFrame
    ) -> None:
        """Crea hoja de resumen ejecutivo"""
        ws = wb.active
        ws.title = "Resumen Ejecutivo"

        # Encabezado
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "📊 RESUMEN EJECUTIVO - PLANNING DIARIO"
        apply_style_dict(title_cell, ChedrauiStyles.get_title_style())
        ws.row_dimensions[1].height = 40

        # Información general
        ws['A3'] = f"CEDIS: {self.cedis}"
        ws['A3'].font = Font(name='Arial', size=12, bold=True)
        ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        # KPIs
        ws['A6'] = "INDICADORES PRINCIPALES"
        ws['A6'].font = Font(name='Arial', size=14, bold=True)
        ws['A6'].fill = PatternFill(
            start_color=ChedrauiColors.SUBHEADER,
            end_color=ChedrauiColors.SUBHEADER,
            fill_type='solid'
        )
        ws.merge_cells('A6:C6')

        kpis = [
            ("Total Órdenes de Compra:", len(df_oc) if not df_oc.empty else 0, None),
            ("Total ASN Activos:", len(df_asn) if not df_asn.empty else 0, None),
            ("Errores Detectados:", len(df_errors) if not df_errors.empty else 0,
             ChedrauiColors.ERROR_RED if len(df_errors) > 0 else ChedrauiColors.OK_GREEN),
        ]

        row = 7
        for label, value, color in kpis:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(name='Arial', size=11)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
            if color:
                ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True, color=color)
            row += 1

        # Footer
        self.apply_corporate_footer(ws, row + 5)

    def _apply_severity_formatting(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        start_row: int
    ) -> None:
        """Aplica formato según severidad"""
        if 'Severidad' not in df.columns:
            return

        sev_col = df.columns.get_loc('Severidad') + 1

        for row_idx in range(start_row + 1, start_row + len(df) + 1):
            sev_cell = ws.cell(row=row_idx, column=sev_col)
            sev_value = str(sev_cell.value).lower() if sev_cell.value else ""

            style = ChedrauiStyles.get_status_style(sev_value)
            for col_idx in range(1, len(df.columns) + 1):
                apply_style_dict(ws.cell(row=row_idx, column=col_idx), style)


class DistributionTemplate(ExcelTemplate):
    """
    Template para reportes de distribuciones
    """

    def create_report(
        self,
        df_distributions: pd.DataFrame,
        oc_number: str,
        filename: str = None
    ) -> str:
        """
        Crea reporte de distribuciones por OC

        Args:
            df_distributions: DataFrame con distribuciones
            oc_number: Número de OC
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Distribuciones_OC_{oc_number}_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Distribuciones"

        # Encabezado
        start_row = self.apply_corporate_header(
            ws,
            title=f"DISTRIBUCIONES - OC {oc_number}",
            subtitle="Detalle por Tienda y SKU"
        )

        # Tabla principal
        self.create_table(ws, df_distributions, start_row)
        self.auto_adjust_columns(ws)
        self.freeze_panes(ws, f"A{start_row + 1}")
        self.add_autofilter(ws)

        # Hoja de pivot por tienda
        if not df_distributions.empty and 'TIENDA' in df_distributions.columns:
            self._create_store_pivot(wb, df_distributions)

        self.apply_corporate_footer(ws)
        return self.save(filename)

    def _create_store_pivot(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Crea pivot por tienda"""
        ws = wb.create_sheet("Por Tienda")

        qty_col = None
        for col in ['DISTR_QTY', 'CANTIDAD', 'QTY']:
            if col in df.columns:
                qty_col = col
                break

        if qty_col is None:
            return

        pivot = df.groupby('TIENDA')[qty_col].sum().reset_index()
        pivot.columns = ['Tienda', 'Total Unidades']
        pivot = pivot.sort_values('Total Unidades', ascending=False)

        start_row = self.apply_corporate_header(
            ws, "DISTRIBUCIÓN POR TIENDA",
            merge_cols=2
        )

        self.create_table(ws, pivot, start_row)
        self.auto_adjust_columns(ws)


class ErrorReportTemplate(ExcelTemplate):
    """
    Template para reportes de errores
    """

    def create_report(
        self,
        df_errors: pd.DataFrame,
        filename: str = None
    ) -> str:
        """
        Crea reporte de errores detectados

        Args:
            df_errors: DataFrame con errores
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Errores_Detectados_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Errores"

        start_row = self.apply_corporate_header(
            ws,
            title="REPORTE DE ERRORES Y ALERTAS",
            subtitle="Errores detectados durante la validación"
        )

        self.create_table(ws, df_errors, start_row)

        # Formatear por severidad
        if 'Severidad' in df_errors.columns:
            sev_col = df_errors.columns.get_loc('Severidad') + 1
            for row_idx in range(start_row + 1, start_row + len(df_errors) + 1):
                sev_value = str(ws.cell(row=row_idx, column=sev_col).value).lower()
                style = ChedrauiStyles.get_status_style(sev_value)
                for col_idx in range(1, len(df_errors.columns) + 1):
                    apply_style_dict(ws.cell(row=row_idx, column=col_idx), style)

        self.auto_adjust_columns(ws)
        self.apply_corporate_footer(ws)
        return self.save(filename)


class ASNStatusTemplate(ExcelTemplate):
    """
    Template para reportes de status de ASN
    """

    def create_report(
        self,
        df_asn: pd.DataFrame,
        filename: str = None
    ) -> str:
        """
        Crea reporte de status ASN

        Args:
            df_asn: DataFrame con datos de ASN
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"ASN_Status_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Status ASN"

        start_row = self.apply_corporate_header(
            ws,
            title="REPORTE DE STATUS ASN",
            subtitle="Advanced Shipping Notices activos"
        )

        self.create_table(ws, df_asn, start_row)
        self.auto_adjust_columns(ws)
        self.freeze_panes(ws, f"A{start_row + 1}")
        self.add_autofilter(ws)
        self.apply_corporate_footer(ws)

        return self.save(filename)


class InventoryTemplate(ExcelTemplate):
    """
    Template para reportes de inventario
    """

    def create_report(
        self,
        df_inventory: pd.DataFrame,
        filename: str = None
    ) -> str:
        """
        Crea reporte de inventario

        Args:
            df_inventory: DataFrame con datos de inventario
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Inventario_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Inventario"

        start_row = self.apply_corporate_header(
            ws,
            title="REPORTE DE INVENTARIO",
            subtitle="Estado actual del inventario en almacén"
        )

        end_row, _ = self.create_table(ws, df_inventory, start_row)

        # Formato condicional para stock bajo
        if 'STOCK' in df_inventory.columns:
            stock_col = df_inventory.columns.get_loc('STOCK') + 1
            col_letter = get_column_letter(stock_col)
            self.add_conditional_formatting(ws, [{
                'range': f'{col_letter}{start_row + 1}:{col_letter}{end_row}',
                'type': 'color_scale',
                'params': {
                    'start_type': 'min',
                    'start_color': ChedrauiColors.ERROR_RED,
                    'mid_type': 'percentile',
                    'mid_value': 50,
                    'mid_color': ChedrauiColors.WARNING_YELLOW,
                    'end_type': 'max',
                    'end_color': ChedrauiColors.OK_GREEN
                }
            }])

        self.auto_adjust_columns(ws)
        self.apply_corporate_footer(ws)
        return self.save(filename)


class AuditTemplate(ExcelTemplate):
    """
    Template para reportes de auditoría
    """

    def create_report(
        self,
        df_audit: pd.DataFrame,
        audit_type: str = "General",
        filename: str = None
    ) -> str:
        """
        Crea reporte de auditoría

        Args:
            df_audit: DataFrame con datos de auditoría
            audit_type: Tipo de auditoría
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Auditoria_{audit_type}_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Auditoría"

        start_row = self.apply_corporate_header(
            ws,
            title=f"REPORTE DE AUDITORÍA - {audit_type.upper()}",
            subtitle="Registro de acciones y cambios"
        )

        self.create_table(ws, df_audit, start_row)
        self.auto_adjust_columns(ws)
        self.freeze_panes(ws, f"A{start_row + 1}")
        self.apply_corporate_footer(ws)

        return self.save(filename)


class KPITemplate(ExcelTemplate):
    """
    Template para reportes de KPIs
    """

    def create_report(
        self,
        kpis: Dict[str, Any],
        filename: str = None
    ) -> str:
        """
        Crea reporte de KPIs

        Args:
            kpis: Diccionario con KPIs
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"KPIs_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "KPIs"

        # Encabezado
        start_row = self.apply_corporate_header(
            ws,
            title="INDICADORES CLAVE DE DESEMPEÑO (KPIs)",
            subtitle="Métricas de operación"
        )

        # Crear tarjetas de KPI
        row = start_row + 1
        col = 1

        for kpi_name, kpi_data in kpis.items():
            value = kpi_data.get('value', 0)
            target = kpi_data.get('target', None)
            unit = kpi_data.get('unit', '')
            trend = kpi_data.get('trend', 'neutral')

            # Nombre del KPI
            ws.cell(row=row, column=col, value=kpi_name)
            ws.cell(row=row, column=col).font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=ChedrauiColors.SUBHEADER,
                end_color=ChedrauiColors.SUBHEADER,
                fill_type='solid'
            )
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)

            # Valor
            row += 1
            value_cell = ws.cell(row=row, column=col, value=f"{value}{unit}")

            if trend == 'positive':
                style = ChedrauiStyles.get_kpi_positive_style()
            elif trend == 'negative':
                style = ChedrauiStyles.get_kpi_negative_style()
            else:
                style = ChedrauiStyles.get_kpi_neutral_style()

            apply_style_dict(value_cell, style)

            # Target si existe
            if target:
                row += 1
                ws.cell(row=row, column=col, value=f"Meta: {target}{unit}")
                ws.cell(row=row, column=col).font = Font(name='Arial', size=9, italic=True)

            row += 2
            if row > 20:
                row = start_row + 1
                col += 3

        self.auto_adjust_columns(ws)
        self.apply_corporate_footer(ws)
        return self.save(filename)


class ComparativeTemplate(ExcelTemplate):
    """
    Template para reportes comparativos
    """

    def create_report(
        self,
        df_current: pd.DataFrame,
        df_previous: pd.DataFrame,
        comparison_type: str = "Período",
        filename: str = None
    ) -> str:
        """
        Crea reporte comparativo entre dos períodos

        Args:
            df_current: DataFrame período actual
            df_previous: DataFrame período anterior
            comparison_type: Tipo de comparación
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Comparativo_{comparison_type}_{timestamp}.xlsx"

        wb = self.create_workbook()

        # Hoja comparativa
        ws = wb.active
        ws.title = "Comparativo"

        start_row = self.apply_corporate_header(
            ws,
            title=f"REPORTE COMPARATIVO - {comparison_type.upper()}",
            subtitle="Análisis de variaciones"
        )

        # Datos actuales
        ws.cell(row=start_row, column=1, value="PERÍODO ACTUAL")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        self.create_table(ws, df_current, start_row + 1)

        # Datos anteriores
        ws_prev = wb.create_sheet("Período Anterior")
        self.apply_corporate_header(ws_prev, "PERÍODO ANTERIOR")
        self.create_table(ws_prev, df_previous, 5)
        self.auto_adjust_columns(ws_prev)

        self.auto_adjust_columns(ws)
        self.apply_corporate_footer(ws)
        return self.save(filename)


class ExecutiveDashboardTemplate(ExcelTemplate):
    """
    Template para dashboard ejecutivo
    """

    def create_report(
        self,
        kpis: Dict[str, Any],
        df_summary: pd.DataFrame,
        charts_data: Dict[str, pd.DataFrame] = None,
        filename: str = None
    ) -> str:
        """
        Crea dashboard ejecutivo completo

        Args:
            kpis: Diccionario con KPIs principales
            df_summary: DataFrame con resumen de datos
            charts_data: Datos para gráficos
            filename: Nombre del archivo

        Returns:
            str: Ruta del archivo generado
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"Dashboard_Ejecutivo_{timestamp}.xlsx"

        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Dashboard"

        # Header grande
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"📊 DASHBOARD EJECUTIVO - CEDIS {self.cedis}"
        apply_style_dict(title_cell, ChedrauiStyles.get_title_style())
        ws.row_dimensions[1].height = 50

        ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True)

        # Sección de KPIs
        ws['A4'] = "INDICADORES CLAVE"
        ws['A4'].font = Font(name='Arial', size=14, bold=True)
        ws['A4'].fill = PatternFill(
            start_color=ChedrauiColors.HEADER_RED,
            end_color=ChedrauiColors.HEADER_RED,
            fill_type='solid'
        )
        ws['A4'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        ws.merge_cells('A4:J4')

        # KPIs en fila
        col = 1
        for name, data in kpis.items():
            value = data.get('value', 0)
            unit = data.get('unit', '')
            trend = data.get('trend', 'neutral')

            ws.cell(row=5, column=col, value=name)
            ws.cell(row=5, column=col).font = Font(size=10, bold=True)
            ws.cell(row=5, column=col).alignment = Alignment(horizontal='center')

            value_cell = ws.cell(row=6, column=col, value=f"{value}{unit}")
            if trend == 'positive':
                value_cell.font = Font(size=16, bold=True, color=ChedrauiColors.OK_GREEN)
            elif trend == 'negative':
                value_cell.font = Font(size=16, bold=True, color=ChedrauiColors.ERROR_RED)
            else:
                value_cell.font = Font(size=16, bold=True, color=ChedrauiColors.CHEDRAUI_BLUE)
            value_cell.alignment = Alignment(horizontal='center')

            col += 2

        # Tabla resumen
        if not df_summary.empty:
            ws['A9'] = "RESUMEN DE DATOS"
            ws['A9'].font = Font(name='Arial', size=12, bold=True)
            self.create_table(ws, df_summary, 10)

        self.auto_adjust_columns(ws)
        self.apply_corporate_footer(ws)

        return self.save(filename)


# ═══════════════════════════════════════════════════════════════
# EXPORTACIÓN
# ═══════════════════════════════════════════════════════════════

__all__ = [
    'OCValidationTemplate',
    'DailyPlanningTemplate',
    'DistributionTemplate',
    'ErrorReportTemplate',
    'ASNStatusTemplate',
    'InventoryTemplate',
    'AuditTemplate',
    'KPITemplate',
    'ComparativeTemplate',
    'ExecutiveDashboardTemplate',
]
