"""
═══════════════════════════════════════════════════════════════
TESTS PARA MÓDULO DE EXPORTACIÓN
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Tests unitarios para el gestor de exportación de reportes.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pytest
import pandas as pd
from pathlib import Path
import tempfile
import os
import json
import zipfile

# Importar módulos a testear
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from modules.export_manager import (
    ExportManager, ExportConfig, ExportHistory,
    quick_export, export_with_compression
)


class TestExportConfig:
    """Tests para la clase ExportConfig"""

    def test_default_config(self):
        """Verifica valores por defecto de configuración"""
        config = ExportConfig()

        assert config.include_timestamp is True
        assert config.compression is False
        assert config.encoding == 'utf-8'

    def test_custom_output_dir(self):
        """Verifica configuración con directorio personalizado"""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(output_dir=Path(tmpdir))
            assert config.output_dir == Path(tmpdir)


class TestExportHistory:
    """Tests para la clase ExportHistory"""

    def test_add_export(self):
        """Verifica agregar exportación al historial"""
        history = ExportHistory()

        history.add_export(
            filename="test.xlsx",
            format_type="xlsx",
            size_bytes=1024,
            rows=100,
            success=True
        )

        exports = history.get_history()
        assert len(exports) == 1
        assert exports[0]['filename'] == "test.xlsx"
        assert exports[0]['success'] is True

    def test_add_failed_export(self):
        """Verifica agregar exportación fallida"""
        history = ExportHistory()

        history.add_export(
            filename="error.xlsx",
            format_type="xlsx",
            size_bytes=0,
            rows=0,
            success=False,
            error="Test error"
        )

        exports = history.get_history()
        assert exports[0]['success'] is False
        assert exports[0]['error'] == "Test error"

    def test_get_summary(self):
        """Verifica obtener resumen de exportaciones"""
        history = ExportHistory()

        history.add_export("file1.xlsx", "xlsx", 1024, 10, True)
        history.add_export("file2.csv", "csv", 512, 10, True)
        history.add_export("error.xlsx", "xlsx", 0, 0, False, "Error")

        summary = history.get_summary()

        assert summary['total_exports'] == 3
        assert summary['successful'] == 2
        assert summary['failed'] == 1


class TestExportManager:
    """Tests para la clase ExportManager"""

    @pytest.fixture
    def temp_output_dir(self):
        """Fixture para directorio temporal de salida"""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield Path(tmpdir)

    @pytest.fixture
    def export_manager(self, temp_output_dir):
        """Fixture para crear ExportManager con directorio temporal"""
        config = ExportConfig(output_dir=temp_output_dir)
        return ExportManager(config)

    @pytest.fixture
    def sample_df(self):
        """Fixture para crear DataFrame de prueba"""
        return pd.DataFrame({
            'Column1': ['A', 'B', 'C'],
            'Column2': [1, 2, 3],
            'Column3': [10.5, 20.5, 30.5]
        })

    def test_export_to_excel(self, export_manager, sample_df):
        """Verifica exportación a Excel"""
        result = export_manager.export_to_excel(
            sample_df, "test_export"
        )

        assert os.path.exists(result)
        assert result.endswith('.xlsx')

        # Verificar contenido
        from openpyxl import load_workbook
        wb = load_workbook(result)
        ws = wb.active
        assert ws['A1'].value == 'Column1'

    def test_export_to_excel_with_title(self, export_manager, sample_df):
        """Verifica exportación a Excel con título"""
        result = export_manager.export_to_excel(
            sample_df, "test_with_title",
            title="Test Report Title"
        )

        assert os.path.exists(result)

    def test_export_to_csv(self, export_manager, sample_df):
        """Verifica exportación a CSV"""
        result = export_manager.export_to_csv(sample_df, "test_export")

        assert os.path.exists(result)
        assert result.endswith('.csv')

        # Verificar contenido
        df_loaded = pd.read_csv(result)
        assert len(df_loaded) == 3
        assert 'Column1' in df_loaded.columns

    def test_export_to_csv_custom_separator(self, export_manager, sample_df):
        """Verifica exportación a CSV con separador personalizado"""
        result = export_manager.export_to_csv(
            sample_df, "test_semicolon",
            separator=';'
        )

        with open(result, 'r') as f:
            content = f.read()
            assert ';' in content

    def test_export_to_json(self, export_manager, sample_df):
        """Verifica exportación a JSON"""
        result = export_manager.export_to_json(sample_df, "test_export")

        assert os.path.exists(result)
        assert result.endswith('.json')

        # Verificar contenido
        with open(result, 'r') as f:
            data = json.load(f)
            assert len(data) == 3

    def test_export_multiple_sheets(self, export_manager, sample_df):
        """Verifica exportación con múltiples hojas"""
        dfs_dict = {
            'Sheet1': sample_df,
            'Sheet2': sample_df.copy(),
            'Sheet3': pd.DataFrame({'X': [1, 2], 'Y': [3, 4]})
        }

        result = export_manager.export_multiple_sheets(dfs_dict, "multi_sheet")

        assert os.path.exists(result)

        # Verificar hojas
        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert len(wb.sheetnames) == 3

    def test_compress_reports(self, export_manager, sample_df, temp_output_dir):
        """Verifica compresión de reportes"""
        # Crear algunos archivos
        file1 = export_manager.export_to_excel(sample_df, "file1")
        file2 = export_manager.export_to_csv(sample_df, "file2")

        # Comprimir
        result = export_manager.compress_reports([file1, file2], "archive")

        assert os.path.exists(result)
        assert result.endswith('.zip')

        # Verificar contenido del ZIP
        with zipfile.ZipFile(result, 'r') as zipf:
            names = zipf.namelist()
            assert len(names) == 2

    def test_export_to_multiple_formats(self, export_manager, sample_df):
        """Verifica exportación a múltiples formatos"""
        results = export_manager.export_to_multiple_formats(
            sample_df, "multi_format",
            formats=['xlsx', 'csv', 'json']
        )

        assert 'xlsx' in results
        assert 'csv' in results
        assert 'json' in results

        for fmt, path in results.items():
            if not path.startswith('Error'):
                assert os.path.exists(path)

    def test_get_export_history(self, export_manager, sample_df):
        """Verifica obtener historial de exportaciones"""
        export_manager.export_to_excel(sample_df, "test1")
        export_manager.export_to_csv(sample_df, "test2")

        history = export_manager.get_export_history()
        assert len(history) == 2

    def test_get_export_summary(self, export_manager, sample_df):
        """Verifica obtener resumen de exportaciones"""
        export_manager.export_to_excel(sample_df, "test1")
        export_manager.export_to_csv(sample_df, "test2")

        summary = export_manager.get_export_summary()
        assert summary['total_exports'] == 2
        assert summary['successful'] == 2


class TestQuickExport:
    """Tests para la función quick_export"""

    @pytest.fixture
    def sample_df(self):
        return pd.DataFrame({
            'A': [1, 2, 3],
            'B': ['x', 'y', 'z']
        })

    def test_quick_export_xlsx(self, sample_df):
        """Verifica exportación rápida a Excel"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Cambiar al directorio temporal
            original_dir = os.getcwd()
            os.chdir(tmpdir)

            try:
                result = quick_export(sample_df, "quick_test", "xlsx")
                assert os.path.exists(result)
            finally:
                os.chdir(original_dir)

    def test_quick_export_csv(self, sample_df):
        """Verifica exportación rápida a CSV"""
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = os.getcwd()
            os.chdir(tmpdir)

            try:
                result = quick_export(sample_df, "quick_test", "csv")
                assert os.path.exists(result)
            finally:
                os.chdir(original_dir)

    def test_quick_export_invalid_format(self, sample_df):
        """Verifica manejo de formato inválido"""
        with pytest.raises(ValueError):
            quick_export(sample_df, "test", "invalid_format")


class TestExportWithCompression:
    """Tests para la función export_with_compression"""

    def test_export_and_compress(self):
        """Verifica exportación y compresión"""
        df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})

        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = os.getcwd()
            os.chdir(tmpdir)

            try:
                result = export_with_compression(
                    df, "compressed_test",
                    formats=['xlsx', 'csv']
                )
                assert os.path.exists(result)
                assert result.endswith('.zip')
            finally:
                os.chdir(original_dir)


class TestExportManagerIntegration:
    """Tests de integración para ExportManager"""

    def test_full_export_workflow(self):
        """Test de flujo completo de exportación"""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(output_dir=Path(tmpdir))
            manager = ExportManager(config)

            # Crear datos
            df_main = pd.DataFrame({
                'OC': ['OC001', 'OC002'],
                'Cantidad': [100, 200],
                'Status': ['OK', 'Pendiente']
            })

            df_details = pd.DataFrame({
                'SKU': ['SKU1', 'SKU2', 'SKU3'],
                'Tienda': ['T1', 'T1', 'T2'],
                'Qty': [50, 30, 120]
            })

            # Exportar a Excel con múltiples hojas
            multi_result = manager.export_multiple_sheets(
                {'Resumen': df_main, 'Detalles': df_details},
                "informe_completo",
                title="Informe Completo"
            )
            assert os.path.exists(multi_result)

            # Exportar a CSV
            csv_result = manager.export_to_csv(df_main, "resumen")
            assert os.path.exists(csv_result)

            # Comprimir todo
            zip_result = manager.compress_reports(
                [multi_result, csv_result],
                "paquete_informes"
            )
            assert os.path.exists(zip_result)

            # Verificar historial
            history = manager.get_export_history()
            assert len(history) == 3

            # Verificar resumen
            summary = manager.get_export_summary()
            assert summary['successful'] == 3


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
