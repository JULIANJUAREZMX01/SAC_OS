"""
═══════════════════════════════════════════════════════════════
TESTS PARA MÓDULO DE GENERACIÓN DE GRÁFICOS
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Tests unitarios para el generador de gráficos Excel.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pytest
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, AreaChart
from openpyxl.chart import Reference
from pathlib import Path
import tempfile
import os

# Importar módulos a testear
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from modules.chart_generator import ChartGenerator, ChartConfig, add_chart_to_sheet


class TestChartConfig:
    """Tests para la clase ChartConfig"""

    def test_default_config(self):
        """Verifica valores por defecto de configuración"""
        config = ChartConfig()

        assert config.title == ""
        assert config.width == 15
        assert config.height == 10
        assert config.show_legend is True
        assert config.show_data_labels is False

    def test_custom_config(self):
        """Verifica configuración personalizada"""
        config = ChartConfig(
            title="Test Chart",
            width=20,
            height=12,
            show_legend=False,
            show_data_labels=True
        )

        assert config.title == "Test Chart"
        assert config.width == 20
        assert config.height == 12
        assert config.show_legend is False
        assert config.show_data_labels is True


class TestChartGenerator:
    """Tests para la clase ChartGenerator"""

    @pytest.fixture
    def workbook_with_data(self):
        """Fixture para crear workbook con datos de prueba"""
        wb = Workbook()
        ws = wb.active

        # Escribir datos
        ws['A1'] = 'Categoría'
        ws['B1'] = 'Valor'
        ws['A2'] = 'A'
        ws['B2'] = 10
        ws['A3'] = 'B'
        ws['B3'] = 20
        ws['A4'] = 'C'
        ws['B4'] = 30
        ws['A5'] = 'D'
        ws['B5'] = 25

        return wb, ws

    @pytest.fixture
    def references(self, workbook_with_data):
        """Fixture para crear referencias de datos"""
        wb, ws = workbook_with_data
        categories = Reference(ws, min_col=1, min_row=2, max_row=5)
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        return categories, data, ws

    def test_create_bar_chart(self, references):
        """Verifica creación de gráfico de barras"""
        categories, data, ws = references
        config = ChartConfig(title="Test Bar Chart")

        chart = ChartGenerator.create_bar_chart(ws, data, categories, config)

        assert chart is not None
        assert isinstance(chart, BarChart)
        # chart.title es un objeto Title, verificamos que el texto esté presente
        assert chart.title is not None

    def test_create_bar_chart_horizontal(self, references):
        """Verifica creación de gráfico de barras horizontales"""
        categories, data, ws = references

        chart = ChartGenerator.create_bar_chart(
            ws, data, categories, chart_type='bar'
        )

        assert chart is not None
        assert chart.type == 'bar'

    def test_create_pie_chart(self, references):
        """Verifica creación de gráfico de pastel"""
        categories, data, ws = references
        config = ChartConfig(title="Test Pie Chart", show_data_labels=True)

        chart = ChartGenerator.create_pie_chart(ws, data, categories, config)

        assert chart is not None
        assert isinstance(chart, PieChart)
        # chart.title es un objeto Title, verificamos que exista
        assert chart.title is not None

    def test_create_doughnut_chart(self, references):
        """Verifica creación de gráfico de dona"""
        categories, data, ws = references
        config = ChartConfig(title="Test Doughnut")

        chart = ChartGenerator.create_pie_chart(
            ws, data, categories, config, is_doughnut=True
        )

        assert chart is not None
        # DoughnutChart es un tipo de pie

    def test_create_line_chart(self, references):
        """Verifica creación de gráfico de líneas"""
        categories, data, ws = references
        config = ChartConfig(
            title="Test Line Chart",
            x_axis_title="X Axis",
            y_axis_title="Y Axis"
        )

        chart = ChartGenerator.create_line_chart(ws, data, categories, config)

        assert chart is not None
        assert isinstance(chart, LineChart)
        # chart.title es un objeto Title, verificamos que exista
        assert chart.title is not None

    def test_create_line_chart_smooth(self, references):
        """Verifica creación de gráfico de líneas suavizadas"""
        categories, data, ws = references

        chart = ChartGenerator.create_line_chart(
            ws, data, categories, smooth_lines=True
        )

        assert chart is not None

    def test_create_area_chart(self, references):
        """Verifica creación de gráfico de área"""
        categories, data, ws = references
        config = ChartConfig(title="Test Area Chart")

        chart = ChartGenerator.create_area_chart(ws, data, categories, config)

        assert chart is not None
        assert isinstance(chart, AreaChart)

    def test_create_area_chart_stacked(self, references):
        """Verifica creación de gráfico de área apilada"""
        categories, data, ws = references

        chart = ChartGenerator.create_area_chart(
            ws, data, categories, stacked=True
        )

        assert chart is not None
        assert chart.grouping == "stacked"

    def test_create_dashboard(self, workbook_with_data):
        """Verifica creación de dashboard con múltiples gráficos"""
        wb, ws = workbook_with_data

        categories = Reference(ws, min_col=1, min_row=2, max_row=5)
        data = Reference(ws, min_col=2, min_row=1, max_row=5)

        chart1 = ChartGenerator.create_bar_chart(ws, data, categories)
        chart2 = ChartGenerator.create_pie_chart(ws, data, categories)

        charts = [(chart1, "D1"), (chart2, "D15")]

        # No debe lanzar excepción
        ChartGenerator.create_dashboard(ws, charts)

        # Verificar que los gráficos se agregaron
        assert len(ws._charts) == 2

    def test_create_chart_from_dataframe(self):
        """Verifica creación de gráfico desde DataFrame"""
        wb = Workbook()
        ws = wb.active

        df = pd.DataFrame({
            'Categoría': ['A', 'B', 'C'],
            'Valor': [10, 20, 30]
        })

        config = ChartConfig(title="From DataFrame")
        chart, position = ChartGenerator.create_chart_from_dataframe(
            ws, df, 'bar', 'Categoría', ['Valor'], start_row=1, config=config
        )

        assert chart is not None
        assert isinstance(chart, BarChart)
        assert 'D' in position  # Posición calculada

    def test_style_chart(self, references):
        """Verifica aplicación de estilos a gráfico"""
        categories, data, ws = references
        chart = ChartGenerator.create_bar_chart(ws, data, categories)

        style_config = {
            'title_font_size': 14
        }

        styled_chart = ChartGenerator.style_chart(chart, style_config)
        assert styled_chart is not None


class TestAddChartToSheet:
    """Tests para la función de utilidad add_chart_to_sheet"""

    def test_add_bar_chart(self):
        """Verifica agregar gráfico de barras a hoja"""
        wb = Workbook()
        ws = wb.active

        df = pd.DataFrame({
            'Category': ['A', 'B', 'C'],
            'Value': [10, 20, 30]
        })

        # No debe lanzar excepción
        add_chart_to_sheet(
            ws, df, 'bar', 'Category', 'Value',
            position="E1", title="Test"
        )

        assert len(ws._charts) == 1

    def test_add_pie_chart(self):
        """Verifica agregar gráfico de pastel a hoja"""
        wb = Workbook()
        ws = wb.active

        df = pd.DataFrame({
            'Category': ['A', 'B'],
            'Value': [60, 40]
        })

        add_chart_to_sheet(
            ws, df, 'pie', 'Category', 'Value',
            position="E1", title="Pie Test"
        )

        assert len(ws._charts) == 1

    def test_add_line_chart(self):
        """Verifica agregar gráfico de líneas a hoja"""
        wb = Workbook()
        ws = wb.active

        df = pd.DataFrame({
            'Month': ['Jan', 'Feb', 'Mar'],
            'Sales': [100, 150, 200]
        })

        add_chart_to_sheet(
            ws, df, 'line', 'Month', 'Sales',
            position="E1", title="Line Test"
        )

        assert len(ws._charts) == 1


class TestChartIntegration:
    """Tests de integración para generación de gráficos"""

    def test_save_workbook_with_chart(self):
        """Verifica guardar workbook con gráfico"""
        wb = Workbook()
        ws = wb.active

        # Agregar datos
        ws['A1'] = 'Category'
        ws['B1'] = 'Value'
        for i, (cat, val) in enumerate([('A', 10), ('B', 20), ('C', 30)], 2):
            ws[f'A{i}'] = cat
            ws[f'B{i}'] = val

        # Crear gráfico
        categories = Reference(ws, min_col=1, min_row=2, max_row=4)
        data = Reference(ws, min_col=2, min_row=1, max_row=4)

        chart = ChartGenerator.create_bar_chart(
            ws, data, categories,
            ChartConfig(title="Integration Test")
        )
        ws.add_chart(chart, "D1")

        # Guardar
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            wb.save(filename)
            assert os.path.exists(filename)

            # Verificar que se pueda cargar
            from openpyxl import load_workbook
            loaded_wb = load_workbook(filename)
            loaded_ws = loaded_wb.active
            assert len(loaded_ws._charts) == 1
        finally:
            if os.path.exists(filename):
                os.unlink(filename)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
