"""
═══════════════════════════════════════════════════════════════
TESTS PARA MÓDULO DE TEMPLATES EXCEL
Sistema de Automatización de Consultas - Chedraui CEDIS
═══════════════════════════════════════════════════════════════

Tests unitarios para el sistema de templates Excel.

Desarrollado por: Julián Alexander Juárez Alvarado (ADMJAJA)
Jefe de Sistemas - CEDIS Cancún 427
═══════════════════════════════════════════════════════════════
"""

import pytest
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
from pathlib import Path
import tempfile
import os

# Importar módulos a testear
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from modules.excel_styles import ChedrauiStyles, ChedrauiColors, apply_style_dict
from modules.excel_templates.base_template import ExcelTemplate
from modules.excel_templates.report_templates import (
    OCValidationTemplate,
    DailyPlanningTemplate,
    DistributionTemplate,
    ErrorReportTemplate,
    ASNStatusTemplate,
)


class TestChedrauiStyles:
    """Tests para la clase ChedrauiStyles"""

    def test_colors_defined(self):
        """Verifica que los colores corporativos estén definidos"""
        assert ChedrauiColors.HEADER_RED == "E31837"
        assert ChedrauiColors.OK_GREEN == "92D050"
        assert ChedrauiColors.ERROR_RED == "FF0000"
        assert ChedrauiColors.WARNING_YELLOW == "FFC000"
        assert ChedrauiColors.INFO_BLUE == "B4C7E7"

    def test_get_header_style(self):
        """Verifica que el estilo de encabezado se genere correctamente"""
        style = ChedrauiStyles.get_header_style()
        assert style is not None
        assert style.font.bold is True
        assert style.font.color.rgb == "00FFFFFF"

    def test_get_subheader_style(self):
        """Verifica que el estilo de subencabezado se genere correctamente"""
        style = ChedrauiStyles.get_subheader_style()
        assert style is not None
        assert style.font.bold is True

    def test_get_data_style(self):
        """Verifica que el estilo de datos se genere correctamente"""
        style = ChedrauiStyles.get_data_style()
        assert style is not None
        assert style.font.name == 'Arial'
        assert style.font.size == 10

    def test_get_status_style_ok(self):
        """Verifica el estilo para status OK"""
        style = ChedrauiStyles.get_status_style('OK')
        assert style is not None
        assert 'fill' in style
        assert style['fill'].start_color.rgb == f"00{ChedrauiColors.OK_GREEN}"

    def test_get_status_style_critical(self):
        """Verifica el estilo para status crítico"""
        style = ChedrauiStyles.get_status_style('CRÍTICO')
        assert style is not None
        assert 'fill' in style

    def test_get_status_style_warning(self):
        """Verifica el estilo para status warning"""
        style = ChedrauiStyles.get_status_style('alerta')
        assert style is not None

    def test_number_formats(self):
        """Verifica que los formatos numéricos estén definidos"""
        assert ChedrauiStyles.NUMBER_FORMAT_CURRENCY == '"$"#,##0.00'
        assert ChedrauiStyles.NUMBER_FORMAT_PERCENT == '0.00%'
        assert ChedrauiStyles.NUMBER_FORMAT_DATE == 'DD/MM/YYYY'


class TestExcelTemplate:
    """Tests para la clase base ExcelTemplate"""

    @pytest.fixture
    def template(self):
        """Fixture para crear un template de prueba"""
        return ExcelTemplate(cedis="TEST", author="Test Author")

    @pytest.fixture
    def sample_df(self):
        """Fixture para crear un DataFrame de prueba"""
        return pd.DataFrame({
            'Column1': ['A', 'B', 'C'],
            'Column2': [1, 2, 3],
            'Column3': [10.5, 20.5, 30.5]
        })

    def test_create_workbook(self, template):
        """Verifica que se pueda crear un workbook"""
        wb = template.create_workbook()
        assert wb is not None
        assert isinstance(wb, Workbook)

    def test_apply_corporate_header(self, template):
        """Verifica que el encabezado corporativo se aplique"""
        wb = template.create_workbook()
        ws = wb.active

        start_row = template.apply_corporate_header(
            ws, title="Test Report", subtitle="Test Subtitle"
        )

        assert ws['A1'].value == "CHEDRAUI"
        assert ws['A2'].value == "Test Report"
        assert start_row == 4

    def test_create_table(self, template, sample_df):
        """Verifica que se pueda crear una tabla"""
        wb = template.create_workbook()
        ws = wb.active

        end_row, end_col = template.create_table(ws, sample_df, start_row=1)

        assert end_row == 4  # 1 header + 3 data rows
        assert end_col == 3  # 3 columns
        assert ws.cell(row=1, column=1).value == 'Column1'

    def test_create_table_empty_df(self, template):
        """Verifica el manejo de DataFrame vacío"""
        wb = template.create_workbook()
        ws = wb.active
        empty_df = pd.DataFrame()

        end_row, end_col = template.create_table(ws, empty_df, start_row=1)

        assert ws['A1'].value == "Sin datos disponibles"

    def test_auto_adjust_columns(self, template, sample_df):
        """Verifica que el ajuste de columnas funcione"""
        wb = template.create_workbook()
        ws = wb.active
        template.create_table(ws, sample_df, start_row=1)

        # No debe lanzar excepción
        template.auto_adjust_columns(ws)

    def test_freeze_panes(self, template):
        """Verifica que los paneles se congelen"""
        wb = template.create_workbook()
        ws = wb.active

        template.freeze_panes(ws, "A5")
        assert ws.freeze_panes == "A5"

    def test_save_workbook(self, template, sample_df):
        """Verifica que el workbook se guarde correctamente"""
        wb = template.create_workbook()
        ws = wb.active
        template.create_table(ws, sample_df, start_row=1)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            result = template.save(filename)
            assert result == filename
            assert os.path.exists(filename)
        finally:
            if os.path.exists(filename):
                os.unlink(filename)


class TestOCValidationTemplate:
    """Tests para OCValidationTemplate"""

    @pytest.fixture
    def template(self):
        return OCValidationTemplate(cedis="TEST")

    @pytest.fixture
    def validation_df(self):
        return pd.DataFrame({
            'OC': ['OC001', 'OC002', 'OC003'],
            'Total_OC': [100, 200, 150],
            'Total_Distro': [100, 180, 200],
            'DIFERENCIA': [0, 20, -50],
            'STATUS': ['OK', 'Distro incompleta', 'Distro excedente']
        })

    def test_create_report(self, template, validation_df):
        """Verifica la creación de reporte de validación OC"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            result = template.create_report(validation_df, filename)
            assert result == filename
            assert os.path.exists(filename)
        finally:
            if os.path.exists(filename):
                os.unlink(filename)

    def test_create_report_auto_filename(self, template, validation_df):
        """Verifica generación de nombre automático"""
        result = template.create_report(validation_df)

        try:
            assert result.startswith("Validacion_OC_")
            assert result.endswith(".xlsx")
            assert os.path.exists(result)
        finally:
            if os.path.exists(result):
                os.unlink(result)


class TestDailyPlanningTemplate:
    """Tests para DailyPlanningTemplate"""

    @pytest.fixture
    def template(self):
        return DailyPlanningTemplate(cedis="TEST")

    @pytest.fixture
    def planning_data(self):
        df_oc = pd.DataFrame({
            'OC': ['OC001', 'OC002'],
            'PROVEEDOR': ['Prov1', 'Prov2'],
            'CANTIDAD': [100, 200]
        })
        df_asn = pd.DataFrame({
            'ASN': ['ASN001'],
            'STATUS': ['Pendiente']
        })
        df_errors = pd.DataFrame({
            'Tipo': ['Error1'],
            'Severidad': ['CRÍTICO'],
            'Mensaje': ['Test error']
        })
        return df_oc, df_asn, df_errors

    def test_create_report(self, template, planning_data):
        """Verifica la creación de reporte diario de Planning"""
        df_oc, df_asn, df_errors = planning_data

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            result = template.create_report(df_oc, df_asn, df_errors, filename)
            assert result == filename
            assert os.path.exists(filename)
        finally:
            if os.path.exists(filename):
                os.unlink(filename)


class TestDistributionTemplate:
    """Tests para DistributionTemplate"""

    @pytest.fixture
    def template(self):
        return DistributionTemplate(cedis="TEST")

    @pytest.fixture
    def distribution_df(self):
        return pd.DataFrame({
            'SKU': ['SKU001', 'SKU002', 'SKU001'],
            'TIENDA': ['T001', 'T001', 'T002'],
            'DISTR_QTY': [50, 30, 20]
        })

    def test_create_report(self, template, distribution_df):
        """Verifica la creación de reporte de distribuciones"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            result = template.create_report(distribution_df, "OC123", filename)
            assert result == filename
            assert os.path.exists(filename)
        finally:
            if os.path.exists(filename):
                os.unlink(filename)


class TestErrorReportTemplate:
    """Tests para ErrorReportTemplate"""

    @pytest.fixture
    def template(self):
        return ErrorReportTemplate(cedis="TEST")

    @pytest.fixture
    def errors_df(self):
        return pd.DataFrame({
            'Tipo': ['Error1', 'Error2'],
            'Severidad': ['CRÍTICO', 'ALTO'],
            'Mensaje': ['Error crítico', 'Error alto'],
            'Módulo': ['MOD1', 'MOD2']
        })

    def test_create_report(self, template, errors_df):
        """Verifica la creación de reporte de errores"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            result = template.create_report(errors_df, filename)
            assert result == filename
            assert os.path.exists(filename)
        finally:
            if os.path.exists(filename):
                os.unlink(filename)


class TestASNStatusTemplate:
    """Tests para ASNStatusTemplate"""

    @pytest.fixture
    def template(self):
        return ASNStatusTemplate(cedis="TEST")

    @pytest.fixture
    def asn_df(self):
        return pd.DataFrame({
            'ASN': ['ASN001', 'ASN002'],
            'STATUS': ['Recibido', 'Pendiente'],
            'PROVEEDOR': ['Prov1', 'Prov2']
        })

    def test_create_report(self, template, asn_df):
        """Verifica la creación de reporte de ASN"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            result = template.create_report(asn_df, filename)
            assert result == filename
            assert os.path.exists(filename)
        finally:
            if os.path.exists(filename):
                os.unlink(filename)


# ═══════════════════════════════════════════════════════════════
# TESTS DE INTEGRACIÓN
# ═══════════════════════════════════════════════════════════════

class TestTemplateIntegration:
    """Tests de integración para templates"""

    def test_full_workflow(self):
        """Test de flujo completo de generación de reporte"""
        # Crear datos de prueba
        df = pd.DataFrame({
            'OC': ['OC001', 'OC002', 'OC003'],
            'CANTIDAD': [100, 200, 150],
            'STATUS': ['OK', 'Incompleta', 'OK']
        })

        # Crear template y generar reporte
        template = ExcelTemplate(cedis="Integration Test")
        wb = template.create_workbook()
        ws = wb.active

        # Aplicar encabezado
        start_row = template.apply_corporate_header(
            ws, "Integration Test Report", "Subtitle"
        )

        # Crear tabla
        end_row, end_col = template.create_table(ws, df, start_row)

        # Aplicar formato adicional
        template.auto_adjust_columns(ws)
        template.freeze_panes(ws, f"A{start_row + 1}")
        template.apply_corporate_footer(ws)

        # Guardar
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filename = f.name

        try:
            result = template.save(filename)
            assert os.path.exists(result)

            # Verificar contenido
            from openpyxl import load_workbook
            loaded_wb = load_workbook(filename)
            loaded_ws = loaded_wb.active

            assert loaded_ws['A1'].value == "CHEDRAUI"
            assert loaded_ws['A2'].value == "Integration Test Report"
        finally:
            if os.path.exists(filename):
                os.unlink(filename)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
